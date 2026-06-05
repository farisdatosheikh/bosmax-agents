from __future__ import annotations

import argparse
import json
import sys
from functools import lru_cache
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from product_copy_router import (
    WORKBOOK_PATH,
    load_products,
    load_registry as load_router_registry,
    load_workbook_index,
    normalize,
    route_product,
)
from resolver_runtime import (
    ResolverError,
    resolve_avatar_context_id,
    resolve_avatar_pool,
    resolve_copywriting_id,
)


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
CONTRACT_PATH = ROOT / "registries" / "product_on_the_fly_video_prompt_contract.yaml"

COPY_PACK_COLUMNS = [
    "Row_ID",
    "Family_Code",
    "Family_Name",
    "Product_ID_Optional",
    "Product_Name_Optional",
    "Category",
    "Sub_Category",
    "Product_Type",
    "Product_Size",
    "Product_Scale",
    "Copywriting_Formula",
    "Hook",
    "Pain_or_Friction",
    "USP_1",
    "USP_2",
    "USP_3",
    "CTA",
    "Status",
]


@lru_cache(maxsize=1)
def load_contract() -> dict[str, Any]:
    with CONTRACT_PATH.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


@lru_cache(maxsize=1)
def load_copy_pack_index() -> dict[str, Any]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    approved_by_product_id: dict[str, list[dict[str, str]]] = {}

    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("PRODUCT_"):
            continue

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue

        header_map = {header: idx for idx, header in enumerate(headers) if header}
        if "Status" not in header_map:
            continue

        for row in rows:
            status = str(row[header_map["Status"]] or "").strip()
            if status != "APPROVED":
                continue

            record = {"sheet_name": sheet_name}
            for column in COPY_PACK_COLUMNS:
                if column in header_map:
                    record[column] = str(row[header_map[column]] or "").strip()
            product_id = normalize(record.get("Product_ID_Optional"))
            if product_id:
                approved_by_product_id.setdefault(product_id, []).append(record)

    for records in approved_by_product_id.values():
        records.sort(key=lambda item: item.get("Row_ID", ""))

    return {"approved_by_product_id": approved_by_product_id}


def find_product_record(payload: dict[str, Any], route_result: dict[str, Any], products: dict[str, Any]) -> dict[str, Any] | None:
    product_id_key = normalize(route_result.get("product_id") or payload.get("product_id"))
    product_name_key = normalize(route_result.get("product_name") or payload.get("product_name"))
    return products["by_id"].get(product_id_key) or products["by_name"].get(product_name_key)


def find_registered_copy_pack(payload: dict[str, Any], route_result: dict[str, Any]) -> dict[str, str] | None:
    copy_index = load_copy_pack_index()
    product_id_key = normalize(route_result.get("product_id") or payload.get("product_id"))
    if not product_id_key:
        return None
    records = copy_index["approved_by_product_id"].get(product_id_key, [])
    return records[0] if records else None


def select_prompt_module_status(route_mode: str, route_result: dict[str, Any], template: dict[str, Any]) -> str:
    if route_mode == "REVIEW_ONLY_PRODUCT":
        return template["prompt_module_status_blocked"]
    if route_result.get("ready_for_generation"):
        return template["prompt_module_status_ready"]
    return template.get("prompt_module_status_hold", "HOLD_MISSING_INPUT")


def format_scalar(value: Any, default: str = "none") -> str:
    text = str(value or "").strip()
    return text if text else default


def format_list(values: list[Any]) -> str:
    items = [str(value).strip() for value in values if str(value or "").strip()]
    return ", ".join(items) if items else "none"


def append_shared_context(lines: list[str], payload: dict[str, Any], route_result: dict[str, Any]) -> None:
    lines.extend(
        [
            f"product_name: {format_scalar(route_result.get('product_name') or payload.get('product_name'))}",
            f"category: {format_scalar(payload.get('category'))}",
            f"target_user: {format_scalar(payload.get('target_user'))}",
            f"main_problem_solved: {format_scalar(payload.get('main_problem_solved'))}",
            f"main_benefit: {format_scalar(payload.get('main_benefit'))}",
            f"product_format: {format_scalar(payload.get('product_format'))}",
            f"product_size_or_scale: {format_scalar(payload.get('product_size_or_scale'))}",
            f"platform: {format_scalar(payload.get('platform'))}",
            f"language: {format_scalar(payload.get('language'))}",
            f"compliance_class: {format_scalar(payload.get('compliance_class'))}",
            f"visual_reference_status: {format_scalar(payload.get('visual_reference_status'))}",
        ]
    )


def resolve_runtime_contract(
    payload: dict[str, Any],
    route_result: dict[str, Any],
    product_record: dict[str, Any] | None,
    copy_pack: dict[str, str] | None,
) -> dict[str, Any]:
    manual_override = normalize(payload.get("manual_override")).lower() in {"1", "true", "yes"}
    review_status = normalize(payload.get("review_status"))

    expected_template_lane = (
        normalize(payload.get("template_lane"))
        or normalize(payload.get("silo"))
        or normalize((product_record or {}).get("silo"))
        or normalize((copy_pack or {}).get("Type_of_Content"))
    )
    expected_template_silo = (
        normalize(payload.get("silo_key"))
        or normalize((copy_pack or {}).get("Silo_Key"))
        or normalize(((product_record or {}).get("dialogue_authority") or {}).get("silo_id"))
    )
    expected_product_family = normalize(payload.get("product_family") or route_result.get("family_code"))
    expected_camera_style = normalize(payload.get("camera_style"))
    physics_class = normalize(payload.get("physics_class"))

    resolved: dict[str, Any] = {}

    if normalize(payload.get("copywriting_mode")) == "auto_resolve":
        copywriting_id = normalize(payload.get("copywriting_id"))
        if not copywriting_id:
            raise ResolverError("Copywriting Mode AUTO_RESOLVE requires copywriting_id.")
        resolved["copywriting"] = resolve_copywriting_id(
            copywriting_id,
            expected_template_lane=expected_template_lane,
            expected_template_silo=expected_template_silo,
            manual_override=manual_override,
            review_status=review_status,
        )

    avatar_mode = normalize(payload.get("avatar_mode"))
    if avatar_mode == "auto_resolve":
        avatar_context_id = normalize(payload.get("avatar_context_id"))
        if not avatar_context_id:
            raise ResolverError("Avatar Mode AUTO_RESOLVE requires avatar_context_id.")
        resolved["avatar_context"] = resolve_avatar_context_id(
            avatar_context_id,
            expected_template_silo=expected_template_lane or expected_template_silo,
            expected_product_family=expected_product_family,
            expected_camera_style=expected_camera_style,
            physics_class=physics_class,
            manual_override=manual_override,
            review_status=review_status,
        )
    elif avatar_mode == "auto_rotate":
        avatar_pool_id = normalize(payload.get("avatar_pool_id"))
        rotation_rule = normalize(payload.get("rotation_rule"))
        batch_count = int(payload.get("batch_count") or 0)
        if not avatar_pool_id:
            raise ResolverError("Avatar Mode AUTO_ROTATE requires avatar_pool_id.")
        if not rotation_rule:
            raise ResolverError("Avatar Mode AUTO_ROTATE requires rotation_rule.")
        resolved["avatar_pool"] = resolve_avatar_pool(
            avatar_pool_id,
            batch_count=batch_count,
            rotation_rule=rotation_rule,
            expected_template_silo=expected_template_lane or expected_template_silo,
            expected_product_family=expected_product_family,
            expected_camera_style=expected_camera_style,
            physics_class=physics_class,
            manual_override=manual_override,
            review_status=review_status,
        )

    return resolved


def append_resolver_output(lines: list[str], payload: dict[str, Any], resolved_contract: dict[str, Any]) -> None:
    copywriting = resolved_contract.get("copywriting")
    if copywriting:
        lines.extend(
            [
                f"copywriting_id: {format_scalar(copywriting.get('copywriting_id'))}",
                f"copywriting_mode: {format_scalar(payload.get('copywriting_mode'))}",
                f"resolved_copy_formula: {format_scalar(copywriting.get('submode_formula'))}",
                f"resolved_copy_hook: {format_scalar(copywriting.get('hook'))}",
                f"resolved_copy_pain_or_friction: {format_scalar(copywriting.get('pain_or_friction'))}",
                f"resolved_copy_usp_1: {format_scalar(copywriting.get('usp_1'))}",
                f"resolved_copy_usp_2: {format_scalar(copywriting.get('usp_2'))}",
                f"resolved_copy_usp_3: {format_scalar(copywriting.get('usp_3'))}",
                f"resolved_copy_cta: {format_scalar(copywriting.get('cta'))}",
            ]
        )

    avatar_context = resolved_contract.get("avatar_context")
    if avatar_context:
        lines.extend(
            [
                f"avatar_context_id: {format_scalar(avatar_context.get('avatar_context_id'))}",
                f"avatar_mode: {format_scalar(payload.get('avatar_mode'))}",
                f"resolved_persona_id: {format_scalar(avatar_context.get('persona_id'))}",
                f"resolved_wardrobe_id: {format_scalar(avatar_context.get('wardrobe_id'))}",
                f"resolved_scene_context_id: {format_scalar(avatar_context.get('scene_context_id'))}",
                f"resolved_mannequin_id: {format_scalar(avatar_context.get('mannequin_id'))}",
                f"resolved_camera_style: {format_scalar(format_list(avatar_context.get('camera_style_allowed') or []))}",
            ]
        )

    avatar_pool = resolved_contract.get("avatar_pool")
    if avatar_pool:
        pool = avatar_pool["pool"]
        first_selection = avatar_pool["sequence"][0] if avatar_pool["sequence"] else {}
        lines.extend(
            [
                f"avatar_pool_id: {format_scalar(pool.get('pool_id'))}",
                f"avatar_mode: {format_scalar(payload.get('avatar_mode'))}",
                f"batch_count: {format_scalar(payload.get('batch_count'))}",
                f"rotation_rule: {format_scalar(payload.get('rotation_rule'))}",
                f"rotation_sequence: {format_scalar(', '.join(avatar_pool['sequence_ids']))}",
                f"resolved_persona_id: {format_scalar(first_selection.get('persona_id'))}",
                f"resolved_wardrobe_id: {format_scalar(first_selection.get('wardrobe_id'))}",
                f"resolved_scene_context_id: {format_scalar(first_selection.get('scene_context_id'))}",
                f"resolved_mannequin_id: {format_scalar(first_selection.get('mannequin_id'))}",
            ]
        )


def build_registered_output(
    payload: dict[str, Any],
    route_result: dict[str, Any],
    template: dict[str, Any],
    product_record: dict[str, Any] | None,
    copy_pack: dict[str, str] | None,
    resolved_contract: dict[str, Any],
) -> str:
    resolved_copywriting = resolved_contract.get("copywriting")
    repo_authority_parts = [product_record["source_file"]] if product_record else []
    if copy_pack:
        repo_authority_parts.append(f"{copy_pack['sheet_name']}::{copy_pack['Row_ID']}")
    if resolved_copywriting:
        repo_authority_parts.append(f"registries/copywriting_id_resolver.yaml::{resolved_copywriting['copywriting_id']}")
    if resolved_contract.get("avatar_context"):
        repo_authority_parts.append(
            f"registries/avatar_context_rotation.yaml::{resolved_contract['avatar_context']['avatar_context_id']}"
        )
    if resolved_contract.get("avatar_pool"):
        repo_authority_parts.append(
            f"registries/avatar_context_rotation.yaml::{resolved_contract['avatar_pool']['pool']['pool_id']}"
        )

    lines = [
        "[BOSMAX_ROUTE_MANUAL_OUTPUT]",
        "module: PRODUCT_ON_THE_FLY_VIDEO_PROMPT",
        "route_mode: REGISTERED_PRODUCT",
        f"router_output_status: {template['router_output_status']}",
        f"prompt_module_status: {template['prompt_module_status_ready']}",
        f"session_scope: {template['session_scope']}",
        f"authority_tier: {template['authority_tier']}",
        f"repo_authority: {format_list(repo_authority_parts)}",
        f"product_id: {format_scalar(route_result.get('product_id') or payload.get('product_id'))}",
        f"product_name: {format_scalar(route_result.get('product_name') or payload.get('product_name'))}",
        f"category: {format_scalar((product_record or {}).get('category') or payload.get('category'))}",
        f"sub_category: {format_scalar((product_record or {}).get('sub_category') or payload.get('sub_category'))}",
        f"copy_pack_row_id: {format_scalar((copy_pack or {}).get('Row_ID'))}",
        f"copy_formula: {format_scalar((resolved_copywriting or {}).get('submode_formula') or (copy_pack or {}).get('Copywriting_Formula'))}",
        f"copy_hook: {format_scalar((resolved_copywriting or {}).get('hook') or (copy_pack or {}).get('Hook'))}",
        f"copy_pain_or_friction: {format_scalar((resolved_copywriting or {}).get('pain_or_friction') or (copy_pack or {}).get('Pain_or_Friction'))}",
        f"copy_usp_1: {format_scalar((resolved_copywriting or {}).get('usp_1') or (copy_pack or {}).get('USP_1'))}",
        f"copy_usp_2: {format_scalar((resolved_copywriting or {}).get('usp_2') or (copy_pack or {}).get('USP_2'))}",
        f"copy_usp_3: {format_scalar((resolved_copywriting or {}).get('usp_3') or (copy_pack or {}).get('USP_3'))}",
        f"copy_cta: {format_scalar((resolved_copywriting or {}).get('cta') or (copy_pack or {}).get('CTA'))}",
        f"product_scale_anchor: {format_scalar((copy_pack or {}).get('Product_Scale'))}",
        "operator_request:",
    ]
    append_resolver_output(lines, payload, resolved_contract)
    for rule in template["operator_rules"]:
        lines.append(f"- {rule}")
    lines.extend(
        [
            "- Generate the video prompt from the registered product truth and approved copy pack only.",
            "[/BOSMAX_ROUTE_MANUAL_OUTPUT]",
        ]
    )
    return "\n".join(lines)


def build_family_output(payload: dict[str, Any], route_result: dict[str, Any], template: dict[str, Any]) -> str:
    prompt_module_status = select_prompt_module_status("FAMILY_MATCHED_PRODUCT", route_result, template)
    lines = [
        "[BOSMAX_ROUTE_MANUAL_OUTPUT]",
        "module: PRODUCT_ON_THE_FLY_VIDEO_PROMPT",
        "route_mode: FAMILY_MATCHED_PRODUCT",
        f"router_output_status: {template['router_output_status']}",
        f"prompt_module_status: {prompt_module_status}",
        f"session_scope: {template['session_scope']}",
        f"authority_tier: {template['authority_tier']}",
        f"family_sheet: {format_scalar(route_result.get('family_sheet'))}",
        f"family_code: {format_scalar(route_result.get('family_code'))}",
        f"minimum_intake_missing: {format_list(route_result.get('minimum_intake_missing', []))}",
    ]
    append_shared_context(lines, payload, route_result)
    lines.append("operator_request:")
    for rule in template["operator_rules"]:
        lines.append(f"- {rule}")
    lines.append("- Generate one family-lane video prompt from the current intake only.")
    lines.append("[/BOSMAX_ROUTE_MANUAL_OUTPUT]")
    return "\n".join(lines)


def build_on_the_fly_output(payload: dict[str, Any], route_result: dict[str, Any], template: dict[str, Any]) -> str:
    prompt_module_status = select_prompt_module_status("ON_THE_FLY_PRODUCT", route_result, template)
    lines = [
        "[BOSMAX_ROUTE_MANUAL_OUTPUT]",
        "module: PRODUCT_ON_THE_FLY_VIDEO_PROMPT",
        "route_mode: ON_THE_FLY_PRODUCT",
        f"router_output_status: {template['router_output_status']}",
        f"prompt_module_status: {prompt_module_status}",
        f"session_scope: {template['session_scope']}",
        f"authority_tier: {template['authority_tier']}",
        f"minimum_intake_missing: {format_list(route_result.get('minimum_intake_missing', []))}",
    ]
    append_shared_context(lines, payload, route_result)
    lines.append("operator_request:")
    for rule in template["operator_rules"]:
        lines.append(f"- {rule}")
    lines.append("- Generate one AI-ready video request block from this intake and keep it session-only.")
    lines.append("[/BOSMAX_ROUTE_MANUAL_OUTPUT]")
    return "\n".join(lines)


def build_review_only_output(payload: dict[str, Any], route_result: dict[str, Any], template: dict[str, Any]) -> str:
    lines = [
        "[BOSMAX_ROUTE_MANUAL_OUTPUT]",
        "module: PRODUCT_ON_THE_FLY_VIDEO_PROMPT",
        "route_mode: REVIEW_ONLY_PRODUCT",
        f"router_output_status: {template['router_output_status']}",
        f"prompt_module_status: {template['prompt_module_status_blocked']}",
        f"session_scope: {template['session_scope']}",
        f"authority_tier: {template['authority_tier']}",
        f"product_name: {format_scalar(route_result.get('product_name') or payload.get('product_name'))}",
        f"category: {format_scalar(payload.get('category'))}",
        f"compliance_class: {format_scalar(payload.get('compliance_class'))}",
        f"minimum_intake_missing: {format_list(route_result.get('minimum_intake_missing', []))}",
        f"block_reason: {format_list(route_result.get('notes', []))}",
        "review_action:",
    ]
    for rule in template["operator_rules"]:
        lines.append(f"- {rule}")
    lines.append("- STOP. Do not generate a final AI-ready video prompt for this route.")
    lines.append("[/BOSMAX_ROUTE_MANUAL_OUTPUT]")
    return "\n".join(lines)


def build_manual_request(payload: dict[str, Any]) -> dict[str, Any]:
    contract = load_contract()
    router_registry = load_router_registry()
    products = load_products()
    workbook_index = load_workbook_index()
    route_result = route_product(payload, registry=router_registry, products=products, workbook_index=workbook_index)
    route_mode = route_result["route_mode"]
    product_record = find_product_record(payload, route_result, products)
    copy_pack = find_registered_copy_pack(payload, route_result) if route_mode == "REGISTERED_PRODUCT" else None
    resolved_contract: dict[str, Any] = {}

    if route_mode != "REVIEW_ONLY_PRODUCT":
        try:
            resolved_contract = resolve_runtime_contract(payload, route_result, product_record, copy_pack)
        except ResolverError as exc:
            route_result = {
                **route_result,
                "route_mode": "REVIEW_ONLY_PRODUCT",
                "output_status": contract["route_templates"]["REVIEW_ONLY_PRODUCT"]["router_output_status"],
                "ready_for_generation": False,
                "requires_review": True,
                "notes": [*route_result.get("notes", []), str(exc)],
            }
            route_mode = "REVIEW_ONLY_PRODUCT"

    template = contract["route_templates"][route_mode]

    if route_mode == "REGISTERED_PRODUCT":
        manual_output = build_registered_output(payload, route_result, template, product_record, copy_pack, resolved_contract)
    elif route_mode == "FAMILY_MATCHED_PRODUCT":
        manual_output = build_family_output(payload, route_result, template)
    elif route_mode == "ON_THE_FLY_PRODUCT":
        manual_output = build_on_the_fly_output(payload, route_result, template)
    else:
        manual_output = build_review_only_output(payload, route_result, template)

    prompt_module_status = select_prompt_module_status(route_mode, route_result, template)
    return {
        "route_result": route_result,
        "prompt_module_status": prompt_module_status,
        "manual_output_field_candidates": contract["authority"]["downstream_output_field_candidates"],
        "manual_output": manual_output,
        "copy_pack_row_id": (copy_pack or {}).get("Row_ID", ""),
        "resolved_contract": resolved_contract,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build route-specific AI-ready manual request blocks for BOSMAX video runs.")
    parser.add_argument("--sample", default="")
    parser.add_argument("--output-format", choices=("text", "json"), default="text")
    parser.add_argument("--product-id", default="")
    parser.add_argument("--product-name", default="")
    parser.add_argument("--category", default="")
    parser.add_argument("--sub-category", default="")
    parser.add_argument("--target-user", default="")
    parser.add_argument("--main-problem-solved", default="")
    parser.add_argument("--main-benefit", default="")
    parser.add_argument("--product-format", default="")
    parser.add_argument("--product-size-or-scale", default="")
    parser.add_argument("--platform", default="")
    parser.add_argument("--language", default="")
    parser.add_argument("--compliance-class", default="")
    parser.add_argument("--visual-reference-status", default="")
    parser.add_argument("--template-lane", default="")
    parser.add_argument("--silo", default="")
    parser.add_argument("--silo-key", default="")
    parser.add_argument("--product-family", default="")
    parser.add_argument("--camera-style", default="")
    parser.add_argument("--physics-class", default="")
    parser.add_argument("--copywriting-id", default="")
    parser.add_argument("--copywriting-mode", default="")
    parser.add_argument("--avatar-context-id", default="")
    parser.add_argument("--avatar-pool-id", default="")
    parser.add_argument("--avatar-mode", default="")
    parser.add_argument("--batch-count", default="")
    parser.add_argument("--rotation-rule", default="")
    parser.add_argument("--manual-override", default="")
    parser.add_argument("--review-status", default="")
    return parser.parse_args()


def build_payload_from_args(args: argparse.Namespace) -> dict[str, Any]:
    contract = load_contract()
    if args.sample:
        sample = contract.get("sample_payloads", {}).get(args.sample)
        if sample is None:
            raise SystemExit(f"Unknown sample payload: {args.sample}")
        return dict(sample)

    return {
        "product_id": args.product_id,
        "product_name": args.product_name,
        "category": args.category,
        "sub_category": args.sub_category,
        "target_user": args.target_user,
        "main_problem_solved": args.main_problem_solved,
        "main_benefit": args.main_benefit,
        "product_format": args.product_format,
        "product_size_or_scale": args.product_size_or_scale,
        "platform": args.platform,
        "language": args.language,
        "compliance_class": args.compliance_class,
        "visual_reference_status": args.visual_reference_status,
        "template_lane": args.template_lane,
        "silo": args.silo,
        "silo_key": args.silo_key,
        "product_family": args.product_family,
        "camera_style": args.camera_style,
        "physics_class": args.physics_class,
        "copywriting_id": args.copywriting_id,
        "copywriting_mode": args.copywriting_mode,
        "avatar_context_id": args.avatar_context_id,
        "avatar_pool_id": args.avatar_pool_id,
        "avatar_mode": args.avatar_mode,
        "batch_count": args.batch_count,
        "rotation_rule": args.rotation_rule,
        "manual_override": args.manual_override,
        "review_status": args.review_status,
    }


def main() -> None:
    args = parse_args()
    payload = build_payload_from_args(args)
    result = build_manual_request(payload)

    if args.output_format == "json":
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return

    print(result["manual_output"])


if __name__ == "__main__":
    main()
