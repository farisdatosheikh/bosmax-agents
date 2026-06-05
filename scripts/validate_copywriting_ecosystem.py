from __future__ import annotations

import sys
from collections import Counter
from pathlib import Path

import yaml
from openpyxl import load_workbook

from stealth_copy_authority import (
    AUTHORITY_MAP_PATH,
    FORBIDDEN_DIRECT_PRONOUNS,
    FORBIDDEN_MEDICAL_CLAIMS,
    GENERIC_BANNED_PHRASES,
    STEALTH_SOURCE_HEADERS,
    STEALTH_WORKBOOK_SHEETS,
    contains_term,
    validate_stealth_blob,
)

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx"
REGISTRY_PATH = ROOT / "registries" / "dialogue_budget_corridor.yaml"
MWCB_TAXONOMY_PATH = ROOT / "registries" / "mwcb_copywriting_angle_taxonomy.yaml"
PAIN_HEADER = "Pain_or_Friction"
REQUIRED_PRIORITY_SHEETS = [
    "PRODUCT_BOSMAX_SERUM",
    "PRODUCT_MW_CAP_BURUNG",
    "FAMILY_MALE_EXT_OIL",
    "FAMILY_TRAD_REMEDY_OIL",
]
VALID_FORMULAS = {"AIDA", "PAS", "HSO", "HPAS", "SAVAGE_HPAS"}
VALID_LANES = {"DIRECT", "STEALTH"}
PLACEHOLDER_TOKENS = ("tbd", "placeholder", "hook 1", "usp 1", "cta 1", "lorem ipsum")

# MWCB taxonomy constants
VALID_MCA_IDS = {
    "MWCB-MCA01", "MWCB-MCA02", "MWCB-MCA03", "MWCB-MCA04",
    "MWCB-MCA05", "MWCB-MCA06", "MWCB-MCA07", "REVIEW_ONLY", "UNASSIGNED",
}
VALID_COMPLIANCE_RISKS = {"GREEN", "YELLOW", "HIGH", "REVIEW_ONLY", "UNASSIGNED"}
MWCB_SHEETS = {"PRODUCT_MW_CAP_BURUNG", "FAMILY_TRAD_REMEDY_OIL"}
APPROVED_STATUSES = {"APPROVED", "LOCKED"}
MWCB_FORBIDDEN_COPY_TERMS = (
    "roll-on",
    "roll on",
    "roller",
    "rollerball",
    "pump",
    "spray",
    "dropper",
    "ubat",
    "sembuh",
    "merawat",
    "penawar",
    "antiseptik",
    "antibakteria",
    "selamat untuk bayi",
    "selamat untuk kanak-kanak",
    "buka saluran hidung",
    "longgarkan kahak",
    "mengecutkan rahim",
    "mengempiskan perut",
)
PHASE1_MARKER = "PHASE_1_MWCB|"
PHASE1_MCA_IDS = {"MWCB-MCA01", "MWCB-MCA03", "MWCB-MCA07"}
PHASE1_FORMULAS = {"PAS", "AIDA", "HSO", "HPAS"}
PHASE1_CONTEXTS = {"CTX-A_SEGARA", "CTX-B_RUTIN", "CTX-C_SYOR"}


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def fail(message: str) -> None:
    print(f"FAIL: {message}")
    sys.exit(1)


def validate_workbook():
    if not WORKBOOK_PATH.exists():
        fail(f"Workbook missing: {WORKBOOK_PATH}")

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    missing_sheets = [name for name in REQUIRED_PRIORITY_SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        fail(f"Required sheets missing: {', '.join(missing_sheets)}")

    row_counts: dict[str, int] = {}
    for sheet_name in workbook.sheetnames:
        if not (sheet_name.startswith("PRODUCT_") or sheet_name.startswith("FAMILY_")):
            continue

        ws = workbook[sheet_name]
        headers = [normalize(cell.value) for cell in ws[1]]
        if "Hook" not in headers or "CTA" not in headers:
            continue
        header_index = {header: idx for idx, header in enumerate(headers)}
        required = {"Row_ID", "Hook", "USP_1", "USP_2", "USP_3", "CTA", "Copywriting_Formula", "Silo_Key", "Type_of_Content", PAIN_HEADER}
        if sheet_name in STEALTH_WORKBOOK_SHEETS:
            required.update(STEALTH_SOURCE_HEADERS)
        missing_headers = [header for header in required if header not in header_index]
        if missing_headers:
            fail(f"{sheet_name} missing headers: {', '.join(missing_headers)}")

        rows_checked = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            hook = normalize(row[header_index["Hook"]])
            usp_1 = normalize(row[header_index["USP_1"]])
            usp_2 = normalize(row[header_index["USP_2"]])
            usp_3 = normalize(row[header_index["USP_3"]])
            cta = normalize(row[header_index["CTA"]])
            formula = normalize(row[header_index["Copywriting_Formula"]])
            silo_key = normalize(row[header_index["Silo_Key"]])
            lane = normalize(row[header_index["Type_of_Content"]])
            pain = normalize(row[header_index[PAIN_HEADER]])
            row_id = normalize(row[header_index["Row_ID"]])
            authority_source = normalize(row[header_index["Authority_Source"]]) if "Authority_Source" in header_index else ""

            populated = any((hook, usp_1, usp_2, usp_3, cta))
            if not populated:
                continue

            rows_checked += 1
            if not pain:
                fail(f"{sheet_name} {row_id} has content but blank {PAIN_HEADER}")
            if formula not in VALID_FORMULAS:
                fail(f"{sheet_name} {row_id}: invalid formula '{formula}'")
            if lane not in VALID_LANES:
                fail(f"{sheet_name} {row_id}: invalid lane '{lane}'")
            if not silo_key:
                fail(f"{sheet_name} {row_id}: blank Silo_Key")

            row_blob = " | ".join([hook, pain, usp_1, usp_2, usp_3, cta]).lower()
            if any(token in row_blob for token in PLACEHOLDER_TOKENS):
                fail(f"{sheet_name} {row_id}: obvious placeholder text detected")

            if sheet_name == "PRODUCT_BOSMAX_SERUM" and lane != "STEALTH":
                fail(f"{sheet_name} {row_id}: BOSMAX Serum must remain STEALTH")
            if sheet_name == "PRODUCT_MW_CAP_BURUNG" and lane != "DIRECT":
                fail(f"{sheet_name} {row_id}: MW Cap Burung must remain DIRECT")
            if sheet_name == "FAMILY_MALE_EXT_OIL" and lane != "STEALTH":
                fail(f"{sheet_name} {row_id}: FAMILY_MALE_EXT_OIL must remain STEALTH")
            if sheet_name == "FAMILY_TRAD_REMEDY_OIL" and lane != "DIRECT":
                fail(f"{sheet_name} {row_id}: FAMILY_TRAD_REMEDY_OIL must remain DIRECT")

            if sheet_name in STEALTH_WORKBOOK_SHEETS:
                if "SCRIPT_REGISTRY_UNIFIED.md" not in authority_source or "SCRIPT_VARIANT_LIBRARY.md" not in authority_source:
                    fail(f"{sheet_name} {row_id}: stealth authority source missing script-registry lock")

                source_metadata = {
                    header: normalize(row[header_index[header]])
                    for header in STEALTH_SOURCE_HEADERS
                }
                blanks = [header for header, value in source_metadata.items() if not value]
                if blanks:
                    fail(f"{sheet_name} {row_id}: blank stealth source metadata {', '.join(blanks)}")

                row_blob = " | ".join([hook, pain, usp_1, usp_2, usp_3, cta])
                issues = validate_stealth_blob(row_blob)
                if issues:
                    fail(f"{sheet_name} {row_id}: {'; '.join(issues)}")

                lowered = row_blob.lower()
                if any(contains_term(lowered, pronoun) for pronoun in FORBIDDEN_DIRECT_PRONOUNS):
                    fail(f"{sheet_name} {row_id}: forbidden direct pronoun detected")
                if any(contains_term(lowered, claim) for claim in FORBIDDEN_MEDICAL_CLAIMS):
                    fail(f"{sheet_name} {row_id}: forbidden medical claim detected")
                if any(contains_term(lowered, phrase) for phrase in GENERIC_BANNED_PHRASES):
                    fail(f"{sheet_name} {row_id}: generic packaging phrase detected")

        row_counts[sheet_name] = rows_checked

    return row_counts


def validate_registry() -> None:
    if not REGISTRY_PATH.exists():
        fail(f"Dialogue budget registry missing: {REGISTRY_PATH}")

    with REGISTRY_PATH.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh)

    corridors = data.get("corridors", [])
    if not isinstance(corridors, list) or not corridors:
        fail("Dialogue budget registry has no corridors list")

    bm_brisk = {
        entry["duration_seconds"]: entry
        for entry in corridors
        if entry.get("language") == "BM" and entry.get("pace_class") == "BRISK_UGC"
    }
    required_durations = [6, 8, 10, 12, 15, 16, 20, 24, 30]
    missing = [str(duration) for duration in required_durations if duration not in bm_brisk]
    if missing:
        fail(f"BM/BRISK_UGC durations missing: {', '.join(missing)}")

    for duration in required_durations:
        entry = bm_brisk[duration]
        chain = [
            entry["minimum_words"],
            entry["target_min_words"],
            entry["target_max_words"],
            entry["safe_max_words"],
            entry["hard_ceiling_words"],
        ]
        if chain != sorted(chain):
            fail(f"Budget monotonicity failed for duration {duration}s: {chain}")


def validate_stealth_authority_map() -> None:
    if not AUTHORITY_MAP_PATH.exists():
        fail(f"Stealth authority map missing: {AUTHORITY_MAP_PATH}")

    with AUTHORITY_MAP_PATH.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh)

    bosmax = (data or {}).get("products", {}).get("BOSMAX_SERUM")
    if not bosmax:
        fail("Stealth authority map missing products.BOSMAX_SERUM")

    if bosmax.get("copywriting_mode") != "script_registry_and_variant_only":
        fail("Stealth authority map copywriting_mode must remain script_registry_and_variant_only")

    metadata_headers = bosmax.get("workbook_source_metadata", [])
    if list(metadata_headers) != list(STEALTH_SOURCE_HEADERS):
        fail("Stealth authority map workbook_source_metadata drift detected")


def validate_mwcb_taxonomy() -> None:
    """Validate MWCB MCA taxonomy registry and workbook column compliance."""
    if not MWCB_TAXONOMY_PATH.exists():
        fail(f"MWCB taxonomy registry missing: {MWCB_TAXONOMY_PATH}")

    with MWCB_TAXONOMY_PATH.open("r", encoding="utf-8") as fh:
        taxonomy = yaml.safe_load(fh)

    # 1. All 7 MCA IDs must exist
    mca_list = taxonomy.get("mca_list", [])
    mca_ids_in_registry = {entry["mca_id"] for entry in mca_list}
    required_mca_ids = {
        "MWCB-MCA01", "MWCB-MCA02", "MWCB-MCA03", "MWCB-MCA04",
        "MWCB-MCA05", "MWCB-MCA06", "MWCB-MCA07",
    }
    missing_mca = required_mca_ids - mca_ids_in_registry
    if missing_mca:
        fail(f"MWCB taxonomy missing MCA IDs: {', '.join(sorted(missing_mca))}")

    # 2. Phase 1 MCAs must be exactly MWCB-MCA01, MWCB-MCA03, MWCB-MCA07
    phase_1_in_registry = set(taxonomy.get("phase_1_ready_mcas", []))
    expected_phase_1 = {"MWCB-MCA01", "MWCB-MCA03", "MWCB-MCA07"}
    if phase_1_in_registry != expected_phase_1:
        fail(
            f"MWCB taxonomy phase_1_ready_mcas mismatch. "
            f"Expected: {sorted(expected_phase_1)}, got: {sorted(phase_1_in_registry)}"
        )

    # 3. REVIEW_ONLY use cases must be present and quarantined
    review_use_cases = taxonomy.get("review_only_use_cases", [])
    if not review_use_cases:
        fail("MWCB taxonomy missing review_only_use_cases — quarantine not defined")
    review_uc_ids = {uc["use_case_id"] for uc in review_use_cases}
    required_review_ucs = {
        "RO-UC-01", "RO-UC-02", "RO-UC-03", "RO-UC-04",
        "RO-UC-05", "RO-UC-06", "RO-UC-07",
    }
    missing_ucs = required_review_ucs - review_uc_ids
    if missing_ucs:
        fail(f"MWCB taxonomy missing review-only use cases: {', '.join(sorted(missing_ucs))}")

    # 4 + 5. Workbook sheets must have MCA_ID and Compliance_Risk columns
    if not WORKBOOK_PATH.exists():
        fail(f"Workbook missing: {WORKBOOK_PATH}")
    wb = load_workbook(WORKBOOK_PATH, data_only=True)

    for sheet_name in MWCB_SHEETS:
        if sheet_name not in wb.sheetnames:
            fail(f"Required MWCB sheet missing from workbook: {sheet_name}")
        ws = wb[sheet_name]
        headers = [normalize(cell.value) for cell in ws[1]]
        if "MCA_ID" not in headers:
            fail(f"{sheet_name}: missing required column MCA_ID")
        if "Compliance_Risk" not in headers:
            fail(f"{sheet_name}: missing required column Compliance_Risk")

        header_index = {h: i for i, h in enumerate(headers)}
        mca_col = header_index["MCA_ID"]
        risk_col = header_index["Compliance_Risk"]
        status_col = header_index.get("Status")

        for row in ws.iter_rows(min_row=2, values_only=True):
            mca_id = normalize(row[mca_col])
            compliance_risk = normalize(row[risk_col])
            status = normalize(row[status_col]) if status_col is not None else ""

            # 6. Non-empty MCA_ID must be a recognised value
            if mca_id and mca_id not in VALID_MCA_IDS:
                fail(
                    f"{sheet_name}: invalid MCA_ID '{mca_id}' — "
                    f"must be one of {sorted(VALID_MCA_IDS)}"
                )

            # 7. Non-empty Compliance_Risk must be a recognised value
            if compliance_risk and compliance_risk not in VALID_COMPLIANCE_RISKS:
                fail(
                    f"{sheet_name}: invalid Compliance_Risk '{compliance_risk}' — "
                    f"must be one of {sorted(VALID_COMPLIANCE_RISKS)}"
                )

            # 8. REVIEW_ONLY Compliance_Risk rows must not be APPROVED or LOCKED
            if compliance_risk == "REVIEW_ONLY" and status in APPROVED_STATUSES:
                fail(
                    f"{sheet_name}: row with Compliance_Risk=REVIEW_ONLY "
                    f"must not have Status={status}"
                )

            # 9. Reject forbidden MWCB copy terms in any approved row
            if status in APPROVED_STATUSES and sheet_name in MWCB_SHEETS:
                copy_fields = []
                for col_name in ("Hook", "Pain_or_Friction", "USP_1", "USP_2", "USP_3", "CTA"):
                    col_idx = header_index.get(col_name)
                    if col_idx is not None:
                        copy_fields.append(normalize(row[col_idx]))
                blob = " | ".join(copy_fields).lower()
                for term in MWCB_FORBIDDEN_COPY_TERMS:
                    if term.lower() in blob:
                        fail(
                            f"{sheet_name}: approved row contains forbidden MWCB "
                            f"copy term '{term}'"
                        )


def validate_mwcb_phase1_distribution() -> None:
    if not MWCB_TAXONOMY_PATH.exists():
        fail(f"MWCB taxonomy registry missing: {MWCB_TAXONOMY_PATH}")

    with MWCB_TAXONOMY_PATH.open("r", encoding="utf-8") as fh:
        taxonomy = yaml.safe_load(fh)

    forbidden_review_phrases: set[str] = set()
    for use_case in taxonomy.get("review_only_use_cases", []):
        forbidden_review_phrases.update(normalize(item).lower() for item in use_case.get("forbidden_phrases", []))

    if not WORKBOOK_PATH.exists():
        fail(f"Workbook missing: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    ws = wb["PRODUCT_MW_CAP_BURUNG"]
    headers = [normalize(cell.value) for cell in ws[1]]
    header_index = {header: idx for idx, header in enumerate(headers)}

    required_headers = {
        "Row_ID",
        "Angle_ID",
        "MCA_ID",
        "Compliance_Risk",
        "Hook_ID",
        "CTA_ID",
        "Hook",
        "Pain_or_Friction",
        "USP_1",
        "USP_2",
        "USP_3",
        "CTA",
        "Copywriting_Formula",
        "Status",
        "Notes",
    }
    missing_headers = [header for header in required_headers if header not in header_index]
    if missing_headers:
        fail(f"PRODUCT_MW_CAP_BURUNG missing headers for Phase 1 validation: {', '.join(missing_headers)}")

    phase1_rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        notes = normalize(row[header_index["Notes"]])
        if not notes.startswith(PHASE1_MARKER):
            continue
        phase1_rows.append(
            {
                "row_id": normalize(row[header_index["Row_ID"]]),
                "angle_id": normalize(row[header_index["Angle_ID"]]),
                "mca_id": normalize(row[header_index["MCA_ID"]]),
                "risk": normalize(row[header_index["Compliance_Risk"]]),
                "hook_id": normalize(row[header_index["Hook_ID"]]),
                "cta_id": normalize(row[header_index["CTA_ID"]]),
                "hook": normalize(row[header_index["Hook"]]),
                "pain": normalize(row[header_index["Pain_or_Friction"]]),
                "usp_1": normalize(row[header_index["USP_1"]]),
                "usp_2": normalize(row[header_index["USP_2"]]),
                "usp_3": normalize(row[header_index["USP_3"]]),
                "cta": normalize(row[header_index["CTA"]]),
                "formula": normalize(row[header_index["Copywriting_Formula"]]),
                "status": normalize(row[header_index["Status"]]),
                "notes": notes,
            }
        )

    if len(phase1_rows) != 36:
        fail(f"MWCB Phase 1 row count mismatch. Expected 36, found {len(phase1_rows)}")

    row_ids = [row["row_id"] for row in phase1_rows]
    angle_ids = [row["angle_id"] for row in phase1_rows]
    hook_ids = [row["hook_id"] for row in phase1_rows]
    cta_ids = [row["cta_id"] for row in phase1_rows]
    for label, values in {
        "Row_ID": row_ids,
        "Angle_ID": angle_ids,
        "Hook_ID": hook_ids,
        "CTA_ID": cta_ids,
    }.items():
        if len(values) != len(set(values)):
            fail(f"MWCB Phase 1 duplicate {label} detected")

    expected_row_ids = [f"PRODUCT_MW_CAP_BURUNG_R{index:03d}" for index in range(31, 67)]
    if sorted(row_ids) != expected_row_ids:
        fail("MWCB Phase 1 Row_ID range mismatch; expected PRODUCT_MW_CAP_BURUNG_R031..R066")

    mca_counts = Counter(row["mca_id"] for row in phase1_rows)
    expected_mca_counts = Counter({mca_id: 12 for mca_id in PHASE1_MCA_IDS})
    if mca_counts != expected_mca_counts:
        fail(f"MWCB Phase 1 MCA distribution mismatch: {mca_counts}")

    formula_counts = Counter(row["formula"] for row in phase1_rows)
    expected_formula_counts = Counter({formula: 9 for formula in PHASE1_FORMULAS})
    if formula_counts != expected_formula_counts:
        fail(f"MWCB Phase 1 formula distribution mismatch: {formula_counts}")

    context_counts: Counter[str] = Counter()
    for row in phase1_rows:
        if row["mca_id"] not in PHASE1_MCA_IDS:
            fail(f"MWCB Phase 1 row has invalid MCA_ID: {row['mca_id']}")
        if row["risk"] != "GREEN":
            fail(f"MWCB Phase 1 row must have Compliance_Risk=GREEN: {row['row_id']}")
        if row["formula"] not in PHASE1_FORMULAS:
            fail(f"MWCB Phase 1 row has invalid formula '{row['formula']}': {row['row_id']}")
        if row["status"] != "APPROVED":
            fail(f"MWCB Phase 1 row must have Status=APPROVED: {row['row_id']}")

        parts = {}
        for segment in row["notes"].split("|")[1:]:
            if "=" not in segment:
                continue
            key, value = segment.split("=", 1)
            parts[key] = value
        ctx = parts.get("CTX")
        if ctx not in PHASE1_CONTEXTS:
            fail(f"MWCB Phase 1 row has invalid context marker '{ctx}': {row['row_id']}")
        if parts.get("MCA") != row["mca_id"]:
            fail(f"MWCB Phase 1 notes/MCA mismatch: {row['row_id']}")
        if parts.get("FORMULA") != row["formula"]:
            fail(f"MWCB Phase 1 notes/formula mismatch: {row['row_id']}")
        context_counts[ctx] += 1

        blob = " | ".join(
            [row["hook"], row["pain"], row["usp_1"], row["usp_2"], row["usp_3"], row["cta"], row["notes"]]
        ).lower()
        for term in MWCB_FORBIDDEN_COPY_TERMS:
            if term in blob:
                fail(f"MWCB Phase 1 row contains forbidden MWCB term '{term}': {row['row_id']}")
        for phrase in forbidden_review_phrases:
            if phrase and phrase in blob:
                fail(f"MWCB Phase 1 row contains review-only phrase '{phrase}': {row['row_id']}")

    expected_context_counts = Counter({context: 12 for context in PHASE1_CONTEXTS})
    if context_counts != expected_context_counts:
        fail(f"MWCB Phase 1 context distribution mismatch: {context_counts}")


def main() -> None:
    row_counts = validate_workbook()
    validate_registry()
    validate_stealth_authority_map()
    validate_mwcb_taxonomy()
    validate_mwcb_phase1_distribution()

    print("VALIDATION PASSED")
    print(f"Workbook: {WORKBOOK_PATH}")
    print(f"Registry: {REGISTRY_PATH}")
    print(f"Stealth Authority Map: {AUTHORITY_MAP_PATH}")
    print(f"MWCB Taxonomy: {MWCB_TAXONOMY_PATH}")
    for sheet_name in sorted(row_counts):
        if row_counts[sheet_name]:
            print(f"{sheet_name}: {row_counts[sheet_name]} populated rows validated")


if __name__ == "__main__":
    main()
