from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from resolver_runtime import (
    ResolverError,
    assert_avatar_context_pack_runtime,
    resolve_avatar_context_id,
    resolve_avatar_pool,
)
from build_avatar_context_resolver import (
    COMMAND_CENTRE_AVATAR_SHEET_ALIAS,
    COMMAND_CENTRE_POOL_SHEET_ALIAS,
)


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "BOSMAX_AVATAR_CONTEXT_RESOLVER_v1.xlsx"
REGISTRY_PATH = ROOT / "registries" / "avatar_context_rotation.yaml"

REQUIRED_SHEETS = {
    "AVATAR_CONTEXT_PACKS",
    "AVATAR_PERSONA_INDEX",
    "VIEW_ALIASES",
    "MANNEQUIN_POSE_LIBRARY",
    "SCENE_CONTEXT_LIBRARY",
    "ROTATION_POOLS",
    COMMAND_CENTRE_AVATAR_SHEET_ALIAS,
    COMMAND_CENTRE_POOL_SHEET_ALIAS,
    "NOTION_EXPORT_VIEW",
    "VALIDATION_RULES",
    "CHANGELOG",
}
EXPECTED_VIEW_ALIASES = {
    "NOTION_COMMAND_CENTRE_AVATAR_ID_VIEW": COMMAND_CENTRE_AVATAR_SHEET_ALIAS,
    "NOTION_COMMAND_CENTRE_AVATAR_POOL_VIEW": COMMAND_CENTRE_POOL_SHEET_ALIAS,
}
FORBIDDEN_NOTION_HEADERS = {
    "Prompt_Fragment_Source",
    "Internal_Notes",
    "Compatible_Physics_Classes",
}
EXPECTED_NOTION_HEADERS = [
    "Avatar_Context_ID",
    "Display_Name",
    "Persona_Label",
    "Gender",
    "Age_Range",
    "Silo_Allowed",
    "Product_Family_Allowed",
    "Scene_Label",
    "Mannequin_Label",
    "Camera_Style_Allowed",
    "Status",
    "Safe_Usage_Notes",
]
EXPECTED_COMMAND_CENTRE_AVATAR_HEADERS = [
    "Avatar_Context_ID",
    "Display_Name",
    "Persona_Label",
    "Gender",
    "Age_Range",
    "Silo_Allowed",
    "Product_Family_Allowed",
    "Scene_Label",
    "Mannequin_Label",
    "Camera_Style_Allowed",
    "Status",
    "Safe_Usage_Notes",
]
EXPECTED_COMMAND_CENTRE_POOL_HEADERS = [
    "Pool_ID",
    "Display_Name",
    "Product_ID",
    "Product_Family",
    "Silo",
    "Rotation_Mode",
    "No_Repeat_Window",
    "Minimum_Approved_Count",
    "Status",
    "Safe_Usage_Notes",
]
FORBIDDEN_COMMAND_CENTRE_HEADERS = {
    "Prompt_Fragment_Source",
    "Internal_Notes",
    "Compatible_Physics_Classes",
    "Allowed_Avatar_Context_IDs",
    "Runtime_Allowed",
}


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def fail(message: str) -> None:
    print(f"FAIL: {message}")
    sys.exit(1)


def expect(condition: bool, message: str) -> None:
    if not condition:
        fail(message)


def load_yaml_registry() -> dict[str, Any]:
    with REGISTRY_PATH.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def validate_workbook_shape() -> None:
    expect(WORKBOOK_PATH.exists(), f"Workbook missing: {WORKBOOK_PATH}")
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    expect(REQUIRED_SHEETS.issubset(set(workbook.sheetnames)), "Avatar resolver workbook missing required sheets")

    notion_sheet = workbook["NOTION_EXPORT_VIEW"]
    notion_headers = [normalize(cell.value) for cell in next(notion_sheet.iter_rows(max_row=1))]
    expect(notion_headers == EXPECTED_NOTION_HEADERS, "NOTION_EXPORT_VIEW headers drift detected")
    expect(not FORBIDDEN_NOTION_HEADERS.intersection(set(notion_headers)), "NOTION_EXPORT_VIEW exposes forbidden fields")

    command_avatar_sheet = workbook[COMMAND_CENTRE_AVATAR_SHEET_ALIAS]
    command_avatar_headers = [normalize(cell.value) for cell in next(command_avatar_sheet.iter_rows(max_row=1))]
    expect(command_avatar_headers == EXPECTED_COMMAND_CENTRE_AVATAR_HEADERS, f"{COMMAND_CENTRE_AVATAR_SHEET_ALIAS} headers drift detected")
    expect(not FORBIDDEN_COMMAND_CENTRE_HEADERS.intersection(set(command_avatar_headers)), f"{COMMAND_CENTRE_AVATAR_SHEET_ALIAS} exposes forbidden fields")

    command_pool_sheet = workbook[COMMAND_CENTRE_POOL_SHEET_ALIAS]
    command_pool_headers = [normalize(cell.value) for cell in next(command_pool_sheet.iter_rows(max_row=1))]
    expect(command_pool_headers == EXPECTED_COMMAND_CENTRE_POOL_HEADERS, f"{COMMAND_CENTRE_POOL_SHEET_ALIAS} headers drift detected")
    expect(not FORBIDDEN_COMMAND_CENTRE_HEADERS.intersection(set(command_pool_headers)), f"{COMMAND_CENTRE_POOL_SHEET_ALIAS} exposes forbidden fields")

    alias_sheet = workbook["VIEW_ALIASES"]
    alias_rows = list(alias_sheet.iter_rows(min_row=2, values_only=True))
    alias_map = {normalize(row[0]): normalize(row[1]) for row in alias_rows if normalize(row[0])}
    expect(alias_map == EXPECTED_VIEW_ALIASES, "VIEW_ALIASES sheet drift detected")
    print("workbook shape ok")


def validate_registry_shape(registry: dict[str, Any]) -> None:
    expect(REGISTRY_PATH.exists(), f"Registry missing: {REGISTRY_PATH}")
    expect(registry.get("registry_id") == "BOSMAX_AVATAR_CONTEXT_ROTATION_V1", "registry_id drift detected")
    expect(bool(registry.get("avatar_context_packs")), "avatar_context_packs cannot be empty")
    expect(bool(registry.get("rotation_pools")), "rotation_pools cannot be empty")
    expect(bool(registry.get("notion_command_centre_avatar_id_view")), "notion_command_centre_avatar_id_view cannot be empty")
    expect(bool(registry.get("notion_command_centre_avatar_pool_view")), "notion_command_centre_avatar_pool_view cannot be empty")
    expect(registry.get("samples", {}).get("sample_avatar_context_id") == "BOSMAX_AVP_0001", "sample_avatar_context_id drift detected")
    expect(registry.get("samples", {}).get("sample_avatar_pool_id") == "BOSMAX_MALE_STEALTH_POOL_001", "sample_avatar_pool_id drift detected")
    expect(registry.get("samples", {}).get("sample_mwcb_avatar_context_id") == "MWCB_DIRECT_AVP_0001", "sample_mwcb_avatar_context_id drift detected")
    expect(registry.get("samples", {}).get("sample_mwcb_avatar_pool_id") == "MWCB_TRAD_REMEDY_POOL_001", "sample_mwcb_avatar_pool_id drift detected")
    expect(registry.get("runtime_contract", {}).get("default_notion_flow") == "COMMAND_CENTRE_PLUG_AND_PLAY", "default_notion_flow drift detected")
    legacy = registry.get("legacy_expert_mode", {})
    expect(legacy.get("label") == "LEGACY_EXPERT_MODE", "legacy expert label drift detected")
    expect(legacy.get("manual_override_posture") == "MANUAL_OVERRIDE_REVIEW_ONLY", "manual override posture drift detected")
    alias_map = {
        normalize(row.get("public_view_id")): normalize(row.get("workbook_sheet_alias"))
        for row in registry.get("workbook_sheet_aliases", [])
    }
    expect(alias_map == EXPECTED_VIEW_ALIASES, "workbook_sheet_aliases registry drift detected")
    print("registry shape ok")


def validate_runtime_records(registry: dict[str, Any]) -> None:
    packs = registry.get("avatar_context_packs", [])
    pool_index = {normalize(pool.get("pool_id")): pool for pool in registry.get("rotation_pools", [])}
    expect("BOSMAX_MALE_STEALTH_POOL_001" in pool_index, "BOSMAX_MALE_STEALTH_POOL_001 missing")
    expect("MWCB_TRAD_REMEDY_POOL_001" in pool_index, "MWCB_TRAD_REMEDY_POOL_001 missing")

    for pack in packs:
        assert_avatar_context_pack_runtime(
            pack,
            expected_template_silo=(pack.get("silo_allowed") or [""])[0],
            expected_product_family=(pack.get("product_family_allowed") or [""])[0],
            expected_camera_style=(pack.get("camera_style_allowed") or [""])[0],
            physics_class=(pack.get("compatible_physics_classes") or ["CLASS_A"])[0],
        )

    pool = pool_index["BOSMAX_MALE_STEALTH_POOL_001"]
    allowed_ids = pool.get("allowed_avatar_context_ids", [])
    stealth_ids = [
        normalize(pack.get("avatar_context_id"))
        for pack in packs
        if "STEALTH" in [normalize(item) for item in pack.get("silo_allowed", [])]
    ]
    expect(sorted(allowed_ids) == sorted(stealth_ids), "Stealth pool should include every STEALTH avatar context pack")
    expect(pool.get("no_repeat_window") == max(1, len(stealth_ids) - 1), "Stealth no-repeat window drift detected")
    expect(pool.get("minimum_approved_count") == len(stealth_ids), "Stealth minimum approved count drift detected")

    mwcb_pool = pool_index["MWCB_TRAD_REMEDY_POOL_001"]
    mwcb_allowed_ids = mwcb_pool.get("allowed_avatar_context_ids", [])
    direct_ids = [
        normalize(pack.get("avatar_context_id"))
        for pack in packs
        if "DIRECT" in [normalize(item) for item in pack.get("silo_allowed", [])]
    ]
    expect(sorted(mwcb_allowed_ids) == sorted(direct_ids), "MWCB pool should include every DIRECT avatar context pack")
    expect(mwcb_pool.get("no_repeat_window") == 1, "MWCB pool no-repeat window drift detected")
    expect(mwcb_pool.get("minimum_approved_count") == len(direct_ids), "MWCB minimum approved count drift detected")

    command_avatar_rows = registry.get("notion_command_centre_avatar_id_view", [])
    command_avatar_headers = set(command_avatar_rows[0].keys())
    expect("prompt_fragment_source" not in command_avatar_headers, "Command Centre avatar view must exclude prompt fragment sources")
    expect("internal_notes" not in command_avatar_headers, "Command Centre avatar view must exclude internal notes")

    command_pool_rows = registry.get("notion_command_centre_avatar_pool_view", [])
    command_pool_headers = set(command_pool_rows[0].keys())
    expect("allowed_avatar_context_ids" not in command_pool_headers, "Command Centre pool view must exclude raw avatar ID lists")
    if any(normalize(row.get("status")) == "SEED_READY" for row in command_avatar_rows + command_pool_rows):
        staging_rows = [row for row in command_avatar_rows + command_pool_rows if normalize(row.get("status")) == "SEED_READY"]
        expect(
            all("STAGING_ONLY" in normalize(row.get("safe_usage_notes")) for row in staging_rows),
            "SEED_READY Command Centre avatar rows must be clearly marked as staging",
        )
    print("runtime record invariants ok")


def validate_negative_paths(registry: dict[str, Any]) -> None:
    sample = resolve_avatar_context_id(
        "BOSMAX_AVP_0001",
        expected_template_silo="STEALTH",
        expected_product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
        expected_camera_style="UGC_IPHONE_RAW",
        physics_class="CLASS_A",
    )
    expect(sample["avatar_context_id"] == "BOSMAX_AVP_0001", "Sample avatar context resolution failed")

    direct_sample = resolve_avatar_context_id(
        "MWCB_DIRECT_AVP_0001",
        expected_template_silo="DIRECT",
        expected_product_family="FAMILY_TRADITIONAL_REMEDY_OIL",
        expected_camera_style="UGC_IPHONE_RAW",
        physics_class="CLASS_A",
    )
    expect(direct_sample["avatar_context_id"] == "MWCB_DIRECT_AVP_0001", "MWCB direct avatar context resolution failed")

    try:
        resolve_avatar_context_id("BOSMAX_AVP_9999", expected_template_silo="STEALTH")
    except ResolverError:
        pass
    else:
        fail("Missing Avatar_Context_ID must fail closed")

    try:
        resolve_avatar_context_id(
            "BOSMAX_AVP_0001",
            expected_template_silo="DIRECT",
            expected_product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_A",
        )
    except ResolverError:
        pass
    else:
        fail("Silo mismatch must fail closed")

    try:
        resolve_avatar_context_id(
            "BOSMAX_AVP_0001",
            expected_template_silo="STEALTH",
            expected_product_family="FAMILY_TRAD_REMEDY_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_A",
        )
    except ResolverError:
        pass
    else:
        fail("Product family mismatch must fail closed")

    try:
        resolve_avatar_context_id(
            "BOSMAX_AVP_0001",
            expected_template_silo="STEALTH",
            expected_product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
            expected_camera_style="CINEMATIC_PRO",
            physics_class="CLASS_A",
        )
    except ResolverError:
        pass
    else:
        fail("Camera style mismatch must fail closed")

    try:
        resolve_avatar_context_id(
            "BOSMAX_AVP_0001",
            expected_template_silo="STEALTH",
            expected_product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_B",
        )
    except ResolverError:
        pass
    else:
        fail("Physics mismatch must fail closed")

    try:
        resolve_avatar_context_id(
            "MWCB_DIRECT_AVP_0001",
            expected_template_silo="STEALTH",
            expected_product_family="FAMILY_TRADITIONAL_REMEDY_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_A",
        )
    except ResolverError:
        pass
    else:
        fail("DIRECT avatar context must fail closed on STEALTH mismatch")

    invalid_status_pack = dict(sample)
    invalid_status_pack["status"] = "DRAFT"
    try:
        assert_avatar_context_pack_runtime(
            invalid_status_pack,
            expected_template_silo="STEALTH",
            expected_product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_A",
        )
    except ResolverError:
        pass
    else:
        fail("Invalid avatar context status must fail closed")

    try:
        resolve_avatar_pool(
            "BOSMAX_MALE_STEALTH_POOL_001",
            batch_count=20,
            rotation_rule="ROUND_ROBIN_NO_REPEAT",
            expected_template_silo="STEALTH",
            expected_product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_A",
        )
    except ResolverError as exc:
        fail(f"Pool rotation positive sample failed unexpectedly: {exc}")

    try:
        resolve_avatar_pool(
            "MWCB_TRAD_REMEDY_POOL_001",
            batch_count=4,
            rotation_rule="ROUND_ROBIN_NO_REPEAT",
            expected_template_silo="DIRECT",
            expected_product_family="FAMILY_TRADITIONAL_REMEDY_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_A",
        )
    except ResolverError as exc:
        fail(f"MWCB direct pool rotation positive sample failed unexpectedly: {exc}")

    try:
        resolve_avatar_pool(
            "BOSMAX_MALE_STEALTH_POOL_001",
            batch_count=20,
            rotation_rule="INVALID_RULE",
            expected_template_silo="STEALTH",
            expected_product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_A",
        )
    except ResolverError:
        pass
    else:
        fail("Invalid rotation rule must fail closed")

    shallow_pool = dict(registry["rotation_pools"][0])
    shallow_pool["minimum_approved_count"] = len(shallow_pool["allowed_avatar_context_ids"]) + 1
    try:
        resolve_avatar_pool(
            shallow_pool["pool_id"],
            batch_count=20,
            rotation_rule="ROUND_ROBIN_NO_REPEAT",
            expected_template_silo="STEALTH",
            expected_product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
            expected_camera_style="UGC_IPHONE_RAW",
            physics_class="CLASS_A",
            registry={
                **registry,
                "rotation_pools": [shallow_pool],
            },
        )
    except ResolverError:
        pass
    else:
        fail("Pool with insufficient approved count must fail closed")

    print("negative fail-closed checks ok")


def main() -> None:
    registry = load_yaml_registry()
    validate_workbook_shape()
    validate_registry_shape(registry)
    validate_runtime_records(registry)
    validate_negative_paths(registry)
    print("VALIDATION PASSED: avatar context resolver")


if __name__ == "__main__":
    main()
