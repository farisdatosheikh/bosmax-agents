from __future__ import annotations

from collections import Counter
from pathlib import Path
import re
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx"
SHEET_NAME = "PRODUCT_MW_CAP_BURUNG"
EXPECTED_EXISTING_ROWS = 30
PHASE1_START = 31

AUTHORITY_SOURCE = (
    "registries/mwcb_copywriting_angle_taxonomy.yaml + "
    "products/CAP_BURUNG_MINYAK.yaml label benefits"
)
PRODUCT_SCALE = (
    "EXACTLY a small pocket-size 30ml rectangular clear glass medicated oil bottle "
    "with a red ribbed cap, shorter than the palm and only two fingers wide"
)
REQUIRED_HEADERS = [
    "Row_ID",
    "Family_Code",
    "Family_Name",
    "Product_ID_Optional",
    "Product_Name_Optional",
    "SKU_Optional",
    "Category",
    "Sub_Category",
    "Product_Type",
    "UOM",
    "Product_Size",
    "Product_Scale",
    "Type_of_Content",
    "Silo_Key",
    "Angle_ID",
    "MCA_ID",
    "Compliance_Risk",
    "Angle",
    "Hook_ID",
    "Hook",
    "Pain_or_Friction",
    "USP_1",
    "USP_2",
    "USP_3",
    "CTA_ID",
    "CTA",
    "Copywriting_Formula",
    "Authority_Source",
    "Fastmoss_Reference",
    "Status",
    "Notes",
]
FORBIDDEN_TERMS = (
    "roll-on",
    "roll on",
    "roller",
    "rollerball",
    "ubat",
    "sembuh",
    "merawat",
    "penawar",
    "rawatan medis",
    "antiseptik",
    "antibakteria",
    "selamat untuk bayi",
    "selamat untuk kanak-kanak",
    "buka saluran hidung",
    "longgarkan kahak",
    "mengecutkan rahim",
    "mengempiskan perut",
    "sakit dada",
    "kencing malam",
    "bayi kembung",
    "luka kecil",
    "sakit gigi",
    "bengkak gusi",
    "confirm",
    "guaranteed",
    "100%",
    "hilang terus",
    "pulih",
    "doctor",
    "klinik",
    "farmasi",
    "lebih kuat dari ubat",
)


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def fail(message: str) -> None:
    print(f"ERROR: {message}")
    sys.exit(1)


COPY_BLUEPRINTS = {
    "MWCB-MCA01": {
        "angle": "Kepala Berat & Stress Harian",
        "contexts": {
            "CTX-A_SEGARA": {
                "pain": "Bila kepala mula terasa berat dan serabut, fokus terus terganggu dan kerja rasa makin lambat.",
                "usp_1": "Aroma herba yang menenangkan bantu kepala rasa lebih lega bila disapu pada kawasan sesuai.",
                "usp_2": "Petua warisan keluarga sejak 1958 buat orang lebih yakin simpan sebagai peneman harian.",
                "usp_3": "Botol 30ML kecil mudah capai dari beg, laci meja, atau kereta bila keadaan datang tiba-tiba.",
            },
            "CTX-B_RUTIN": {
                "pain": "Bila rutin harian sentiasa padat, rasa tegang di kepala cepat ganggu mood dan momentum.",
                "usp_1": "Sapu sikit masa rasa penat menghadap skrin supaya kepala rasa lebih ringan dan segar semula.",
                "usp_2": "Aroma herba yang familiar sesuai dijadikan ritual ringkas sebelum sambung rutin seterusnya.",
                "usp_3": "Saiz 30ML yang praktikal memang senang disimpan dalam beg kerja atau pouch harian.",
            },
            "CTX-C_SYOR": {
                "pain": "Ramai cuma cari benda ni bila kepala dah rasa penuh, padahal lebih senang bila awal-awal dah standby.",
                "usp_1": "Minyak warisan ini selalu disyor sebab aroma herbanya buat rasa lebih selesa tanpa cerita berlebihan.",
                "usp_2": "Nama yang turun-temurun dalam keluarga bantu bina trust bila nak kenalkan pada orang lain.",
                "usp_3": "Botol kecil dan praktikal memudahkan orang beli satu untuk diri sendiri dan satu lagi untuk orang rumah.",
            },
        },
        "hooks": {
            "PAS": {
                "CTX-A_SEGARA": "Kepala dah mula rasa berat? Sapu sikit dulu sebelum rasa serabut tu makin mengganggu.",
                "CTX-B_RUTIN": "Kalau kepala selalu rasa tegang lepas hadap rutin panjang, lebih baik standby awal-awal.",
                "CTX-C_SYOR": "Ramai orang syor benda ni bila kepala cepat rasa penuh masa hari sibuk.",
            },
            "AIDA": {
                "CTX-A_SEGARA": "Tengah buat kerja tapi kepala rasa padat sangat? Capai yang kecil ni dulu.",
                "CTX-B_RUTIN": "Ada sebab ramai orang kerja simpan satu botol kecil macam ni dekat meja.",
                "CTX-C_SYOR": "Sekali orang rumah syor minyak ni, baru faham kenapa ramai tak tinggal botol kecil ni.",
            },
            "HSO": {
                "CTX-A_SEGARA": "Baru beberapa jam hadap skrin, kepala terus rasa tegang. Masa tu benda ni memang berguna.",
                "CTX-B_RUTIN": "Ada orang sebelum sambung rutin petang, dia sapu dulu minyak warisan yang memang selalu standby.",
                "CTX-C_SYOR": "Kawan ofis selalu hulur botol ni bila nampak muka orang dah mula serabut.",
            },
            "HPAS": {
                "CTX-A_SEGARA": "Bila kepala berat datang mengejut, memang terus rasa tak selesa nak sambung apa pun.",
                "CTX-B_RUTIN": "Kalau hari-hari kepala rasa tegang, jangan tunggu sampai rutin terus tunggang-langgang.",
                "CTX-C_SYOR": "Orang yang pernah cuba biasanya terus faham kenapa minyak ni selalu jadi barang syor.",
            },
        },
        "ctas": {
            "PAS": {
                "CTX-A_SEGARA": "Add to Cart dan simpan satu dekat tempat yang paling senang kau capai.",
                "CTX-B_RUTIN": "Order sekarang supaya rutin harian ada satu standby yang praktikal.",
                "CTX-C_SYOR": "Ambil satu untuk diri sendiri dan kongsi satu lagi dengan orang rumah.",
            },
            "AIDA": {
                "CTX-A_SEGARA": "Masukkan dalam cart sekarang dan biar satu botol standby masa perlukan nanti.",
                "CTX-B_RUTIN": "Simpan satu dalam beg kerja sebelum rutin esok bermula.",
                "CTX-C_SYOR": "Klik beg kuning dan syorkan terus pada orang yang selalu hadap hari sibuk.",
            },
            "HSO": {
                "CTX-A_SEGARA": "Scroll bawah dan standby satu botol kecil ni sekarang.",
                "CTX-B_RUTIN": "Tambah satu ke cart supaya rutin petang rasa lebih tersusun.",
                "CTX-C_SYOR": "Dapatkan sekarang kalau kau pun nak ada standby yang senang disyor.",
            },
            "HPAS": {
                "CTX-A_SEGARA": "Terus simpan satu dalam cart sebelum kepala rasa makin tak selesa.",
                "CTX-B_RUTIN": "Order hari ni dan letak siap-siap dalam beg atau laci harian.",
                "CTX-C_SYOR": "Beli sekarang kalau kau nak benda yang memang senang diperkenalkan pada orang lain.",
            },
        },
    },
    "MWCB-MCA03": {
        "angle": "Perut Tak Selesa & Wanita Harian",
        "contexts": {
            "CTX-A_SEGARA": {
                "pain": "Bila perut rasa tak selesa atau kembung ringan datang tiba-tiba, semua rancangan terus rasa tergendala.",
                "usp_1": "Sapuan dan urutan lembut beri rasa hangat yang buat perut rasa lebih selesa.",
                "usp_2": "Aroma herba yang familiar bantu tenangkan momen tak selesa tanpa bahasa yang keterlaluan.",
                "usp_3": "Botol 30ML kecil memang senang bawa dalam beg tangan atau pouch harian.",
            },
            "CTX-B_RUTIN": {
                "pain": "Rasa tak selesa yang datang berulang memang mudah kacau mood bila jadual seharian dah penuh.",
                "usp_1": "Sapu sikit bila rasa perut tak selesa atau bila datang bulan supaya badan rasa lebih tenang.",
                "usp_2": "Petua warisan wanita dalam keluarga buat orang lebih mudah percaya untuk simpan sebagai standby.",
                "usp_3": "Saiz kecil dan praktikal memudahkan ia kekal ada dalam beg, meja sisi, atau kereta.",
            },
            "CTX-C_SYOR": {
                "pain": "Ramai hanya teringat cari benda ni bila perjalanan dah bermula atau bila rasa tak selesa dah datang duluan.",
                "usp_1": "Ramai syor minyak ni untuk rasa mual semasa perjalanan dan rasa tak selesa harian yang ringan.",
                "usp_2": "Warisan keluarga sejak 1958 bagi rasa yakin bila nak rekomen pada kawan atau saudara.",
                "usp_3": "Botol 30ML kecil dan mudah simpan memang sesuai dibawa ke mana-mana tanpa makan ruang.",
            },
        },
        "hooks": {
            "PAS": {
                "CTX-A_SEGARA": "Perut dah mula rasa tak selesa? Sapu dan urut lembut dulu sebelum mood terus berubah.",
                "CTX-B_RUTIN": "Kalau rasa tak selesa ni selalu datang, lagi baik ada satu standby dalam beg.",
                "CTX-C_SYOR": "Benda macam ni memang selalu orang perempuan syor sesama sendiri bila rasa tak selesa datang.",
            },
            "AIDA": {
                "CTX-A_SEGARA": "Tengah keluar tapi perut tiba-tiba rasa tak selesa? Capai yang kecil ni dulu.",
                "CTX-B_RUTIN": "Ada sebab ramai wanita suka simpan satu botol kecil ni sebagai standby harian.",
                "CTX-C_SYOR": "Sekali kawan syor minyak ni untuk travel, terus faham kenapa ia selalu ada dalam beg.",
            },
            "HSO": {
                "CTX-A_SEGARA": "Masa tengah jalan, perut mula rasa tak sedap. Waktu tu memang sedap ada standby kecil macam ni.",
                "CTX-B_RUTIN": "Ada yang memang tak tunggu rasa tak selesa datang, dia terus simpan awal-awal dalam pouch.",
                "CTX-C_SYOR": "Kawan travel selalu pesan, kalau nak perjalanan lebih tenang bawa sekali minyak warisan ni.",
            },
            "HPAS": {
                "CTX-A_SEGARA": "Bila perut rasa tak selesa datang tiba-tiba, memang susah nak terus tenang dan fokus.",
                "CTX-B_RUTIN": "Kalau rasa tak selesa ni selalu ganggu rutin, jangan tunggu sampai badan terus penat.",
                "CTX-C_SYOR": "Bila dah jumpa benda yang senang bantu rasa lebih selesa, memang mudah nak syor pada orang lain.",
            },
        },
        "ctas": {
            "PAS": {
                "CTX-A_SEGARA": "Add to Cart dan simpan satu dalam beg untuk kegunaan yang lebih praktikal.",
                "CTX-B_RUTIN": "Order sekarang supaya pouch harian kau ada satu standby yang mudah capai.",
                "CTX-C_SYOR": "Ambil satu untuk diri sendiri dan kongsi satu lagi dengan kawan rapat.",
            },
            "AIDA": {
                "CTX-A_SEGARA": "Masukkan dalam cart sekarang sebelum keluar rumah tanpa standby.",
                "CTX-B_RUTIN": "Klik beg kuning dan jadikan ia sebahagian daripada kit harian kau.",
                "CTX-C_SYOR": "Dapatkan sekarang kalau kau nak standby yang memang senang diperkenalkan pada orang lain.",
            },
            "HSO": {
                "CTX-A_SEGARA": "Scroll bawah dan standby satu botol kecil ni untuk perjalanan seterusnya.",
                "CTX-B_RUTIN": "Tambah satu ke cart supaya rutin harian rasa lebih bersedia.",
                "CTX-C_SYOR": "Beli sekarang dan simpan siap-siap dalam beg yang selalu ikut kau keluar.",
            },
            "HPAS": {
                "CTX-A_SEGARA": "Terus order satu sebelum rasa tak selesa datang tanpa persediaan.",
                "CTX-B_RUTIN": "Simpan satu sekarang dan bagi rutin harian ada pilihan yang lebih mudah capai.",
                "CTX-C_SYOR": "Add to Cart kalau kau nak standby yang memang orang senang faham bila disyor.",
            },
        },
    },
    "MWCB-MCA07": {
        "angle": "Satu Botol Standby Rumah",
        "contexts": {
            "CTX-A_SEGARA": {
                "pain": "Bila rumah tiba-tiba perlukan sesuatu yang cepat dicapai, memang rasa susah kalau semuanya tak ready.",
                "usp_1": "Satu botol minyak warisan untuk banyak situasi harian yang perlukan rasa lebih selesa dan praktikal.",
                "usp_2": "Sejak 1958 ramai keluarga simpan sebagai petua rumah yang senang dicapai bila perlu.",
                "usp_3": "Botol 30ML kecil mudah letak di dapur, ruang tamu, beg, atau kereta.",
            },
            "CTX-B_RUTIN": {
                "pain": "Rumah yang tak ada standby selalunya buat orang tercari-cari benda yang patutnya dah siap lebih awal.",
                "usp_1": "Mudah dijadikan rutin simpan satu dekat rumah, satu dalam kereta, dan satu lagi dalam beg.",
                "usp_2": "Warisan keluarga yang practical buat orang rasa lebih yakin untuk simpan lama-lama sebagai standby.",
                "usp_3": "Saiz 30ML kecil, mudah bawa, dan tak makan ruang dalam tempat simpanan harian.",
            },
            "CTX-C_SYOR": {
                "pain": "Ramai baru teringat tentang minyak warisan macam ni lepas orang lain tunjuk betapa senangnya ada satu botol standby.",
                "usp_1": "Minyak warisan ini selalu disyor sebab satu botol saja dah cukup praktikal untuk situasi harian ringan.",
                "usp_2": "Nama yang dah lama dikenali dalam keluarga buat orang senang percaya dan senang hadiahkan.",
                "usp_3": "Botol kecil, mudah simpan, dan memang sesuai dibeli lebih dari satu untuk rumah dan kereta.",
            },
        },
        "hooks": {
            "PAS": {
                "CTX-A_SEGARA": "Rumah kena ada satu botol standby macam ni sebelum datang momen yang memerlukan tindakan cepat.",
                "CTX-B_RUTIN": "Kalau rumah asyik tercari-cari standby, memang sudah masa simpan satu botol yang betul.",
                "CTX-C_SYOR": "Benda macam ni selalunya orang syor sebab sekali simpan, terus rasa praktikalnya.",
            },
            "AIDA": {
                "CTX-A_SEGARA": "Satu botol kecil je, tapi selalu jadi benda pertama yang orang capai bila rumah perlukan standby.",
                "CTX-B_RUTIN": "Ada sebab ramai keluarga suka ada lebih dari satu botol kecil macam ni di rumah.",
                "CTX-C_SYOR": "Sekali orang rumah kenalkan minyak ni, terus nampak kenapa ia sesuai jadi standby keluarga.",
            },
            "HSO": {
                "CTX-A_SEGARA": "Tengah sibuk, tiba-tiba perlukan sesuatu yang cepat dicapai. Masa tu botol kecil ni memang terasa gunanya.",
                "CTX-B_RUTIN": "Ada keluarga yang memang simpan siap-siap dekat rumah dan kereta sebab tak suka kelam-kabut bila perlu.",
                "CTX-C_SYOR": "Biasanya orang yang dah lama simpan minyak ni akan pesan benda sama, elok ada satu botol standby.",
            },
            "HPAS": {
                "CTX-A_SEGARA": "Bila rumah perlukan standby tapi tak ada, memang terus terasa susah dan lambat nak urus.",
                "CTX-B_RUTIN": "Kalau benda kecil macam ni pun selalu tiada masa diperlukan, rutin rumah memang mudah jadi serabut.",
                "CTX-C_SYOR": "Sekali orang tunjuk betapa praktikalnya satu botol standby ni, memang mudah terus masuk senarai beli.",
            },
        },
        "ctas": {
            "PAS": {
                "CTX-A_SEGARA": "Add to Cart dan letak satu terus di tempat standby rumah.",
                "CTX-B_RUTIN": "Order sekarang dan lengkapkan rutin simpanan rumah dengan satu botol praktikal.",
                "CTX-C_SYOR": "Ambil satu untuk rumah sendiri dan satu lagi untuk orang yang kau selalu ingatkan.",
            },
            "AIDA": {
                "CTX-A_SEGARA": "Masukkan dalam cart sekarang supaya rumah ada satu standby yang senang dicapai.",
                "CTX-B_RUTIN": "Klik beg kuning dan simpan satu dekat rumah sebelum rutin esok bermula.",
                "CTX-C_SYOR": "Dapatkan sekarang kalau kau nak hadiah praktikal yang memang orang akan guna.",
            },
            "HSO": {
                "CTX-A_SEGARA": "Scroll bawah dan standby satu botol kecil ni dekat rumah hari ini.",
                "CTX-B_RUTIN": "Tambah satu ke cart dan bagi rumah kau satu standby yang lebih tersusun.",
                "CTX-C_SYOR": "Beli sekarang dan teruskan syor yang memang senang diterima orang rumah.",
            },
            "HPAS": {
                "CTX-A_SEGARA": "Terus order satu sebelum rumah sekali lagi kena cari standby saat akhir.",
                "CTX-B_RUTIN": "Simpan satu sekarang dan jadikan rumah lebih bersedia setiap hari.",
                "CTX-C_SYOR": "Add to Cart kalau kau nak satu botol standby yang memang senang disyor dan dihadiahkan.",
            },
        },
    },
}


def build_phase1_rows() -> list[dict[str, str]]:
    formulas = ("PAS", "AIDA", "HSO", "HPAS")
    contexts = ("CTX-A_SEGARA", "CTX-B_RUTIN", "CTX-C_SYOR")
    mca_order = ("MWCB-MCA01", "MWCB-MCA03", "MWCB-MCA07")
    rows: list[dict[str, str]] = []
    counter = PHASE1_START

    for mca_id in mca_order:
        blueprint = COPY_BLUEPRINTS[mca_id]
        mca_token = mca_id.split("-")[-1]
        for formula in formulas:
            for ctx in contexts:
                ctx_token = ctx.replace("-", "_")
                angle_id = f"MWCB_P1_{mca_token}_{formula}_{ctx_token}"
                content = blueprint["contexts"][ctx]
                rows.append(
                    {
                        "row_id": f"PRODUCT_MW_CAP_BURUNG_R{counter:03d}",
                        "angle_id": angle_id,
                        "hook_id": f"{angle_id}_HOOK",
                        "cta_id": f"{angle_id}_CTA",
                        "mca_id": mca_id,
                        "compliance_risk": "GREEN",
                        "angle": f"{blueprint['angle']} — {ctx.replace('CTX-', '').replace('_', ' ').title()}",
                        "hook": blueprint["hooks"][formula][ctx],
                        "pain": content["pain"],
                        "usp_1": content["usp_1"],
                        "usp_2": content["usp_2"],
                        "usp_3": content["usp_3"],
                        "cta": blueprint["ctas"][formula][ctx],
                        "formula": formula,
                        "ctx": ctx,
                        "notes": f"PHASE_1_MWCB|MCA={mca_id}|CTX={ctx}|FORMULA={formula}|FRESH_COPY",
                    }
                )
                counter += 1

    return rows


PHASE1_ROWS = build_phase1_rows()


def validate_copy_manifest(rows: list[dict[str, str]]) -> None:
    if len(rows) != 36:
        fail(f"Expected 36 Phase 1 rows, found {len(rows)}")

    mca_counts = Counter(row["mca_id"] for row in rows)
    if mca_counts != Counter({"MWCB-MCA01": 12, "MWCB-MCA03": 12, "MWCB-MCA07": 12}):
        fail(f"MCA distribution mismatch: {mca_counts}")

    formula_counts = Counter(row["formula"] for row in rows)
    if formula_counts != Counter({"PAS": 9, "AIDA": 9, "HSO": 9, "HPAS": 9}):
        fail(f"Formula distribution mismatch: {formula_counts}")

    ctx_counts = Counter(row["ctx"] for row in rows)
    if ctx_counts != Counter({"CTX-A_SEGARA": 12, "CTX-B_RUTIN": 12, "CTX-C_SYOR": 12}):
        fail(f"Context distribution mismatch: {ctx_counts}")

    seen_row_ids: set[str] = set()
    seen_angle_ids: set[str] = set()
    for row in rows:
        if row["row_id"] in seen_row_ids:
            fail(f"Duplicate row_id detected: {row['row_id']}")
        if row["angle_id"] in seen_angle_ids:
            fail(f"Duplicate angle_id detected: {row['angle_id']}")
        seen_row_ids.add(row["row_id"])
        seen_angle_ids.add(row["angle_id"])

        blob = " | ".join(
            [
                row["angle"],
                row["hook"],
                row["pain"],
                row["usp_1"],
                row["usp_2"],
                row["usp_3"],
                row["cta"],
                row["notes"],
            ]
        ).lower()
        for term in FORBIDDEN_TERMS:
            if term in blob:
                fail(f"Forbidden term '{term}' detected in {row['row_id']}")


def get_header_index(ws) -> dict[str, int]:
    headers = [normalize(cell.value) for cell in ws[1]]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        fail(f"{SHEET_NAME} missing required headers: {', '.join(missing)}")
    return {header: idx for idx, header in enumerate(headers)}


def get_canonical_metadata(ws, header_index: dict[str, int]) -> dict[str, str]:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            return {
                key: normalize(row[idx])
                for key, idx in header_index.items()
            }
    fail(f"{SHEET_NAME} has no canonical data row to inherit metadata from")
    return {}


def build_row_values(
    headers: list[str],
    header_index: dict[str, int],
    canonical: dict[str, str],
    pack: dict[str, str],
) -> list[str]:
    row_values = [""] * len(headers)

    inherited_cols = (
        "Family_Code",
        "Family_Name",
        "Category",
        "Sub_Category",
        "Product_Type",
        "UOM",
    )
    fixed_values = {
        "Product_ID_Optional": "CAP_BURUNG_MINYAK",
        "Product_Name_Optional": "Minyak Warisan Cap Burung",
        "SKU_Optional": "30ML_WG40_BOTTLE",
        "Product_Size": "30ML",
        "Product_Scale": PRODUCT_SCALE,
        "Type_of_Content": "DIRECT",
        "Silo_Key": "traditional_remedy_direct",
        "Authority_Source": AUTHORITY_SOURCE,
        "Fastmoss_Reference": "",
        "Status": "APPROVED",
        "Compliance_Risk": "GREEN",
        "MCA_ID": pack["mca_id"],
        "Row_ID": pack["row_id"],
        "Angle_ID": pack["angle_id"],
        "Hook_ID": pack["hook_id"],
        "CTA_ID": pack["cta_id"],
        "Angle": pack["angle"],
        "Hook": pack["hook"],
        "Pain_or_Friction": pack["pain"],
        "USP_1": pack["usp_1"],
        "USP_2": pack["usp_2"],
        "USP_3": pack["usp_3"],
        "CTA": pack["cta"],
        "Copywriting_Formula": pack["formula"],
        "Notes": pack["notes"],
    }

    for col_name in inherited_cols:
        row_values[header_index[col_name]] = canonical[col_name]
    for col_name, value in fixed_values.items():
        row_values[header_index[col_name]] = value

    return row_values


def apply_phase1_rows() -> None:
    if not WORKBOOK_PATH.exists():
        fail(f"Workbook not found: {WORKBOOK_PATH}")

    validate_copy_manifest(PHASE1_ROWS)

    wb = load_workbook(WORKBOOK_PATH)
    if SHEET_NAME not in wb.sheetnames:
        fail(f"Sheet not found: {SHEET_NAME}")
    ws = wb[SHEET_NAME]

    headers = [normalize(cell.value) for cell in ws[1]]
    header_index = get_header_index(ws)
    canonical = get_canonical_metadata(ws, header_index)

    existing_rows = {}
    row_id_col = header_index["Row_ID"]
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_id = normalize(row[row_id_col])
        if row_id:
            if row_id in existing_rows:
                fail(f"Duplicate existing Row_ID detected in workbook: {row_id}")
            existing_rows[row_id] = row_number

    if len(existing_rows) < EXPECTED_EXISTING_ROWS:
        fail(
            f"Expected at least {EXPECTED_EXISTING_ROWS} existing data rows before Phase 1 patch, "
            f"found {len(existing_rows)}"
        )

    for pack in PHASE1_ROWS:
        row_values = build_row_values(headers, header_index, canonical, pack)
        row_number = existing_rows.get(pack["row_id"])
        if row_number is None:
            ws.append(row_values)
            row_number = ws.max_row
            existing_rows[pack["row_id"]] = row_number
        else:
            for idx, value in enumerate(row_values, start=1):
                ws.cell(row=row_number, column=idx).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Patched workbook: {WORKBOOK_PATH}")
    print(f"Updated sheet: {SHEET_NAME}")
    print(f"Phase 1 rows upserted: {len(PHASE1_ROWS)}")


if __name__ == "__main__":
    apply_phase1_rows()
