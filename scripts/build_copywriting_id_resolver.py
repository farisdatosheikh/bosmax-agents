from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SOURCE_WORKBOOK_PATH = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx"
OUTPUT_WORKBOOK_PATH = ROOT / "BOSMAX_COPYWRITING_ID_RESOLVER_v1.xlsx"
OUTPUT_REGISTRY_PATH = ROOT / "registries" / "copywriting_id_resolver.yaml"
PRODUCTS_DIR = ROOT / "products"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

SAFE_NOTION_HEADERS = [
    "Copywriting_ID",
    "Display_Name",
    "Product_Name",
    "Family_Name",
    "Lane",
    "Silo_Key",
    "Submode_Formula",
    "Angle",
    "Hook",
    "USP_1",
    "USP_2",
    "USP_3",
    "CTA",
    "Compliance",
    "Status",
    "Safe_Usage_Notes",
]

REGISTRY_HEADERS = [
    "Copywriting_ID",
    "Display_Name",
    "Product_ID",
    "Product_Name",
    "Family_Code",
    "Family_Name",
    "Lane",
    "Silo_Key",
    "Submode_Formula",
    "Angle_ID",
    "Angle",
    "Hook",
    "Pain_or_Friction",
    "USP_1",
    "USP_2",
    "USP_3",
    "CTA",
    "Compliance",
    "Compliance_Risk",
    "Authority_Source",
    "Source_Script_Node",
    "Source_Variant_Hook_Node",
    "Source_Variant_Problem_Node",
    "Source_Variant_Solution_Node",
    "Source_Variant_CTA_Node",
    "Status",
    "Runtime_Allowed",
    "Notes",
    "Safe_Usage_Notes",
    "Aliases",
    "Legacy_Row_ID",
]

ALIAS_HEADERS = ["Alias", "Normalized_Alias", "Canonical_ID", "Product_ID", "Legacy_Row_ID"]
VALIDATION_RULE_HEADERS = ["Rule_ID", "Rule_Description"]
CHANGELOG_HEADERS = ["Timestamp_UTC", "Change_Note"]

VALIDATION_RULES = [
    ("COPYWRITING_ID_NOT_FOUND", "Fail closed if Copywriting_ID cannot resolve from canonical ID or alias map."),
    ("COPYWRITING_STATUS_INVALID", "Fail closed if Runtime_Allowed is true but Status is not APPROVED, LOCKED, or SEED_READY."),
    ("TEMPLATE_SILO_MISMATCH", "Fail closed if template lane or silo does not match the selected copy pack."),
    ("STEALTH_DIRECT_CROSSOVER", "Fail closed if STEALTH product resolves a DIRECT pack or DIRECT product resolves a STEALTH pack without review."),
    ("BOSMAX_SERUM_PROVENANCE_REQUIRED", "Fail closed if BOSMAX_SERUM STEALTH packs lack required provenance nodes."),
    ("UNSAFE_MANUAL_OVERRIDE", "Fail closed if manual override is present without Needs Compliance Review."),
    ("NOTION_SAFE_EXPORT_ONLY", "Do not expose source paths, raw provenance nodes, or compliance internals in NOTION_EXPORT_VIEW."),
]


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def normalize_alias(value: object) -> str:
    text = normalize(value).upper()
    chars: list[str] = []
    for character in text:
        chars.append(character if character.isalnum() else "_")
    normalized = "".join(chars)
    while "__" in normalized:
        normalized = normalized.replace("__", "_")
    return normalized.strip("_")


def title_fragment(text: str, limit: int = 42) -> str:
    compact = normalize(text).rstrip(".!?")
    return compact[:limit].strip()


def load_products() -> dict[str, dict[str, Any]]:
    products: dict[str, dict[str, Any]] = {}
    for path in sorted(PRODUCTS_DIR.glob("*.yaml")):
        if path.name == "_SCHEMA.yaml":
            continue
        products[path.stem] = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return products


def derive_compliance_risk(row: dict[str, Any], product: dict[str, Any]) -> str:
    workbook_risk = normalize(row.get("Compliance_Risk"))
    if workbook_risk:
        return workbook_risk

    compliance_class = normalize(product.get("compliance_class")).upper()
    lane = normalize(row.get("Type_of_Content")).upper()
    if "STEALTH" in compliance_class or lane == "STEALTH":
        return "STEALTH_REVIEW"
    if compliance_class == "REVIEW_ONLY":
        return "REVIEW_ONLY"
    if compliance_class == "GREEN":
        return "GREEN"
    return "UNDECLARED"


def build_safe_usage_notes(lane: str, compliance: str, product_name: str) -> str:
    if lane == "STEALTH":
        return (
            f"{product_name} STEALTH lane only. Keep repo-approved metaphor copy intact. "
            "Manual override requires Needs Compliance Review."
        )
    return (
        f"{product_name} DIRECT lane only. Do not freestyle replacement copy. "
        "Manual override requires Needs Compliance Review."
    )


def iter_product_rows() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_WORKBOOK_PATH, read_only=True, data_only=True)
    products = load_products()
    sequence_by_product: defaultdict[str, int] = defaultdict(int)
    records: list[dict[str, Any]] = []

    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("PRODUCT_"):
            continue

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [normalize(cell) for cell in next(rows)]
        except StopIteration:
            continue

        header_map = {header: index for index, header in enumerate(headers) if header}
        required_headers = {"Product_ID_Optional", "Hook", "USP_1", "USP_2", "USP_3", "CTA"}
        if not required_headers.issubset(header_map):
            continue

        for row in rows:
            row_data = {
                header: row[index] if index < len(row) else None
                for header, index in header_map.items()
            }
            product_id = normalize(row_data.get("Product_ID_Optional"))
            hook = normalize(row_data.get("Hook"))
            usp_1 = normalize(row_data.get("USP_1"))
            usp_2 = normalize(row_data.get("USP_2"))
            usp_3 = normalize(row_data.get("USP_3"))
            cta = normalize(row_data.get("CTA"))
            if not product_id or not any([hook, usp_1, usp_2, usp_3, cta]):
                continue

            product = products.get(product_id, {})
            sequence_by_product[product_id] += 1
            sequence = sequence_by_product[product_id]
            canonical_id = f"{product_id}_CP_{sequence:04d}"
            product_name = normalize(row_data.get("Product_Name_Optional")) or normalize(product.get("product_name")) or product_id
            lane = normalize(row_data.get("Type_of_Content")).upper()
            silo_key = normalize(row_data.get("Silo_Key"))
            formula = normalize(row_data.get("Copywriting_Formula"))
            angle = normalize(row_data.get("Angle"))
            compliance = normalize(product.get("compliance_class")) or lane
            product_silo = normalize(product.get("silo")).upper() or lane
            aliases = [
                f"{product_id}_{sequence:04d}",
                canonical_id,
                canonical_id.lower(),
            ]
            if product_id == "BOSMAX_SERUM":
                aliases.extend(
                    [
                        f"BOSMAX_{sequence:04d}",
                        f"Bosmax_serum_{sequence:04d}",
                    ]
                )

            notes = normalize(row_data.get("Notes"))
            if not notes:
                notes = f"Generated from {sheet_name} row {normalize(row_data.get('Row_ID'))}."

            record = {
                "copywriting_id": canonical_id,
                "display_name": f"{product_name} | {formula} | {title_fragment(angle, 32)} | {sequence:04d}",
                "product_id": product_id,
                "product_name": product_name,
                "family_code": normalize(row_data.get("Family_Code")),
                "family_name": normalize(row_data.get("Family_Name")),
                "lane": lane,
                "silo_key": silo_key,
                "submode_formula": formula,
                "angle_id": normalize(row_data.get("Angle_ID")),
                "angle": angle,
                "hook": hook,
                "pain_or_friction": normalize(row_data.get("Pain_or_Friction")),
                "usp_1": usp_1,
                "usp_2": usp_2,
                "usp_3": usp_3,
                "cta": cta,
                "compliance": compliance,
                "compliance_risk": derive_compliance_risk(row_data, product),
                "authority_source": normalize(row_data.get("Authority_Source")),
                "source_script_node": normalize(row_data.get("Source_Script_Node")),
                "source_variant_hook_node": normalize(row_data.get("Source_Variant_Hook_Node")),
                "source_variant_problem_node": normalize(row_data.get("Source_Variant_Problem_Node")),
                "source_variant_solution_node": normalize(row_data.get("Source_Variant_Solution_Node")),
                "source_variant_cta_node": normalize(row_data.get("Source_Variant_CTA_Node")),
                "status": normalize(row_data.get("Status")) or "REVIEW_REQUIRED",
                "runtime_allowed": normalize(row_data.get("Status")) in {"APPROVED", "LOCKED", "SEED_READY"},
                "notes": notes,
                "safe_usage_notes": build_safe_usage_notes(lane, compliance, product_name),
                "aliases": sorted(dict.fromkeys(alias for alias in aliases if normalize(alias))),
                "legacy_row_id": normalize(row_data.get("Row_ID")),
                "product_silo": product_silo,
                "require_provenance_nodes": product_id == "BOSMAX_SERUM" and lane == "STEALTH",
                "source_sheet": sheet_name,
            }
            records.append(record)

    records.sort(key=lambda item: item["copywriting_id"])
    return records


def build_registry(records: list[dict[str, Any]]) -> dict[str, Any]:
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    alias_rows: list[dict[str, str]] = []
    notion_rows = []

    for record in records:
        notion_rows.append(
            {
                "copywriting_id": record["copywriting_id"],
                "display_name": record["display_name"],
                "product_name": record["product_name"],
                "family_name": record["family_name"],
                "lane": record["lane"],
                "silo_key": record["silo_key"],
                "submode_formula": record["submode_formula"],
                "angle": record["angle"],
                "hook": record["hook"],
                "usp_1": record["usp_1"],
                "usp_2": record["usp_2"],
                "usp_3": record["usp_3"],
                "cta": record["cta"],
                "compliance": record["compliance"],
                "status": record["status"],
                "safe_usage_notes": record["safe_usage_notes"],
            }
        )

        for alias in record["aliases"]:
            alias_rows.append(
                {
                    "alias": alias,
                    "normalized_alias": normalize_alias(alias),
                    "canonical_id": record["copywriting_id"],
                    "product_id": record["product_id"],
                    "legacy_row_id": record["legacy_row_id"],
                }
            )

    return {
        "schema_version": "1.0",
        "registry_id": "BOSMAX_COPYWRITING_ID_RESOLVER_V1",
        "generated_at_utc": generated_at,
        "source_authority": {
            "workbook": "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx",
            "support_files": [
                "SCRIPT_REGISTRY_UNIFIED.md",
                "SCRIPT_VARIANT_LIBRARY.md",
                "registries/stealth_copy_authority_map.yaml",
                "registries/dialogue_budget_corridor.yaml",
                "BOSMAX_NOTION_COPY_PACK_HANDOFF_v1.md",
            ],
            "notes": "Source workbook remains read-only. Resolver registry is a downstream runtime index only.",
        },
        "runtime_contract": {
            "canonical_id_pattern": r"^[A-Z0-9_]+_CP_[0-9]{4}$",
            "allowed_runtime_statuses": ["APPROVED", "LOCKED", "SEED_READY"],
            "manual_override_requires_review_status": "Needs Compliance Review",
            "notion_safe_fields": [header.lower() for header in SAFE_NOTION_HEADERS],
            "forbidden_notion_fields": [
                "authority_source",
                "source_script_node",
                "source_variant_hook_node",
                "source_variant_problem_node",
                "source_variant_solution_node",
                "source_variant_cta_node",
                "product_silo",
                "require_provenance_nodes",
            ],
        },
        "validation_rules": [
            {"rule_id": rule_id, "description": description}
            for rule_id, description in VALIDATION_RULES
        ],
        "copy_packs": records,
        "alias_map": alias_rows,
        "notion_export_view": notion_rows,
        "samples": {
            "sample_copywriting_id": "BOSMAX_SERUM_CP_0001",
        },
        "counts": {
            "copy_packs": len(records),
            "alias_rows": len(alias_rows),
            "notion_rows": len(notion_rows),
        },
        "changelog": [
            {
                "timestamp_utc": generated_at,
                "change_note": "Generated resolver workbook and YAML registry from approved product workbook rows.",
            }
        ],
    }


def write_yaml_registry(registry: dict[str, Any]) -> None:
    OUTPUT_REGISTRY_PATH.write_text(
        yaml.safe_dump(registry, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )


def style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = WRAP_ALIGNMENT

    for column_cells in worksheet.columns:
        values = [normalize(cell.value) for cell in column_cells[:25]]
        width = min(max((len(value) for value in values), default=10) + 2, 40)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max(width, 12)


def populate_sheet(worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    style_sheet(worksheet)


def write_workbook(registry: dict[str, Any]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    registry_rows = []
    for record in registry["copy_packs"]:
        registry_rows.append(
            [
                record["copywriting_id"],
                record["display_name"],
                record["product_id"],
                record["product_name"],
                record["family_code"],
                record["family_name"],
                record["lane"],
                record["silo_key"],
                record["submode_formula"],
                record["angle_id"],
                record["angle"],
                record["hook"],
                record["pain_or_friction"],
                record["usp_1"],
                record["usp_2"],
                record["usp_3"],
                record["cta"],
                record["compliance"],
                record["compliance_risk"],
                record["authority_source"],
                record["source_script_node"],
                record["source_variant_hook_node"],
                record["source_variant_problem_node"],
                record["source_variant_solution_node"],
                record["source_variant_cta_node"],
                record["status"],
                record["runtime_allowed"],
                record["notes"],
                record["safe_usage_notes"],
                " | ".join(record["aliases"]),
                record["legacy_row_id"],
            ]
        )

    alias_rows = [
        [
            row["alias"],
            row["normalized_alias"],
            row["canonical_id"],
            row["product_id"],
            row["legacy_row_id"],
        ]
        for row in registry["alias_map"]
    ]

    notion_rows = [
        [
            row["copywriting_id"],
            row["display_name"],
            row["product_name"],
            row["family_name"],
            row["lane"],
            row["silo_key"],
            row["submode_formula"],
            row["angle"],
            row["hook"],
            row["usp_1"],
            row["usp_2"],
            row["usp_3"],
            row["cta"],
            row["compliance"],
            row["status"],
            row["safe_usage_notes"],
        ]
        for row in registry["notion_export_view"]
    ]

    rule_rows = [[rule_id, description] for rule_id, description in VALIDATION_RULES]
    changelog_rows = [
        [entry["timestamp_utc"], entry["change_note"]] for entry in registry["changelog"]
    ]

    populate_sheet(workbook.create_sheet("COPY_PACK_REGISTRY"), REGISTRY_HEADERS, registry_rows)
    populate_sheet(workbook.create_sheet("COPY_ID_ALIAS_MAP"), ALIAS_HEADERS, alias_rows)
    populate_sheet(workbook.create_sheet("NOTION_EXPORT_VIEW"), SAFE_NOTION_HEADERS, notion_rows)
    populate_sheet(workbook.create_sheet("VALIDATION_RULES"), VALIDATION_RULE_HEADERS, rule_rows)
    populate_sheet(workbook.create_sheet("CHANGELOG"), CHANGELOG_HEADERS, changelog_rows)

    workbook.save(OUTPUT_WORKBOOK_PATH)


def main() -> None:
    records = iter_product_rows()
    registry = build_registry(records)
    write_yaml_registry(registry)
    write_workbook(registry)
    print(
        f"Generated {OUTPUT_WORKBOOK_PATH.name} and {OUTPUT_REGISTRY_PATH.relative_to(ROOT).as_posix()} "
        f"with {registry['counts']['copy_packs']} copy packs."
    )


if __name__ == "__main__":
    main()
