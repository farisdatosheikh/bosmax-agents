from __future__ import annotations

import argparse
import json
from functools import lru_cache
from pathlib import Path

import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REGISTRY_PATH = ROOT / "registries" / "product_copy_router.yaml"
WORKBOOK_PATH = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx"
PRODUCTS_DIR = ROOT / "products"

def normalize(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return " ".join(text.split())


@lru_cache(maxsize=1)
def load_registry() -> dict:
    with REGISTRY_PATH.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle)


@lru_cache(maxsize=1)
def load_products() -> dict:
    by_id: dict[str, dict] = {}
    by_name: dict[str, dict] = {}

    for path in sorted(PRODUCTS_DIR.glob("*.yaml")):
        if path.name == "_SCHEMA.yaml":
            continue
        with path.open("r", encoding="utf-8") as handle:
            product = yaml.safe_load(handle) or {}
        record = {
            "product_id": product.get("product_id", ""),
            "product_name": product.get("product_name", ""),
            "aliases": product.get("product_aliases", []) or [],
            "category": product.get("category", ""),
            "sub_category": product.get("sub_category", ""),
            "silo": product.get("silo", ""),
            "compliance_class": product.get("compliance_class", ""),
            "product_status": product.get("product_status", ""),
            "dialogue_authority": product.get("dialogue_authority"),
            "source_file": path.relative_to(ROOT).as_posix(),
        }
        if record["product_id"]:
            by_id[normalize(record["product_id"])] = record
        for name in [record["product_name"], *record["aliases"]]:
            if name:
                by_name[normalize(name)] = record

    return {"by_id": by_id, "by_name": by_name}


@lru_cache(maxsize=1)
def load_workbook_index() -> dict:
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    by_id: dict[str, dict] = {}
    by_name: dict[str, dict] = {}
    family_sheets = [name for name in wb.sheetnames if name.startswith("FAMILY_")]

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("PRODUCT_"):
            continue
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            continue
        header_map = {header: idx for idx, header in enumerate(headers) if header}
        for row in rows:
            product_id = row[header_map["Product_ID_Optional"]] if "Product_ID_Optional" in header_map else ""
            product_name = row[header_map["Product_Name_Optional"]] if "Product_Name_Optional" in header_map else ""
            if not product_id and not product_name:
                continue
            record = {
                "product_sheet": sheet_name,
                "family_code": row[header_map["Family_Code"]] if "Family_Code" in header_map else "",
                "family_name": row[header_map["Family_Name"]] if "Family_Name" in header_map else "",
                "product_id": product_id or "",
                "product_name": product_name or "",
                "category": row[header_map["Category"]] if "Category" in header_map else "",
                "sub_category": row[header_map["Sub_Category"]] if "Sub_Category" in header_map else "",
                "type_of_content": row[header_map["Type_of_Content"]] if "Type_of_Content" in header_map else "",
                "silo_key": row[header_map["Silo_Key"]] if "Silo_Key" in header_map else "",
            }
            if record["product_id"]:
                by_id[normalize(record["product_id"])] = record
            if record["product_name"]:
                by_name[normalize(record["product_name"])] = record
            break

    return {"by_id": by_id, "by_name": by_name, "family_sheets": family_sheets}


def resolve_family_sheet_from_code(registry: dict, workbook_index: dict, family_code: str, category_blob: str) -> str | None:
    normalized_code = normalize(family_code)
    for family_sheet, mapping in registry["family_matching"]["keyword_to_family_map"].items():
        if normalized_code in {normalize(family_sheet), normalize(mapping.get("family_code"))}:
            return family_sheet
    for family_sheet in workbook_index["family_sheets"]:
        if normalized_code == normalize(family_sheet):
            return family_sheet
    return find_family_match(registry, workbook_index, category_blob)


def find_family_match(registry: dict, workbook_index: dict, text_blob: str) -> str | None:
    normalized_text = normalize(text_blob)
    for family_sheet, mapping in registry["family_matching"]["keyword_to_family_map"].items():
        if family_sheet not in workbook_index["family_sheets"]:
            continue
        for keyword in mapping.get("keywords", []):
            if normalize(keyword) and normalize(keyword) in normalized_text:
                return family_sheet
    return None


def collect_missing_fields(payload: dict, required_fields: list[str]) -> list[str]:
    missing: list[str] = []
    for field in required_fields:
        if not normalize(payload.get(field)):
            missing.append(field)
    return missing


def detect_review_only(registry: dict, payload: dict) -> list[str]:
    reasons: list[str] = []
    text_parts = [
        payload.get("product_name"),
        payload.get("category"),
        payload.get("sub_category"),
        payload.get("main_problem_solved"),
        payload.get("main_benefit"),
        payload.get("claim_restrictions"),
        payload.get("usage_scenario"),
    ]
    text_blob = normalize(" ".join(part for part in text_parts if part))
    compliance_class = normalize(payload.get("compliance_class")).upper()

    if compliance_class in {value.upper() for value in registry["risk_detection"]["review_only_triggers"]["compliance_classes"]}:
        reasons.append(f"compliance_class={payload.get('compliance_class')}")

    for category in registry["risk_detection"]["high_risk_categories"]:
        if normalize(category) and normalize(category) in text_blob:
            reasons.append(f"high_risk_category={category}")

    for trigger in registry["risk_detection"]["review_only_triggers"]["keywords"]:
        if normalize(trigger) and normalize(trigger) in text_blob:
            reasons.append(f"trigger={trigger}")

    return list(dict.fromkeys(reasons))


def build_registered_result(
    registry: dict,
    product_record: dict | None,
    workbook_record: dict | None,
    workbook_index: dict,
    payload: dict,
    source_tier: str,
) -> dict:
    merged = {}
    merged.update(workbook_record or {})
    merged.update(product_record or {})
    category_blob = " ".join(
        filter(None, [merged.get("category"), merged.get("sub_category"), payload.get("category"), payload.get("sub_category")])
    )
    family_sheet = resolve_family_sheet_from_code(
        registry,
        workbook_index,
        str((workbook_record or {}).get("family_code", "")),
        category_blob,
    )
    copy_authority_parts = []
    if product_record:
        copy_authority_parts.append(product_record["source_file"])
    if workbook_record:
        copy_authority_parts.append(workbook_record["product_sheet"])
    return {
        "route_mode": "REGISTERED_PRODUCT",
        "output_status": registry["output_status"]["REGISTERED"],
        "source_tier": source_tier,
        "source_status": "APPROVED_LIBRARY",
        "ready_for_generation": True,
        "product_id": merged.get("product_id") or payload.get("product_id") or "",
        "product_name": merged.get("product_name") or payload.get("product_name") or "",
        "family_sheet": family_sheet,
        "family_code": (workbook_record or {}).get("family_code", ""),
        "copy_authority": " + ".join(copy_authority_parts),
        "requires_review": False,
        "minimum_intake_missing": [],
        "notes": ["Use product-specific truth and approved product-specific copy packs when present."],
    }


def route_product(payload: dict, registry: dict | None = None, products: dict | None = None, workbook_index: dict | None = None) -> dict:
    registry = registry or load_registry()
    products = products or load_products()
    workbook_index = workbook_index or load_workbook_index()

    product_id_key = normalize(payload.get("product_id"))
    product_name_key = normalize(payload.get("product_name"))

    product_record = products["by_id"].get(product_id_key) or products["by_name"].get(product_name_key)
    workbook_record = workbook_index["by_id"].get(product_id_key) or workbook_index["by_name"].get(product_name_key)

    if product_record or workbook_record:
        source_tier = "product_specific_registry" if product_record else "product_specific_workbook_copy_pack"
        return build_registered_result(registry, product_record, workbook_record, workbook_index, payload, source_tier)

    review_reasons = detect_review_only(registry, payload)
    minimum_fields = registry["minimum_on_the_fly_intake"]["required_fields"]
    missing_fields = collect_missing_fields(payload, minimum_fields)

    if review_reasons:
        return {
            "route_mode": "REVIEW_ONLY_PRODUCT",
            "output_status": registry["output_status"]["REVIEW_ONLY"],
            "source_tier": "review_only_fallback",
            "source_status": "REVIEW_REQUIRED",
            "ready_for_generation": False,
            "product_id": payload.get("product_id", ""),
            "product_name": payload.get("product_name", ""),
            "family_sheet": None,
            "family_code": None,
            "copy_authority": "review_only_fallback",
            "requires_review": True,
            "minimum_intake_missing": missing_fields,
            "notes": review_reasons + ["High-risk lane is fail-closed and must not auto-generate aggressive copy."],
        }

    family_blob = " ".join(
        filter(
            None,
            [
                payload.get("product_name"),
                payload.get("category"),
                payload.get("sub_category"),
                payload.get("product_format"),
            ],
        )
    )
    family_sheet = find_family_match(registry, workbook_index, family_blob)
    if family_sheet:
        mapping = registry["family_matching"]["keyword_to_family_map"].get(family_sheet, {})
        family_missing = [field for field in ("product_name", "visual_reference_status") if not normalize(payload.get(field))]
        return {
            "route_mode": "FAMILY_MATCHED_PRODUCT",
            "output_status": registry["output_status"]["FAMILY_MATCHED"],
            "source_tier": "product_family_workbook_copy_pack",
            "source_status": "FAMILY_LIBRARY",
            "ready_for_generation": not family_missing,
            "product_id": payload.get("product_id", ""),
            "product_name": payload.get("product_name", ""),
            "family_sheet": family_sheet,
            "family_code": mapping.get("family_code", family_sheet),
            "copy_authority": family_sheet,
            "requires_review": False,
            "minimum_intake_missing": family_missing,
            "notes": [
                "Use family-level copy packs only.",
                "Do not pretend product-specific authority exists for this product.",
            ],
        }

    return {
        "route_mode": "ON_THE_FLY_PRODUCT",
        "output_status": registry["output_status"]["ON_THE_FLY"],
        "source_tier": "on_the_fly_intake",
        "source_status": "AD_HOC_GENERATED",
        "ready_for_generation": not missing_fields,
        "product_id": payload.get("product_id", ""),
        "product_name": payload.get("product_name", ""),
        "family_sheet": None,
        "family_code": None,
        "copy_authority": "minimum_on_the_fly_intake",
        "requires_review": False,
        "minimum_intake_missing": missing_fields,
        "notes": [
            "Ad-hoc copy is allowed for this session only.",
            "Do not write on-the-fly output back to the workbook automatically.",
        ],
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Route BOSMAX product copy requests without generating copy.")
    parser.add_argument("--product-id", default="")
    parser.add_argument("--product-name", default="")
    parser.add_argument("--category", default="")
    parser.add_argument("--sub-category", default="")
    parser.add_argument("--target-user", default="")
    parser.add_argument("--main-problem-solved", default="")
    parser.add_argument("--main-benefit", default="")
    parser.add_argument("--product-format", default="")
    parser.add_argument("--product-size-or-scale", default="")
    parser.add_argument("--platform", default="")
    parser.add_argument("--language", default="")
    parser.add_argument("--compliance-class", default="")
    parser.add_argument("--visual-reference-status", default="")
    parser.add_argument("--price", default="")
    parser.add_argument("--offer", default="")
    parser.add_argument("--differentiator", default="")
    parser.add_argument("--competitor-context", default="")
    parser.add_argument("--product-image", default="")
    parser.add_argument("--target-persona", default="")
    parser.add_argument("--claim-restrictions", default="")
    parser.add_argument("--usage-scenario", default="")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    payload = {
        "product_id": args.product_id,
        "product_name": args.product_name,
        "category": args.category,
        "sub_category": args.sub_category,
        "target_user": args.target_user,
        "main_problem_solved": args.main_problem_solved,
        "main_benefit": args.main_benefit,
        "product_format": args.product_format,
        "product_size_or_scale": args.product_size_or_scale,
        "platform": args.platform,
        "language": args.language,
        "compliance_class": args.compliance_class,
        "visual_reference_status": args.visual_reference_status,
        "price": args.price,
        "offer": args.offer,
        "differentiator": args.differentiator,
        "competitor_context": args.competitor_context,
        "product_image": args.product_image,
        "target_persona": args.target_persona,
        "claim_restrictions": args.claim_restrictions,
        "usage_scenario": args.usage_scenario,
    }
    result = route_product(payload)
    print(json.dumps(result, indent=2, ensure_ascii=True))


if __name__ == "__main__":
    main()
