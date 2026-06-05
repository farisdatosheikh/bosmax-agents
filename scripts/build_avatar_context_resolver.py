from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
AVATARS_DIR = ROOT / "avatars"
WARDROBE_RULES_PATH = ROOT / "wardrobes" / "WARDROBE_RULES.yaml"
OUTPUT_WORKBOOK_PATH = ROOT / "BOSMAX_AVATAR_CONTEXT_RESOLVER_v1.xlsx"
OUTPUT_REGISTRY_PATH = ROOT / "registries" / "avatar_context_rotation.yaml"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

CONTEXT_HEADERS = [
    "Avatar_Context_ID",
    "Display_Name",
    "Persona_ID",
    "Base_Archetype",
    "Gender",
    "Age_Range",
    "Silo_Allowed",
    "Product_Family_Allowed",
    "Scene_Context_ID",
    "Mannequin_ID",
    "Wardrobe_ID",
    "Headwear_Style",
    "Camera_Style_Allowed",
    "Prompt_Fragment_Source",
    "Status",
    "Runtime_Allowed",
    "Rotation_Weight",
    "Safe_Usage_Notes",
    "Internal_Notes",
]
PERSONA_HEADERS = [
    "Persona_ID",
    "Persona_Label",
    "Base_Archetype",
    "Assigned_Products",
    "Compatible_Silos",
    "Wardrobe_Count",
    "Prompt_Fragment_Source",
]
MANNEQUIN_HEADERS = [
    "Mannequin_ID",
    "Mannequin_Label",
    "Pose_Summary",
    "Compatible_Physics_Classes",
    "Camera_Style_Allowed",
    "Scene_Context_IDs",
]
SCENE_HEADERS = [
    "Scene_Context_ID",
    "Scene_Label",
    "Base_Scene_Key",
    "Environment_Summary",
    "Silo_Allowed",
    "Camera_Style_Allowed",
    "Safe_Usage_Notes",
]
POOL_HEADERS = [
    "Pool_ID",
    "Display_Name",
    "Product_ID",
    "Product_Family",
    "Silo",
    "Allowed_Avatar_Context_IDs",
    "Rotation_Mode",
    "No_Repeat_Window",
    "Minimum_Approved_Count",
    "Status",
    "Runtime_Allowed",
    "Safe_Usage_Notes",
]
COMMAND_CENTRE_AVATAR_HEADERS = [
    "Avatar_Context_ID",
    "Display_Name",
    "Persona_Label",
    "Gender",
    "Age_Range",
    "Silo_Allowed",
    "Product_Family_Allowed",
    "Scene_Label",
    "Mannequin_Label",
    "Camera_Style_Allowed",
    "Status",
    "Safe_Usage_Notes",
]
COMMAND_CENTRE_POOL_HEADERS = [
    "Pool_ID",
    "Display_Name",
    "Product_ID",
    "Product_Family",
    "Silo",
    "Rotation_Mode",
    "No_Repeat_Window",
    "Minimum_Approved_Count",
    "Status",
    "Safe_Usage_Notes",
]
COMMAND_CENTRE_AVATAR_SHEET_ALIAS = "CC_AVATAR_ID_VIEW"
COMMAND_CENTRE_POOL_SHEET_ALIAS = "CC_AVATAR_POOL_VIEW"
VIEW_ALIAS_HEADERS = ["Public_View_ID", "Workbook_Sheet_Alias", "Description"]
NOTION_HEADERS = [
    "Avatar_Context_ID",
    "Display_Name",
    "Persona_Label",
    "Gender",
    "Age_Range",
    "Silo_Allowed",
    "Product_Family_Allowed",
    "Scene_Label",
    "Mannequin_Label",
    "Camera_Style_Allowed",
    "Status",
    "Safe_Usage_Notes",
]
VALIDATION_RULE_HEADERS = ["Rule_ID", "Rule_Description"]
CHANGELOG_HEADERS = ["Timestamp_UTC", "Change_Note"]

VALIDATION_RULES = [
    ("AVATAR_CONTEXT_ID_NOT_FOUND", "Fail closed if Avatar_Context_ID cannot resolve from the runtime registry."),
    ("AVATAR_POOL_ID_NOT_FOUND", "Fail closed if Avatar_Pool_ID cannot resolve from the runtime registry."),
    ("AVATAR_STATUS_INVALID", "Fail closed if Runtime_Allowed is true but Status is not APPROVED, LOCKED, or SEED_READY."),
    ("AVATAR_SILO_MISMATCH", "Fail closed if template silo is not included in Silo_Allowed."),
    ("AVATAR_PRODUCT_FAMILY_MISMATCH", "Fail closed if product family is not included in Product_Family_Allowed."),
    ("AVATAR_CAMERA_STYLE_MISMATCH", "Fail closed if camera style is not included in Camera_Style_Allowed."),
    ("AVATAR_PHYSICS_MISMATCH", "Fail closed if mannequin pose is incompatible with requested Physics class."),
    ("AVATAR_POOL_REPEAT_WINDOW", "Fail closed if AUTO_ROTATE cannot satisfy the configured no-repeat window."),
    ("UNSAFE_MANUAL_OVERRIDE", "Fail closed if manual override is present without Needs Compliance Review."),
    ("COMMAND_CENTRE_AVATAR_VIEW_ONLY", "Command Centre avatar and pool views must remain frontend-safe selector surfaces only."),
    ("COMMAND_CENTRE_PRODUCT_WORKFLOW_SUPPORT", "Registered Command Centre workflows must expose separate STEALTH and DIRECT avatar support without leaking prompt internals."),
]

VIEW_ALIASES = [
    {
        "public_view_id": "NOTION_COMMAND_CENTRE_AVATAR_ID_VIEW",
        "workbook_sheet_alias": COMMAND_CENTRE_AVATAR_SHEET_ALIAS,
        "description": "Generic beginner avatar selector view across approved runtime packs.",
    },
    {
        "public_view_id": "NOTION_COMMAND_CENTRE_AVATAR_POOL_VIEW",
        "workbook_sheet_alias": COMMAND_CENTRE_POOL_SHEET_ALIAS,
        "description": "Generic beginner avatar pool selector view across approved runtime pools.",
    },
]

SCENE_LIBRARY = {
    "home_indoor": {
        "scene_context_id": "CTX_HOME_STEALTH_RESET_001",
        "scene_label": "Quiet Home Reset Corner",
        "environment_summary": "Warm indoor private corner with low-noise everyday props and product-neutral framing.",
        "safe_usage_notes": "Private indoor stealth lane. Keep framing conversational and product-neutral.",
        "silo_allowed": ["STEALTH"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
    "outdoor_street": {
        "scene_context_id": "CTX_STREET_MORNING_RESET_001",
        "scene_label": "Morning Street Confidence Walk",
        "environment_summary": "Soft daylight outdoor walkway with authentic motion and subtle background depth.",
        "safe_usage_notes": "Outdoor stealth lane. Keep gestures natural and avoid fitness-bro exaggeration.",
        "silo_allowed": ["STEALTH"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
    "office_interior": {
        "scene_context_id": "CTX_OFFICE_CONFIDENCE_CORNER_001",
        "scene_label": "Office Confidence Corner",
        "environment_summary": "Smart indoor office nook with grounded professional texture and light desk cues.",
        "safe_usage_notes": "Professional stealth lane. Keep the setting calm and non-clinical.",
        "silo_allowed": ["STEALTH"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
    "mwcb_home_relief": {
        "scene_context_id": "CTX_HOME_TRAD_REMEDY_001",
        "scene_label": "Household Relief Shelf Reset",
        "environment_summary": "Warm rumah setting with small household storage cues, everyday reachability, and practical family utility framing.",
        "safe_usage_notes": "DIRECT traditional-remedy lane. Keep the bottle within normal household use context and avoid clinical implication.",
        "silo_allowed": ["DIRECT"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
    "mwcb_bag_ready": {
        "scene_context_id": "CTX_TRAVEL_TRAD_REMEDY_001",
        "scene_label": "On-The-Go Remedy Grab",
        "environment_summary": "Natural grab-from-bag or dashboard-side readiness moment showing compact everyday carry practicality.",
        "safe_usage_notes": "DIRECT traditional-remedy lane. Keep usage practical, pocketable, and non-medical.",
        "silo_allowed": ["DIRECT"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
}

MANNEQUIN_LIBRARY = {
    "home_indoor": {
        "mannequin_id": "MAN_STEALTH_CLOSE_HOLD_001",
        "mannequin_label": "Close Hold Reset",
        "pose_summary": "Standing or seated medium-close talk with one natural hand accent and no static mannequin freeze.",
        "compatible_physics_classes": ["CLASS_A"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
    "outdoor_street": {
        "mannequin_id": "MAN_STEALTH_WALK_AND_TALK_001",
        "mannequin_label": "Walk And Talk",
        "pose_summary": "Slow forward or side-walk delivery with light arm swing and relaxed shoulder posture.",
        "compatible_physics_classes": ["CLASS_A"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
    "office_interior": {
        "mannequin_id": "MAN_STEALTH_DESK_REVEAL_001",
        "mannequin_label": "Desk Reveal Confidence",
        "pose_summary": "Controlled standing or leaning office posture with measured hand reveal and grounded eye line.",
        "compatible_physics_classes": ["CLASS_A"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
    "mwcb_home_relief": {
        "mannequin_id": "MAN_DIRECT_BOTTLE_PICKUP_001",
        "mannequin_label": "Bottle Pickup Utility",
        "pose_summary": "Natural household reach, bottle pickup, or side-table explanation gesture with grounded family-man practicality.",
        "compatible_physics_classes": ["CLASS_A"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
    "mwcb_bag_ready": {
        "mannequin_id": "MAN_DIRECT_BAG_REACH_001",
        "mannequin_label": "Bag Reach Practicality",
        "pose_summary": "Simple bag, drawer, or car-pocket reach motion showing compact everyday-carry readiness without exaggerated acting.",
        "compatible_physics_classes": ["CLASS_A"],
        "camera_style_allowed": ["UGC_IPHONE_RAW"],
    },
}


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def load_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8")) or {}


def load_named_personas() -> list[dict[str, Any]]:
    personas = []
    for path in sorted(AVATARS_DIR.glob("*.yaml")):
        if path.name == "_SCHEMA.yaml":
            continue
        data = load_yaml(path)
        if normalize(data.get("persona_type")).upper() != "NAMED_PERSONA":
            continue
        data["_path"] = path
        personas.append(data)
    priority = {"RIZAL": 0, "AZMAN": 1}
    personas.sort(key=lambda item: (priority.get(normalize(item.get("persona_id")).upper(), 99), normalize(item.get("persona_id"))))
    return personas


def derive_headwear_style(outfit: str) -> str:
    lowered = outfit.lower()
    if "songkok" in lowered:
        return "SONGKOK"
    if "kopiah" in lowered:
        return "KOPIAH"
    if "cap" in lowered:
        return "CAP"
    return "NONE"


def build_context_pack(
    *,
    context_id: str,
    persona: dict[str, Any],
    wardrobe_id: str,
    scene_key: str,
    product_family: str,
    usage_note_prefix: str,
) -> dict[str, Any] | None:
    wardrobe_catalogue = persona.get("wardrobe_catalogue", []) or []
    wardrobe = next((item for item in wardrobe_catalogue if normalize(item.get("wardrobe_id")) == wardrobe_id), None)
    if not wardrobe:
        return None

    scene_entry = SCENE_LIBRARY[scene_key]
    mannequin_entry = MANNEQUIN_LIBRARY[scene_key]
    outfit = normalize(wardrobe.get("outfit"))
    persona_id = normalize(persona.get("persona_id"))
    persona_name = normalize(persona.get("persona_name")) or persona_id
    return {
        "avatar_context_id": context_id,
        "display_name": f"{persona_id} | {wardrobe_id} | {scene_entry['scene_label']}",
        "persona_id": persona_id,
        "persona_label": persona_name,
        "base_archetype": normalize(persona.get("base_archetype")),
        "gender": normalize(persona.get("gender")) or "Male",
        "age_range": normalize(persona.get("age_range")),
        "silo_allowed": scene_entry["silo_allowed"],
        "product_family_allowed": [product_family],
        "scene_context_id": scene_entry["scene_context_id"],
        "scene_label": scene_entry["scene_label"],
        "mannequin_id": mannequin_entry["mannequin_id"],
        "mannequin_label": mannequin_entry["mannequin_label"],
        "wardrobe_id": wardrobe_id,
        "headwear_style": derive_headwear_style(outfit),
        "camera_style_allowed": scene_entry["camera_style_allowed"],
        "compatible_physics_classes": mannequin_entry["compatible_physics_classes"],
        "prompt_fragment_source": f"avatars/{persona['_path'].name}::prompt_fragment",
        "status": "APPROVED",
        "runtime_allowed": True,
        "rotation_weight": 1,
        "safe_usage_notes": (
            f"{usage_note_prefix} Persona {persona_id} with {wardrobe_id}. "
            "Manual override requires Needs Compliance Review."
        ),
        "internal_notes": (
            f"Wardrobe sourced from {persona['_path'].name}::{wardrobe_id}; "
            f"scene keyed from {scene_key}; mannequin pose resolved in avatar context builder."
        ),
    }


def build_context_packs() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    personas = load_named_personas()
    context_packs: list[dict[str, Any]] = []
    persona_index: list[dict[str, Any]] = []
    persona_lookup = {normalize(persona.get("persona_id")): persona for persona in personas}

    for persona in personas:
        persona_id = normalize(persona.get("persona_id"))
        assigned_products = [normalize(item) for item in persona.get("assigned_products", [])]
        compatible_silos = [normalize(item).upper() for item in persona.get("compatible_silos", [])]
        wardrobe_catalogue = persona.get("wardrobe_catalogue", []) or []

        persona_index.append(
            {
                "persona_id": persona_id,
                "persona_label": normalize(persona.get("persona_name")) or persona_id,
                "base_archetype": normalize(persona.get("base_archetype")),
                "assigned_products": assigned_products,
                "compatible_silos": compatible_silos,
                "wardrobe_count": len(wardrobe_catalogue),
                "prompt_fragment_source": f"avatars/{persona['_path'].name}::prompt_fragment",
            }
        )

    for persona in personas:
        persona_id = normalize(persona.get("persona_id"))
        assigned_products = [normalize(item) for item in persona.get("assigned_products", [])]
        compatible_silos = [normalize(item).upper() for item in persona.get("compatible_silos", [])]
        if "BOSMAX_SERUM" not in assigned_products or "STEALTH" not in compatible_silos:
            continue

        for wardrobe in persona.get("wardrobe_catalogue", []) or []:
            scene_key = normalize(wardrobe.get("scene_context"))
            wardrobe_id = normalize(wardrobe.get("wardrobe_id"))
            if scene_key not in {"home_indoor", "outdoor_street", "office_interior"}:
                continue
            if scene_key == "home_indoor" and "RAYA" not in wardrobe_id.upper() and "CASUAL" not in wardrobe_id.upper():
                continue
            if scene_key == "outdoor_street" and "ACTIVE" not in wardrobe_id.upper():
                continue
            if scene_key == "office_interior" and "SMART" not in wardrobe_id.upper():
                continue

            pack = build_context_pack(
                context_id=f"BOSMAX_AVP_{len(context_packs) + 1:04d}",
                persona=persona,
                wardrobe_id=wardrobe_id,
                scene_key=scene_key,
                product_family="FAMILY_MALE_EXT_SENSITIVE_OIL",
                usage_note_prefix="STEALTH resolver pack for BOSMAX Serum family.",
            )
            if pack:
                context_packs.append(pack)

    direct_specs = [
        ("MWCB_DIRECT_AVP_0001", "RIZAL", "RIZAL_CASUAL_01", "mwcb_home_relief"),
        ("MWCB_DIRECT_AVP_0002", "AZMAN", "AZMAN_CASUAL_01", "mwcb_bag_ready"),
    ]
    for context_id, persona_id, wardrobe_id, scene_key in direct_specs:
        persona = persona_lookup.get(persona_id)
        if not persona:
            continue
        compatible_silos = [normalize(item).upper() for item in persona.get("compatible_silos", [])]
        assigned_products = [normalize(item) for item in persona.get("assigned_products", [])]
        if "DIRECT" not in compatible_silos or "CAP_BURUNG_MINYAK" not in assigned_products:
            continue
        pack = build_context_pack(
            context_id=context_id,
            persona=persona,
            wardrobe_id=wardrobe_id,
            scene_key=scene_key,
            product_family="FAMILY_TRADITIONAL_REMEDY_OIL",
            usage_note_prefix="DIRECT resolver pack for Minyak Warisan Cap Burung family.",
        )
        if pack:
            context_packs.append(pack)

    context_packs.sort(key=lambda item: item["avatar_context_id"])

    scene_rows = []
    for scene_key, scene_entry in SCENE_LIBRARY.items():
        scene_rows.append(
            {
                "scene_context_id": scene_entry["scene_context_id"],
                "scene_label": scene_entry["scene_label"],
                "base_scene_key": scene_key,
                "environment_summary": scene_entry["environment_summary"],
                "silo_allowed": scene_entry["silo_allowed"],
                "camera_style_allowed": scene_entry["camera_style_allowed"],
                "safe_usage_notes": scene_entry["safe_usage_notes"],
            }
        )

    mannequin_rows = []
    for scene_key, mannequin_entry in MANNEQUIN_LIBRARY.items():
        mannequin_rows.append(
            {
                "mannequin_id": mannequin_entry["mannequin_id"],
                "mannequin_label": mannequin_entry["mannequin_label"],
                "pose_summary": mannequin_entry["pose_summary"],
                "compatible_physics_classes": mannequin_entry["compatible_physics_classes"],
                "camera_style_allowed": mannequin_entry["camera_style_allowed"],
                "scene_context_ids": [SCENE_LIBRARY[scene_key]["scene_context_id"]],
            }
        )

    stealth_context_ids = [
        row["avatar_context_id"]
        for row in context_packs
        if "STEALTH" in row["silo_allowed"]
    ]
    direct_context_ids = [
        row["avatar_context_id"]
        for row in context_packs
        if "DIRECT" in row["silo_allowed"]
    ]

    pools = [{
        "pool_id": "BOSMAX_MALE_STEALTH_POOL_001",
        "display_name": "BOSMAX Male Stealth Pool | BOSMAX Serum Family",
        "product_id": "BOSMAX_SERUM",
        "product_family": "FAMILY_MALE_EXT_SENSITIVE_OIL",
        "silo": "STEALTH",
        "allowed_avatar_context_ids": stealth_context_ids,
        "rotation_mode": "ROUND_ROBIN_NO_REPEAT",
        "no_repeat_window": max(1, len(stealth_context_ids) - 1),
        "minimum_approved_count": len(stealth_context_ids),
        "status": "APPROVED",
        "runtime_allowed": True,
        "safe_usage_notes": (
            "Repo-owned AUTO_ROTATE pool for BOSMAX Serum stealth batch runs. "
            "Notion may display IDs only; manual override requires Needs Compliance Review."
        ),
    }]
    pools.append(
        {
            "pool_id": "MWCB_TRAD_REMEDY_POOL_001",
            "display_name": "MWCB Direct Remedy Pool | Minyak Warisan Cap Burung",
            "product_id": "CAP_BURUNG_MINYAK",
            "product_family": "FAMILY_TRADITIONAL_REMEDY_OIL",
            "silo": "DIRECT",
            "allowed_avatar_context_ids": direct_context_ids,
            "rotation_mode": "ROUND_ROBIN_NO_REPEAT",
            "no_repeat_window": 1,
            "minimum_approved_count": len(direct_context_ids),
            "status": "APPROVED",
            "runtime_allowed": True,
            "safe_usage_notes": (
                "Repo-owned AUTO_ROTATE pool for Minyak Warisan Cap Burung DIRECT batch runs. "
                "Notion may display IDs only; manual override requires Needs Compliance Review."
            ),
        }
    )

    return context_packs, persona_index, mannequin_rows, scene_rows, pools


def build_registry(
    context_packs: list[dict[str, Any]],
    persona_index: list[dict[str, Any]],
    mannequin_rows: list[dict[str, Any]],
    scene_rows: list[dict[str, Any]],
    pools: list[dict[str, Any]],
) -> dict[str, Any]:
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    notion_rows = [
        {
            "avatar_context_id": row["avatar_context_id"],
            "display_name": row["display_name"],
            "persona_label": row["persona_label"],
            "gender": row["gender"],
            "age_range": row["age_range"],
            "silo_allowed": row["silo_allowed"],
            "product_family_allowed": row["product_family_allowed"],
            "scene_label": row["scene_label"],
            "mannequin_label": row["mannequin_label"],
            "camera_style_allowed": row["camera_style_allowed"],
            "status": row["status"],
            "safe_usage_notes": row["safe_usage_notes"],
        }
        for row in context_packs
    ]
    command_centre_avatar_rows = [
        {
            **row,
            "safe_usage_notes": (
                f"STAGING_ONLY — {row['safe_usage_notes']}" if row["status"] == "SEED_READY" else row["safe_usage_notes"]
            ),
        }
        for row in notion_rows
    ]
    command_centre_pool_rows = [
        {
            "pool_id": row["pool_id"],
            "display_name": row["display_name"],
            "product_id": row["product_id"],
            "product_family": row["product_family"],
            "silo": row["silo"],
            "rotation_mode": row["rotation_mode"],
            "no_repeat_window": row["no_repeat_window"],
            "minimum_approved_count": row["minimum_approved_count"],
            "status": row["status"],
            "safe_usage_notes": (
                f"STAGING_ONLY — {row['safe_usage_notes']}" if row["status"] == "SEED_READY" else row["safe_usage_notes"]
            ),
        }
        for row in pools
    ]

    return {
        "schema_version": "1.0",
        "registry_id": "BOSMAX_AVATAR_CONTEXT_ROTATION_V1",
        "generated_at_utc": generated_at,
        "source_authority": {
            "avatar_schema": "avatars/_SCHEMA.yaml",
            "avatar_registry": "avatars/*.yaml",
            "wardrobe_rules": "wardrobes/WARDROBE_RULES.yaml",
            "notes": "Avatar resolver extends runtime context through separate registry outputs only.",
        },
        "runtime_contract": {
            "allowed_runtime_statuses": ["APPROVED", "LOCKED", "SEED_READY"],
            "manual_override_requires_review_status": "Needs Compliance Review",
            "supported_avatar_modes": ["AUTO_RESOLVE", "AUTO_ROTATE"],
            "supported_rotation_rules": ["ROUND_ROBIN_NO_REPEAT"],
            "default_notion_flow": "COMMAND_CENTRE_PLUG_AND_PLAY",
            "legacy_manual_flow": "LEGACY_EXPERT_MODE | MANUAL_OVERRIDE_REVIEW_ONLY",
            "notion_safe_fields": [header.lower() for header in NOTION_HEADERS],
            "command_centre_avatar_fields": [header.lower() for header in COMMAND_CENTRE_AVATAR_HEADERS],
            "command_centre_pool_fields": [header.lower() for header in COMMAND_CENTRE_POOL_HEADERS],
            "forbidden_notion_fields": [
                "prompt_fragment_source",
                "internal_notes",
                "compatible_physics_classes",
            ],
        },
        "workbook_sheet_aliases": VIEW_ALIASES,
        "validation_rules": [
            {"rule_id": rule_id, "description": description}
            for rule_id, description in VALIDATION_RULES
        ],
        "avatar_context_packs": context_packs,
        "persona_index": persona_index,
        "mannequin_pose_library": mannequin_rows,
        "scene_context_library": scene_rows,
        "rotation_pools": pools,
        "notion_export_view": notion_rows,
        "notion_command_centre_avatar_id_view": command_centre_avatar_rows,
        "notion_command_centre_avatar_pool_view": command_centre_pool_rows,
        "legacy_expert_mode": {
            "status": "ACTIVE",
            "label": "LEGACY_EXPERT_MODE",
            "manual_override_posture": "MANUAL_OVERRIDE_REVIEW_ONLY",
            "required_review_status": "Needs Compliance Review",
            "notes": "Manual Avatar/Mannequin/Wardrobe/Scene edits remain expert-only and review-gated.",
        },
        "command_centre_product_workflows": {
            "BOSMAX_SERUM_STEALTH_REGISTERED_PRODUCT": {
                "avatar_context_id": "BOSMAX_AVP_0001",
                "avatar_pool_id": "BOSMAX_MALE_STEALTH_POOL_001",
                "avatar_mode_single": "AUTO_RESOLVE",
                "avatar_mode_batch": "AUTO_ROTATE",
            },
            "MINYAK_WARISAN_CAP_BURUNG_DIRECT_REGISTERED_PRODUCT": {
                "avatar_context_id": "MWCB_DIRECT_AVP_0001",
                "avatar_pool_id": "MWCB_TRAD_REMEDY_POOL_001",
                "avatar_mode_single": "AUTO_RESOLVE",
                "avatar_mode_batch": "AUTO_ROTATE",
            },
        },
        "default_command_centre_batch_template": {
            "platform": "TikTok",
            "mode": "B",
            "engine": "VEO_3_1_LITE",
            "duration": "8s",
            "submode_formula": "SAVAGE_HPAS",
            "silo": "STEALTH",
            "avatar_pool_id": "BOSMAX_MALE_STEALTH_POOL_001",
            "avatar_mode": "AUTO_ROTATE",
            "camera_style": "UGC_IPHONE_RAW",
            "language": "Malay",
            "produk": "BOSMAX Serum 5ML",
            "scale_anchor": "EXACTLY lip balm size, fit into fingers naturally",
            "physics_class": "CLASS_A",
            "copywriting_id": "BOSMAX_SERUM_CP_0001",
            "copywriting_mode": "AUTO_RESOLVE",
            "compliance": "STEALTH_METAPHOR_REQUIRED",
            "batch_count": 20,
            "rotation_rule": "ROUND_ROBIN_NO_REPEAT",
        },
        "samples": {
            "sample_avatar_context_id": "BOSMAX_AVP_0001",
            "sample_avatar_pool_id": "BOSMAX_MALE_STEALTH_POOL_001",
            "sample_mwcb_avatar_context_id": "MWCB_DIRECT_AVP_0001",
            "sample_mwcb_avatar_pool_id": "MWCB_TRAD_REMEDY_POOL_001",
        },
        "counts": {
            "avatar_context_packs": len(context_packs),
            "avatar_pools": len(pools),
            "command_centre_avatar_rows": len(command_centre_avatar_rows),
            "command_centre_pool_rows": len(command_centre_pool_rows),
        },
        "changelog": [
            {
                "timestamp_utc": generated_at,
                "change_note": "Generated avatar context resolver and rotation pools for BOSMAX STEALTH and MWCB DIRECT Command Centre workflows with workbook aliases and beginner selector views.",
            }
        ],
    }


def style_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = WRAP_ALIGNMENT

    for column_cells in worksheet.columns:
        values = [normalize(cell.value) for cell in column_cells[:25]]
        width = min(max((len(value) for value in values), default=10) + 2, 42)
        worksheet.column_dimensions[column_cells[0].column_letter].width = max(width, 12)


def populate_sheet(worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    style_sheet(worksheet)


def write_yaml_registry(registry: dict[str, Any]) -> None:
    OUTPUT_REGISTRY_PATH.write_text(
        yaml.safe_dump(registry, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )


def write_workbook(registry: dict[str, Any]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    context_rows = [
        [
            row["avatar_context_id"],
            row["display_name"],
            row["persona_id"],
            row["base_archetype"],
            row["gender"],
            row["age_range"],
            " | ".join(row["silo_allowed"]),
            " | ".join(row["product_family_allowed"]),
            row["scene_context_id"],
            row["mannequin_id"],
            row["wardrobe_id"],
            row["headwear_style"],
            " | ".join(row["camera_style_allowed"]),
            row["prompt_fragment_source"],
            row["status"],
            row["runtime_allowed"],
            row["rotation_weight"],
            row["safe_usage_notes"],
            row["internal_notes"],
        ]
        for row in registry["avatar_context_packs"]
    ]

    persona_rows = [
        [
            row["persona_id"],
            row["persona_label"],
            row["base_archetype"],
            " | ".join(row["assigned_products"]),
            " | ".join(row["compatible_silos"]),
            row["wardrobe_count"],
            row["prompt_fragment_source"],
        ]
        for row in registry["persona_index"]
    ]
    alias_view_rows = [
        [
            row["public_view_id"],
            row["workbook_sheet_alias"],
            row["description"],
        ]
        for row in registry["workbook_sheet_aliases"]
    ]

    mannequin_rows = [
        [
            row["mannequin_id"],
            row["mannequin_label"],
            row["pose_summary"],
            " | ".join(row["compatible_physics_classes"]),
            " | ".join(row["camera_style_allowed"]),
            " | ".join(row["scene_context_ids"]),
        ]
        for row in registry["mannequin_pose_library"]
    ]

    scene_rows = [
        [
            row["scene_context_id"],
            row["scene_label"],
            row["base_scene_key"],
            row["environment_summary"],
            " | ".join(row["silo_allowed"]),
            " | ".join(row["camera_style_allowed"]),
            row["safe_usage_notes"],
        ]
        for row in registry["scene_context_library"]
    ]

    pool_rows = [
        [
            row["pool_id"],
            row["display_name"],
            row["product_id"],
            row["product_family"],
            row["silo"],
            " | ".join(row["allowed_avatar_context_ids"]),
            row["rotation_mode"],
            row["no_repeat_window"],
            row["minimum_approved_count"],
            row["status"],
            row["runtime_allowed"],
            row["safe_usage_notes"],
        ]
        for row in registry["rotation_pools"]
    ]

    notion_rows = [
        [
            row["avatar_context_id"],
            row["display_name"],
            row["persona_label"],
            row["gender"],
            row["age_range"],
            " | ".join(row["silo_allowed"]),
            " | ".join(row["product_family_allowed"]),
            row["scene_label"],
            row["mannequin_label"],
            " | ".join(row["camera_style_allowed"]),
            row["status"],
            row["safe_usage_notes"],
        ]
        for row in registry["notion_export_view"]
    ]
    command_centre_avatar_rows = [
        [
            row["avatar_context_id"],
            row["display_name"],
            row["persona_label"],
            row["gender"],
            row["age_range"],
            " | ".join(row["silo_allowed"]),
            " | ".join(row["product_family_allowed"]),
            row["scene_label"],
            row["mannequin_label"],
            " | ".join(row["camera_style_allowed"]),
            row["status"],
            row["safe_usage_notes"],
        ]
        for row in registry["notion_command_centre_avatar_id_view"]
    ]
    command_centre_pool_rows = [
        [
            row["pool_id"],
            row["display_name"],
            row["product_id"],
            row["product_family"],
            row["silo"],
            row["rotation_mode"],
            row["no_repeat_window"],
            row["minimum_approved_count"],
            row["status"],
            row["safe_usage_notes"],
        ]
        for row in registry["notion_command_centre_avatar_pool_view"]
    ]

    validation_rows = [[rule_id, description] for rule_id, description in VALIDATION_RULES]
    changelog_rows = [
        [entry["timestamp_utc"], entry["change_note"]] for entry in registry["changelog"]
    ]

    populate_sheet(workbook.create_sheet("AVATAR_CONTEXT_PACKS"), CONTEXT_HEADERS, context_rows)
    populate_sheet(workbook.create_sheet("AVATAR_PERSONA_INDEX"), PERSONA_HEADERS, persona_rows)
    populate_sheet(workbook.create_sheet("VIEW_ALIASES"), VIEW_ALIAS_HEADERS, alias_view_rows)
    populate_sheet(workbook.create_sheet("MANNEQUIN_POSE_LIBRARY"), MANNEQUIN_HEADERS, mannequin_rows)
    populate_sheet(workbook.create_sheet("SCENE_CONTEXT_LIBRARY"), SCENE_HEADERS, scene_rows)
    populate_sheet(workbook.create_sheet("ROTATION_POOLS"), POOL_HEADERS, pool_rows)
    populate_sheet(workbook.create_sheet(COMMAND_CENTRE_AVATAR_SHEET_ALIAS), COMMAND_CENTRE_AVATAR_HEADERS, command_centre_avatar_rows)
    populate_sheet(workbook.create_sheet(COMMAND_CENTRE_POOL_SHEET_ALIAS), COMMAND_CENTRE_POOL_HEADERS, command_centre_pool_rows)
    populate_sheet(workbook.create_sheet("NOTION_EXPORT_VIEW"), NOTION_HEADERS, notion_rows)
    populate_sheet(workbook.create_sheet("VALIDATION_RULES"), VALIDATION_RULE_HEADERS, validation_rows)
    populate_sheet(workbook.create_sheet("CHANGELOG"), CHANGELOG_HEADERS, changelog_rows)
    workbook.save(OUTPUT_WORKBOOK_PATH)


def main() -> None:
    context_packs, persona_index, mannequin_rows, scene_rows, pools = build_context_packs()
    registry = build_registry(context_packs, persona_index, mannequin_rows, scene_rows, pools)
    write_yaml_registry(registry)
    write_workbook(registry)
    print(
        f"Generated {OUTPUT_WORKBOOK_PATH.name} and {OUTPUT_REGISTRY_PATH.relative_to(ROOT).as_posix()} "
        f"with {registry['counts']['avatar_context_packs']} avatar context packs and "
        f"{registry['counts']['avatar_pools']} pools."
    )


if __name__ == "__main__":
    main()
