from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from stealth_copy_authority import ROOT, STEALTH_SOURCE_HEADERS, build_family_sheet_rows, build_product_sheet_rows


WORKBOOK_PATH = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx"
BASE_HEADERS = [
    "Row_ID",
    "Family_Code",
    "Family_Name",
    "Product_ID_Optional",
    "Product_Name_Optional",
    "SKU_Optional",
    "Category",
    "Sub_Category",
    "Product_Type",
    "UOM",
    "Product_Size",
    "Product_Scale",
    "Type_of_Content",
    "Silo_Key",
    "Angle_ID",
    "Angle",
    "Hook_ID",
    "Hook",
    "Pain_or_Friction",
    "USP_1",
    "USP_2",
    "USP_3",
    "CTA_ID",
    "CTA",
    "Copywriting_Formula",
    "Authority_Source",
    *STEALTH_SOURCE_HEADERS,
    "Fastmoss_Reference",
    "Status",
    "Notes",
]


def replace_sheet_rows(sheet_name: str, rows: list[dict[str, object]]) -> int:
    workbook = load_workbook(WORKBOOK_PATH)
    ws = workbook[sheet_name]

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for idx, header in enumerate(BASE_HEADERS, start=1):
        ws.cell(row=1, column=idx, value=header)

    for row in rows:
        ws.append([row.get(header, "") for header in BASE_HEADERS])

    workbook.save(WORKBOOK_PATH)
    workbook.close()
    return len(rows)


def main() -> None:
    product_count = replace_sheet_rows("PRODUCT_BOSMAX_SERUM", build_product_sheet_rows())
    family_count = replace_sheet_rows("FAMILY_MALE_EXT_OIL", build_family_sheet_rows())
    print(f"Patched workbook: {WORKBOOK_PATH}")
    print(f"PRODUCT_BOSMAX_SERUM rows: {product_count}")
    print(f"FAMILY_MALE_EXT_OIL rows: {family_count}")


if __name__ == "__main__":
    main()
