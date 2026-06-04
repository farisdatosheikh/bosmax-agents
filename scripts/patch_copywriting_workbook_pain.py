from __future__ import annotations

import re
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path(r"C:\Users\USER\Desktop\Claude Cowork Bosmax Agents")
WORKBOOK_PATH = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx"
README_SHEET = "README_OR_RULES"
RAW_SHEET = "RAW_FASTMOSS_PRODUCTS"
PAIN_HEADER = "Pain_or_Friction"


def normalize(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def sentence(text: str) -> str:
    cleaned = normalize(text).rstrip(".")
    if not cleaned:
        return ""
    cleaned = cleaned[0].upper() + cleaned[1:]
    return f"{cleaned}."


def load_raw_body_map(workbook) -> dict[str, str]:
    ws = workbook[RAW_SHEET]
    headers = [normalize(cell.value) for cell in ws[1]]
    idx = {header: pos for pos, header in enumerate(headers)}
    mapping: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ref = normalize(row[idx["Fastmoss_Reference"]])
        body = normalize(row[idx["Body"]])
        if ref and body:
            mapping[ref] = body
    return mapping


def extract_fastmoss_pain(body: str) -> str:
    text = normalize(body)
    if not text:
        return ""

    match = re.search(r"Ramai orang ada masalah (.+?)\.", text, flags=re.IGNORECASE)
    if match:
        return sentence(match.group(1))

    match = re.search(r"mula dengan masalah ini:\s*(.+?)(?:\.\s|$)", text, flags=re.IGNORECASE)
    if match:
        candidate = normalize(match.group(1))
        if candidate:
            return sentence(f"Bila {candidate[0].lower() + candidate[1:]}, pembeli terus cari pilihan yang lebih meyakinkan")

    first_sentence = normalize(text.split(". ", 1)[0])
    if first_sentence:
        return sentence(first_sentence)

    return ""


def keyword_match(text: str, mapping: list[tuple[tuple[str, ...], str]], default: str) -> str:
    lowered = normalize(text).lower()
    for keywords, result in mapping:
        if any(keyword in lowered for keyword in keywords):
            return result
    return default


def derive_male_stealth_pain(hook: str, angle: str) -> str:
    mapping = [
        (("enjin panas terus mati",), "Bila momentum baru nak hidup tetapi prestasi cepat drop, keyakinan terus jatuh sebelum suasana sempat jadi selesa."),
        (("tersadai awal", "member lain ride sampai pagi"), "Friction paling pedih bila orang lain masih steady tetapi diri sendiri cepat out dan ego terasa dipijak diam-diam."),
        (("short-circuit", "blackout"), "Saat yang sepatutnya tenang jadi janggal bila respon tiba-tiba terputus dan rasa malu datang lebih laju daripada yakin."),
        (("body sado", "hantu jalanan", "barang display"), "Masalah sebenar muncul bila luaran nampak meyakinkan tetapi prestasi tak ikut rentak, lalu maruah terasa kosong."),
        (("piston lekat",), "Bila respon rasa berat dan susah nak hidupkan momentum, keseluruhan moment terus terasa tersekat."),
        (("starter jam",), "Bila permulaan pun susah nak menjadi, lelaki mudah rasa serba tak kena walaupun cuba kekal tenang."),
        (("malam jumaat",), "Friction dia datang bila masa yang patutnya lancar bertukar jadi sesi mencari alasan untuk tutup rasa segan."),
        (("suara enjin dah kasar",), "Bila tanda awal dah rasa tak smooth, keyakinan pun mula goyah sebelum apa-apa bermula."),
        (("gear dah lekat", "gearbox lekat"), "Masalahnya bukan sekadar perlahan, tetapi rasa seret itu buat rentak terus hilang yakin."),
        (("minyak hitam kering", "pressure minyak hitam low"), "Bila rasa dah makin kurang bertenaga dan cepat drop, lelaki mula fikir macam mana nak pulihkan rentak tanpa kecoh."),
        (("piston tak lancar", "crankshaft abang bengkok"), "Apabila respons terasa tak konsisten, sukar nak kekalkan flow yang buat seseorang rasa benar-benar bersedia."),
        (("rnr",), "Bila tengah jauh dari rutin biasa dan perlu standby, rasa cemas datang bila diri sendiri nampak tak bersedia."),
        (("tiang seri", "tiang tumbang"), "Friction ini menyentuh ego kerana bila tenaga terasa goyah, lelaki mula risau imejnya tak lagi kukuh."),
        (("alternator",), "Bila kuasa macam tak sempat penuh dan cepat padam, setiap cubaan terasa makin menekan keyakinan."),
        (("ekzos",), "Bunyi mungkin ada, tetapi bila hasil sebenar tak keluar, rasa segan itu yang paling susah disorok."),
        (("chassis",), "Bila asas rasa rapuh waktu kena tampung beban, lelaki cepat hilang tenang walaupun cuba nampak steady."),
        (("suspension",), "Masalah ini terasa bila daya tahan tak sekuat yang diharapkan dan prestasi nampak cepat mengalah."),
        (("fuel pump",), "Bila aliran tenaga rasa tersumbat, keseluruhan respon jadi lambat dan menjejaskan keyakinan."),
        (("coolant", "radiator"), "Friction berlaku bila cepat panas tetapi sukar kekalkan konsistensi, lalu lelaki mula rasa serba salah."),
        (("turbo",), "Bila boost yang diharapkan tak datang, rasa lemah itu lebih mengganggu daripada apa yang orang nampak."),
        (("drive shaft",), "Apabila gerakan utama pun rasa tak berjalan, keseluruhan situasi mudah bertukar jadi memalukan."),
        (("brake",), "Bila nak teruskan pun susah dan nak berhenti pun susah, rasa hilang kawalan jadi lebih ketara."),
        (("choke valve",), "Apabila start awal pagi pun semput, lelaki mudah rasa dia tak benar-benar ready bila diperlukan."),
        (("timing belt",), "Bila timing rasa lari dan tak stabil, keyakinan pun hilang kerana rentak tak lagi boleh dijangka."),
        (("diff gear",), "Friction ini terasa bila nak pusing arah pun berat, seolah-olah respon tak lagi ringan seperti yang diharapkan."),
    ]
    default = "Bila moment dah mula berjalan tetapi respon cepat merosot, lelaki biasanya hanya mahu pulihkan yakin tanpa perlu buka cerita sensitif."
    return keyword_match(f"{hook} {angle}", mapping, default)


def derive_traditional_remedy_pain(hook: str, angle: str) -> str:
    mapping = [
        (("kepala berat", "kepala berdenyut", "migrain", "pening"), "Bila kepala mula berat atau berdenyut, kerja mudah terganggu kerana fokus terus merosot."),
        (("badan sengal", "lenguh", "otot", "sendi", "lutut", "pinggang"), "Bila badan sengal atau sendi terasa ketat, pergerakan harian jadi lebih lambat dan memenatkan."),
        (("angin", "kembung", "perut"), "Bila perut masuk angin atau rasa kembung, rutin rumah terus jadi tak selesa dan susah fokus."),
        (("luka", "lebam", "bengkak", "terhantuk", "terkena wap panas"), "Friction datang bila insiden kecil berlaku di rumah tetapi bantuan cepat tak berada dekat tangan."),
        (("sakit gigi", "gusi"), "Bila sakit datang pada waktu susah cari bantuan, malam terasa panjang dan rehat terus terganggu."),
        (("bayi",), "Apabila anak kecil tak selesa dan mula meragam, seluruh rumah pun jadi cemas dan penat serentak."),
        (("mabuk", "mual", "muntah", "travel"), "Masalahnya ialah perjalanan yang sepatutnya lancar jadi meletihkan bila rasa loya datang tiba-tiba."),
        (("kencing malam",), "Bila kerap terjaga waktu malam, badan tak sempat pulih dan esok pagi jadi lesu."),
        (("bersalin", "postnatal", "param", "pantang"), "Badan selepas bersalin perlukan rutin yang konsisten kerana lenguh dan rasa berat mudah berulang sepanjang hari."),
        (("nyamuk", "serangga", "gatal"), "Gangguan kecil seperti gatal atau serangga cepat jadi renyah bila tak ada sapuan standby yang senang dicapai."),
        (("nafas", "hidung sumbat", "selesema", "resdung"), "Bila pernafasan rasa tak lapang, tidur dan tumpuan harian pun ikut terganggu."),
        (("tidur", "mengantuk"), "Friction muncul bila badan mahu berehat tetapi rasa tak selesa atau kepala masih berat."),
        (("warisan", "nenek", "keluarga"), "Ramai orang masih cari minyak warisan bila mahu sesuatu yang terasa familiar, bukan produk yang nampak cantik tetapi tak jelas gunanya."),
        (("multi-fungsi", "satu untuk semua", "14 kegunaan"), "Bila setiap masalah kecil perlukan botol berasingan, rumah jadi serabut dan orang cenderung cari satu standby yang lebih praktikal."),
    ]
    default = "Bila rasa tak selesa datang tanpa banyak amaran, orang biasanya mahu sapuan tradisional yang cepat dicapai dan mudah dimasukkan ke dalam rutin."
    return keyword_match(f"{hook} {angle}", mapping, default)


def derive_tudung_bawal_pain(angle: str, hook: str) -> str:
    mapping = [
        (("awning", "shape"), "Masalah biasa datang bila bawal susah nak bentuk awning kemas dan siap pagi terus makan masa lebih lama."),
        (("cooling", "sejuk", "bernafas", "serap peluh", "all weather"), "Cuaca panas dan peluh mudah buat pemakai rasa rimas bila kain tak cukup sejuk atau cepat lembap."),
        (("minimalist", "printed", "floral", "pastel", "matching"), "Ramai orang pening bila tudung nampak cantik sendiri tetapi susah dipadankan dengan outfit harian."),
        (("no-iron", "anti-kedut", "quick dry", "travel"), "Friction besar selalunya datang pada rutin pagi bila kain mudah berkedut dan perlukan penjagaan lebih daripada masa yang ada."),
        (("office", "teacher", "casual", "versatile"), "Bila mahu nampak kemas sepanjang hari, tudung yang cepat lari bentuk atau tak stabil mudah ganggu keyakinan."),
        (("non-slip", "pin", "durable"), "Masalahnya ialah tudung licin atau cepat rosak bila dipin berkali-kali sehingga gaya tak lagi selesa dikekalkan."),
        (("oversized", "coverage", "non-transparent"), "Ramai pembeli cari coverage yang meyakinkan kerana kain terlalu jarang atau terlalu kecil cepat buat mereka rasa serba salah."),
        (("gift", "packaging"), "Hadiah nampak kurang bermakna bila pembungkusan atau persembahan tak cukup kemas walaupun kainnya cantik."),
        (("soft", "comfort"), "Bila kain terasa kasar atau panas di kulit, pemakaian lama terus jadi meletihkan."),
        (("affordable",), "Friction biasa berlaku bila mahu rupa premium tetapi pilihan yang ada cepat nampak murah atau tak tahan pakai."),
    ]
    default = "Ramai pemakai mahukan tudung yang cepat jadi, selesa lama, dan tak menambah kerja kecil pada rutin harian mereka."
    return keyword_match(f"{angle} {hook}", mapping, default)


def derive_unisex_perfume_pain(angle: str, hook: str) -> str:
    mapping = [
        (("citrus", "fresh", "clean", "laundry", "soap", "sea breeze"), "Masalah biasa ialah bau cepat hilang atau jadi flat, lalu penampilan terasa kurang segar sebelum hari habis."),
        (("hotel", "premium", "classic amber", "sandalwood", "warm", "ginger"), "Friction datang bila orang mahu aura matang dan kemas tetapi wangian sedia ada terasa terlalu biasa atau murah."),
        (("active", "gym", "sweat"), "Bila badan aktif dan mudah berpeluh, bau masam cepat ganggu keyakinan jika semburan tak tahan cukup lama."),
        (("travel", "flight", "vacation"), "Perjalanan panjang selalu buat orang perlukan wangian yang mudah dibawa tanpa menambah beban barang."),
        (("office", "workspace", "presentation"), "Wangian yang terlalu kuat atau terlalu lemah sama-sama jadi masalah bila perlu kekal profesional di ruang kerja."),
        (("gift",), "Ramai orang buntu memilih bau hadiah kerana takut nota terlalu maskulin atau terlalu feminin untuk penerima."),
        (("lavender", "calming", "relaxing"), "Frictionnya ialah minda masih terasa berselerak walaupun orang hanya mahu rutin yang lebih tenang dan bersih."),
        (("minimalist bottle", "spray mechanism"), "Kadang masalah bukan pada bau sahaja, tetapi pada botol yang leceh digunakan atau tak kemas diletakkan."),
        (("non-allergic", "hypoallergenic"), "Kulit sensitif buat orang lebih berhati-hati kerana tidak semua wangian terasa selesa dipakai dekat badan atau fabrik."),
    ]
    default = "Ramai pembeli cari bau yang selamat dipakai selalu kerana wangian yang tak tahan atau tak seimbang cepat hilangkan rasa yakin."
    return keyword_match(f"{angle} {hook}", mapping, default)


def derive_women_perfume_pain(angle: str, hook: str) -> str:
    mapping = [
        (("vanilla", "floral", "rose", "jasmine", "fruity", "coconut"), "Masalah biasa bila wangian manis cepat pudar ialah aura feminin yang diharapkan tak sempat bertahan lama."),
        (("date-night", "romantic", "first date", "allure", "misteri"), "Bila acara penting datang, ramai takut wangian yang dipakai tak cukup meninggalkan impresi yang diinginkan."),
        (("fresh", "clean", "daily", "summer", "office"), "Friction datang bila mahu nampak kemas sepanjang hari tetapi bau cepat hilang selepas beberapa jam."),
        (("budget", "luxury", "high-end"), "Ramai pembeli mahukan rasa eksklusif tanpa perlu bayar mahal, tetapi banyak pilihan murah cepat terasa generik."),
        (("compact", "handbag", "travel"), "Bila rutin padat dan banyak bergerak, perfume besar atau susah dibawa cepat jadi tak praktikal untuk touch-up."),
        (("anti-odor", "gym", "sport"), "Peluh dan aktiviti luar cepat menekan keyakinan jika wangian tak cukup kuat menutup bau badan."),
        (("non-staining", "skin friendly"), "Fabrik cerah dan kulit sensitif buat pembeli lebih berhati-hati kerana salah pilih boleh tinggalkan kesan yang tak diingini."),
        (("signature scent",), "Masalahnya ramai orang masih belum jumpa bau identiti yang terasa khas dan mudah diingati orang lain."),
        (("alert", "focus"), "Apabila hari terasa panjang, orang cenderung cari wangian yang bantu rasa lebih segar dan terjaga."),
        (("rainy day",), "Cuaca lembap mudah buat mood jatuh jika bau yang dipakai terasa terlalu dingin atau tak cukup comforting."),
    ]
    default = "Ramai wanita mahu bau yang kekal cantik sepanjang hari kerana wangian yang cepat hilang mudah buat penampilan terasa tak lengkap."
    return keyword_match(f"{angle} {hook}", mapping, default)


def derive_generic_pain(angle: str, silo_key: str) -> str:
    lowered = f"{normalize(angle).lower()} {normalize(silo_key).lower()}"
    if any(key in lowered for key in ("fit", "comfort", "style")):
        return "Masalah biasa ialah produk nampak menarik di rak tetapi bila dipakai tak cukup selesa, tak mudah dipadankan, atau cepat hilang bentuk."
    if any(key in lowered for key in ("taste", "craving", "food", "drink", "snack", "sauce")):
        return "Friction selalunya datang bila orang mahu sesuatu yang sedap dan mudah, tetapi pilihan sedia ada terasa leceh, tak konsisten, atau kurang memuaskan."
    if any(key in lowered for key in ("problem-solution", "benefit", "supp", "care", "cleanser", "mist", "soap")):
        return "Ramai pembeli datang dengan masalah harian yang jelas, jadi mereka lebih tertarik pada solusi yang terus nampak guna dan hasilnya."
    if any(key in lowered for key in ("room transformation", "storage", "home", "curtain", "bedding")):
        return "Masalah utama biasanya ruang nampak kurang kemas, kurang selesa, atau tak cukup tersusun untuk rutin harian."
    return "Ramai pembeli hanya bertindak bila ada friction yang jelas pada rutin, keselesaan, atau persembahan harian mereka."


def derive_pain(row_map: dict[str, str], raw_body_map: dict[str, str]) -> str:
    raw_ref = normalize(row_map.get("Fastmoss_Reference"))
    hook = normalize(row_map.get("Hook"))
    angle = normalize(row_map.get("Angle"))
    silo_key = normalize(row_map.get("Silo_Key"))

    if raw_ref in raw_body_map:
        extracted = extract_fastmoss_pain(raw_body_map[raw_ref])
        if extracted:
            return extracted

    if silo_key == "male_health_stealth_01":
        return derive_male_stealth_pain(hook, angle)
    if silo_key == "traditional_remedy_direct":
        return derive_traditional_remedy_pain(hook, angle)
    if silo_key == "tudung_bawal_direct":
        return derive_tudung_bawal_pain(angle, hook)
    if silo_key == "unisex_perfume_direct":
        return derive_unisex_perfume_pain(angle, hook)
    if silo_key == "women_perfume_direct":
        return derive_women_perfume_pain(angle, hook)

    return derive_generic_pain(angle, silo_key)


def ensure_pain_column(ws) -> int:
    headers = [normalize(cell.value) for cell in ws[1]]
    header_index = {header: idx + 1 for idx, header in enumerate(headers)}
    hook_col = header_index["Hook"]

    if PAIN_HEADER not in header_index:
        ws.insert_cols(hook_col + 1)
        pain_cell = ws.cell(row=1, column=hook_col + 1)
        pain_cell.value = PAIN_HEADER
        pain_cell.fill = PatternFill("solid", fgColor="70AD47" if ws.title.startswith("FAMILY_") else "C00000")
        source = ws.cell(row=1, column=hook_col)
        pain_cell.font = copy(source.font)
        pain_cell.alignment = copy(source.alignment)
        ws.column_dimensions[pain_cell.column_letter].width = 28

    headers = [normalize(cell.value) for cell in ws[1]]
    header_index = {header: idx + 1 for idx, header in enumerate(headers)}
    return header_index[PAIN_HEADER]


def update_readme_sheet(workbook) -> None:
    ws = workbook[README_SHEET]
    for row in ws.iter_rows(min_row=2, values_only=False):
        rule = normalize(row[0].value)
        if rule == "Row Contract":
            row[1].value = "One row = one coherent angle-hook-pain/friction arc with one USP triplet, one CTA, and one chosen formula."
        elif rule == "Antigravity Scope":
            row[1].value = "Antigravity fills only Type_of_Content, Silo_Key, Angle, Hook, Pain_or_Friction, USP_1-3, CTA, Copywriting_Formula, Notes, and optional IDs/Status inside product or family library sheets."

    existing_rules = {normalize(ws.cell(row=i, column=1).value) for i in range(2, ws.max_row + 1)}
    if "Pain_or_Friction" not in existing_rules:
        ws.append(
            [
                "Pain_or_Friction",
                "Official workbook field after Hook. Every filled copy pack row must state the real pain, inconvenience, or friction before product payoff.",
            ]
        )


def patch_workbook() -> None:
    workbook = load_workbook(WORKBOOK_PATH)
    raw_body_map = load_raw_body_map(workbook)
    updated_counts: dict[str, int] = defaultdict(int)

    update_readme_sheet(workbook)

    for sheet_name in workbook.sheetnames:
        if not (sheet_name.startswith("PRODUCT_") or sheet_name.startswith("FAMILY_")):
            continue

        ws = workbook[sheet_name]
        first_row_headers = [normalize(cell.value) for cell in ws[1]]
        if "Hook" not in first_row_headers or "CTA" not in first_row_headers:
            continue
        pain_col = ensure_pain_column(ws)
        headers = [normalize(cell.value) for cell in ws[1]]
        header_index = {header: idx + 1 for idx, header in enumerate(headers)}

        for row_num in range(2, ws.max_row + 1):
            row_map = {
                header: ws.cell(row=row_num, column=col).value
                for header, col in header_index.items()
            }
            populated = any(normalize(row_map.get(field)) for field in ("Hook", "USP_1", "USP_2", "USP_3", "CTA"))
            if not populated:
                continue

            current_pain = normalize(ws.cell(row=row_num, column=pain_col).value)
            if current_pain:
                continue

            pain = derive_pain(row_map, raw_body_map)
            if pain:
                ws.cell(row=row_num, column=pain_col).value = pain
                updated_counts[sheet_name] += 1

    workbook.save(WORKBOOK_PATH)

    print("Patched workbook:", WORKBOOK_PATH)
    for sheet_name in sorted(updated_counts):
        print(f"{sheet_name}: {updated_counts[sheet_name]} pain rows filled")


if __name__ == "__main__":
    patch_workbook()
