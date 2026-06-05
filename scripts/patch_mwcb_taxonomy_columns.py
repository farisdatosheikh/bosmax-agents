"""
One-time patch: add MCA_ID and Compliance_Risk columns to MWCB workbook sheets.

Inserts two new columns (MCA_ID, Compliance_Risk) after Angle_ID in:
  - PRODUCT_MW_CAP_BURUNG
  - FAMILY_TRAD_REMEDY_OIL

Populates existing seed rows (A01-A20) using the angle_to_mca_map defined
in registries/mwcb_copywriting_angle_taxonomy.yaml.

Safe to re-run: if columns already exist they are left untouched.
"""
from __future__ import annotations

import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx"
TAXONOMY_PATH = ROOT / "registries" / "mwcb_copywriting_angle_taxonomy.yaml"

TARGET_SHEETS = ("PRODUCT_MW_CAP_BURUNG", "FAMILY_TRAD_REMEDY_OIL")
NEW_COLUMNS = ("MCA_ID", "Compliance_Risk")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
WHITE_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def load_angle_map(taxonomy_path: Path) -> dict[str, dict[str, str]]:
    with taxonomy_path.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh)
    return data.get("angle_to_mca_map", {})


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def patch_sheet(ws, angle_map: dict[str, dict[str, str]]) -> int:
    headers = [normalize(cell.value) for cell in ws[1]]

    # Skip if already patched
    if "MCA_ID" in headers and "Compliance_Risk" in headers:
        print(f"  [{ws.title}] Already has MCA_ID + Compliance_Risk — skipping")
        return 0

    # Find insertion point: after Angle_ID
    if "Angle_ID" in headers:
        insert_after = headers.index("Angle_ID") + 1  # 0-based → insert after
        insert_col = insert_after + 1  # openpyxl is 1-based
    elif "Angle" in headers:
        insert_after = headers.index("Angle") + 1
        insert_col = insert_after + 1
    else:
        print(f"  [{ws.title}] WARNING: No Angle_ID or Angle column found — appending at end")
        insert_col = len(headers) + 1

    # Insert two columns at insert_col (MCA_ID first, then Compliance_Risk)
    ws.insert_cols(insert_col, 2)

    # Write headers
    mca_cell = ws.cell(row=1, column=insert_col)
    risk_cell = ws.cell(row=1, column=insert_col + 1)
    mca_cell.value = "MCA_ID"
    risk_cell.value = "Compliance_Risk"
    for cell in (mca_cell, risk_cell):
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = WRAP_ALIGN

    # Refresh headers after insert
    headers = [normalize(cell.value) for cell in ws[1]]
    angle_id_col = headers.index("Angle_ID") if "Angle_ID" in headers else None
    mca_col = headers.index("MCA_ID")
    risk_col = headers.index("Compliance_Risk")

    rows_updated = 0
    for row in ws.iter_rows(min_row=2):
        if angle_id_col is not None:
            angle_id = normalize(row[angle_id_col].value)
        else:
            angle_id = ""

        if angle_id and angle_id in angle_map:
            entry = angle_map[angle_id]
            row[mca_col].value = entry.get("mca_id", "UNASSIGNED")
            row[risk_col].value = entry.get("compliance_risk", "UNASSIGNED")
            rows_updated += 1
        elif any(normalize(cell.value) for cell in row):
            # Populated row without recognized Angle_ID → mark UNASSIGNED
            row[mca_col].value = "UNASSIGNED"
            row[risk_col].value = "UNASSIGNED"

    print(f"  [{ws.title}] Inserted MCA_ID + Compliance_Risk after col {insert_col}; "
          f"{rows_updated} rows populated from angle_map")
    return rows_updated


def main() -> None:
    if not WORKBOOK_PATH.exists():
        print(f"ERROR: Workbook not found: {WORKBOOK_PATH}")
        sys.exit(1)

    if not TAXONOMY_PATH.exists():
        print(f"ERROR: Taxonomy registry not found: {TAXONOMY_PATH}")
        sys.exit(1)

    angle_map = load_angle_map(TAXONOMY_PATH)
    print(f"Loaded angle_to_mca_map: {len(angle_map)} entries")

    wb = load_workbook(WORKBOOK_PATH)

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [{sheet_name}] NOT FOUND in workbook — skipping")
            continue
        patch_sheet(wb[sheet_name], angle_map)

    wb.save(WORKBOOK_PATH)
    print(f"\nWorkbook saved: {WORKBOOK_PATH}")
    print("PATCH COMPLETE")


if __name__ == "__main__":
    main()
