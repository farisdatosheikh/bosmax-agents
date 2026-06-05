from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

from product_copy_router import (
    REGISTRY_PATH,
    WORKBOOK_PATH,
    load_products,
    load_registry,
    load_workbook_index,
    route_product,
)


ROOT = Path(__file__).resolve().parents[1]
REQUIRED_ROUTE_MODES = [
    "REGISTERED_PRODUCT",
    "FAMILY_MATCHED_PRODUCT",
    "ON_THE_FLY_PRODUCT",
    "REVIEW_ONLY_PRODUCT",
]
REQUIRED_SOURCE_HIERARCHY = [
    "product_specific_registry",
    "product_specific_workbook_copy_pack",
    "product_family_workbook_copy_pack",
    "category_mechanic_archetype",
    "on_the_fly_intake",
    "review_only_fallback",
]
REQUIRED_MINIMUM_FIELDS = [
    "product_name",
    "category",
    "target_user",
    "main_problem_solved",
    "main_benefit",
    "product_format",
    "product_size_or_scale",
    "platform",
    "language",
    "compliance_class",
    "visual_reference_status",
]
REQUIRED_RISK_TERMS = [
    "baby",
    "pregnancy",
    "supplement",
    "medical",
    "cure",
    "sembuh",
    "merawat",
    "ubat",
    "penawar",
    "sexual wellness",
    "adult",
    "restricted",
]
REQUIRED_PROMOTION_TRIGGERS = [
    "product used in 2+ sessions",
    "user wants TikTok Shop reusable template",
    "bulk copy packs requested",
    "high-risk product needs permanent quarantine record",
    "recurring product family demand",
]


def fail(message: str) -> None:
    print(f"VALIDATION FAILED: {message}")
    sys.exit(1)


def expect(condition: bool, message: str) -> None:
    if not condition:
        fail(message)


def load_workbook_family_sheets() -> set[str]:
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    return {name for name in wb.sheetnames if name.startswith("FAMILY_")}


def validate_registry_shape(registry: dict) -> None:
    expect(REGISTRY_PATH.exists(), "router registry missing")
    for mode in REQUIRED_ROUTE_MODES:
        expect(mode in registry["route_modes"], f"missing route mode: {mode}")
    print("router registry ok")
    for mode in REQUIRED_ROUTE_MODES:
        print(f"route mode ok: {mode}")

    source_hierarchy = registry.get("source_hierarchy", [])
    for item in REQUIRED_SOURCE_HIERARCHY:
        expect(item in source_hierarchy, f"source hierarchy missing: {item}")

    required_fields = registry["minimum_on_the_fly_intake"]["required_fields"]
    for field in REQUIRED_MINIMUM_FIELDS:
        expect(field in required_fields, f"minimum intake field missing: {field}")

    combined_risk_terms = " ".join(
        registry["risk_detection"]["high_risk_categories"]
        + registry["risk_detection"]["review_only_triggers"]["keywords"]
        + registry["risk_detection"]["forbidden_claim_patterns"]
    ).lower()
    for term in REQUIRED_RISK_TERMS:
        expect(term in combined_risk_terms, f"risk detection missing required term: {term}")

    family_sheets = load_workbook_family_sheets()
    registry_family_sheets = set(registry["family_matching"]["existing_family_sheets"])
    expect("FAMILY_TRAD_REMEDY_OIL" in registry_family_sheets, "FAMILY_TRAD_REMEDY_OIL missing from registry family mapping")
    for family_sheet in registry_family_sheets:
        expect(family_sheet in family_sheets, f"registry family sheet not found in workbook: {family_sheet}")

    triggers = registry["promotion_rules"]["triggers"]
    for trigger in REQUIRED_PROMOTION_TRIGGERS:
        expect(trigger in triggers, f"promotion rule missing: {trigger}")
    print("promotion rules ok")

    authority = registry["authority"]
    expect(authority.get("repo_truth") is True, "repo_truth must be true")
    expect(authority.get("notion_downstream_only") is True, "Notion must remain downstream only")
    expect(authority.get("fastmoss_is_source_signal_not_permission_gate") is True, "FastMoss permission gate posture incorrect")


def validate_routes(registry: dict) -> None:
    products = load_products()
    workbook_index = load_workbook_index()
    registered_cap_burung = route_product(
        {
            "product_id": "CAP_BURUNG_MINYAK",
            "product_name": "Minyak Warisan Cap Burung",
            "category": "traditional remedy oil",
            "compliance_class": "GREEN",
        },
        registry=registry,
        products=products,
        workbook_index=workbook_index,
    )
    expect(registered_cap_burung["route_mode"] == "REGISTERED_PRODUCT", "CAP_BURUNG_MINYAK did not resolve to REGISTERED_PRODUCT")
    print("example route ok: CAP_BURUNG_MINYAK -> REGISTERED_PRODUCT")

    registered_bosmax = route_product(
        {
            "product_id": "BOSMAX_SERUM",
            "product_name": "BOSMAX Serum",
            "category": "men's health oil",
            "compliance_class": "HIGH",
        },
        registry=registry,
        products=products,
        workbook_index=workbook_index,
    )
    expect(registered_bosmax["route_mode"] == "REGISTERED_PRODUCT", "BOSMAX_SERUM did not resolve to REGISTERED_PRODUCT")
    print("example route ok: BOSMAX_SERUM -> REGISTERED_PRODUCT")

    family_match = route_product(
        {
            "product_name": "Minyak Herba Tradisional",
            "category": "traditional remedy oil",
            "compliance_class": "GREEN",
        },
        registry=registry,
        products=products,
        workbook_index=workbook_index,
    )
    expect(family_match["route_mode"] == "FAMILY_MATCHED_PRODUCT", "traditional remedy oil did not resolve to FAMILY_MATCHED_PRODUCT")
    expect(family_match["family_sheet"] == "FAMILY_TRAD_REMEDY_OIL", "traditional remedy oil did not resolve to FAMILY_TRAD_REMEDY_OIL")
    print("example route ok: Minyak Herba Tradisional -> FAMILY_MATCHED_PRODUCT")

    on_the_fly = route_product(
        {
            "product_name": "Portable Blender",
            "category": "kitchen gadget",
            "target_user": "busy home user",
            "main_problem_solved": "quick drink prep",
            "main_benefit": "blend on the go",
            "product_format": "portable blender",
            "product_size_or_scale": "1 unit countertop mini device",
            "platform": "TikTok Shop",
            "language": "BM",
            "compliance_class": "GREEN",
            "visual_reference_status": "USER_PROVIDED",
        },
        registry=registry,
        products=products,
        workbook_index=workbook_index,
    )
    expect(on_the_fly["route_mode"] == "ON_THE_FLY_PRODUCT", "Portable Blender did not resolve to ON_THE_FLY_PRODUCT")
    expect(on_the_fly["ready_for_generation"] is True, "Portable Blender should be generation-ready")
    print("example route ok: Portable Blender -> ON_THE_FLY_PRODUCT")

    review_only = route_product(
        {
            "product_name": "Baby Supplement",
            "category": "baby supplement",
            "target_user": "parents",
            "main_problem_solved": "support immunity",
            "main_benefit": "help baby recover faster",
            "product_format": "powder supplement",
            "product_size_or_scale": "30 sachets",
            "platform": "TikTok Shop",
            "language": "BM",
            "compliance_class": "GREEN",
            "visual_reference_status": "USER_PROVIDED",
        },
        registry=registry,
        products=products,
        workbook_index=workbook_index,
    )
    expect(review_only["route_mode"] == "REVIEW_ONLY_PRODUCT", "Baby Supplement did not resolve to REVIEW_ONLY_PRODUCT")
    print("example route ok: Baby Supplement -> REVIEW_ONLY_PRODUCT")

    high_compliance = route_product(
        {
            "product_name": "Generic Wellness Capsule",
            "category": "wellness product",
            "target_user": "adults",
            "main_problem_solved": "daily support",
            "main_benefit": "feel stronger",
            "product_format": "capsule",
            "product_size_or_scale": "30 capsules",
            "platform": "TikTok Shop",
            "language": "BM",
            "compliance_class": "HIGH",
            "visual_reference_status": "USER_PROVIDED",
        },
        registry=registry,
        products=products,
        workbook_index=workbook_index,
    )
    expect(high_compliance["route_mode"] == "REVIEW_ONLY_PRODUCT", "HIGH compliance class must fail closed into REVIEW_ONLY_PRODUCT")

    incomplete_on_the_fly = route_product(
        {
            "product_name": "Portable Blender",
            "category": "kitchen gadget",
            "compliance_class": "GREEN",
        },
        registry=registry,
        products=products,
        workbook_index=workbook_index,
    )
    expect(incomplete_on_the_fly["route_mode"] == "ON_THE_FLY_PRODUCT", "Incomplete ad-hoc route should still remain ON_THE_FLY_PRODUCT")
    expect(incomplete_on_the_fly["ready_for_generation"] is False, "Incomplete ad-hoc route must not be generation-ready")
    expect("target_user" in incomplete_on_the_fly["minimum_intake_missing"], "Missing intake list is incomplete")
    print("fail-closed rules ok")


def main() -> None:
    registry = load_registry()
    validate_registry_shape(registry)
    validate_routes(registry)
    print("VALIDATION PASSED")


if __name__ == "__main__":
    main()
