from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from resolver_runtime import (
    ResolverError,
    assert_copywriting_pack_runtime,
    normalize_alias,
    resolve_copywriting_id,
)


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "BOSMAX_COPYWRITING_ID_RESOLVER_v1.xlsx"
REGISTRY_PATH = ROOT / "registries" / "copywriting_id_resolver.yaml"

REQUIRED_SHEETS = {
    "COPY_PACK_REGISTRY",
    "COPY_ID_ALIAS_MAP",
    "NOTION_EXPORT_VIEW",
    "VALIDATION_RULES",
    "CHANGELOG",
}
REQUIRED_REGISTRY_HEADERS = {
    "Copywriting_ID",
    "Display_Name",
    "Product_ID",
    "Product_Name",
    "Family_Code",
    "Family_Name",
    "Lane",
    "Silo_Key",
    "Submode_Formula",
    "Angle_ID",
    "Angle",
    "Hook",
    "Pain_or_Friction",
    "USP_1",
    "USP_2",
    "USP_3",
    "CTA",
    "Compliance",
    "Compliance_Risk",
    "Authority_Source",
    "Source_Script_Node",
    "Source_Variant_Hook_Node",
    "Source_Variant_Problem_Node",
    "Source_Variant_Solution_Node",
    "Source_Variant_CTA_Node",
    "Status",
    "Runtime_Allowed",
    "Notes",
}
FORBIDDEN_NOTION_HEADERS = {
    "Authority_Source",
    "Source_Script_Node",
    "Source_Variant_Hook_Node",
    "Source_Variant_Problem_Node",
    "Source_Variant_Solution_Node",
    "Source_Variant_CTA_Node",
    "Compliance_Risk",
}
EXPECTED_NOTION_HEADERS = [
    "Copywriting_ID",
    "Display_Name",
    "Product_Name",
    "Family_Name",
    "Lane",
    "Silo_Key",
    "Submode_Formula",
    "Angle",
    "Hook",
    "USP_1",
    "USP_2",
    "USP_3",
    "CTA",
    "Compliance",
    "Status",
    "Safe_Usage_Notes",
]


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).strip()


def fail(message: str) -> None:
    print(f"FAIL: {message}")
    sys.exit(1)


def expect(condition: bool, message: str) -> None:
    if not condition:
        fail(message)


def load_yaml_registry() -> dict[str, Any]:
    with REGISTRY_PATH.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def validate_workbook_shape() -> None:
    expect(WORKBOOK_PATH.exists(), f"Workbook missing: {WORKBOOK_PATH}")
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    expect(REQUIRED_SHEETS.issubset(set(workbook.sheetnames)), "Resolver workbook missing required sheets")

    registry_sheet = workbook["COPY_PACK_REGISTRY"]
    registry_headers = [normalize(cell.value) for cell in next(registry_sheet.iter_rows(max_row=1))]
    expect(REQUIRED_REGISTRY_HEADERS.issubset(set(registry_headers)), "COPY_PACK_REGISTRY headers drift detected")

    notion_sheet = workbook["NOTION_EXPORT_VIEW"]
    notion_headers = [normalize(cell.value) for cell in next(notion_sheet.iter_rows(max_row=1))]
    expect(notion_headers == EXPECTED_NOTION_HEADERS, "NOTION_EXPORT_VIEW headers drift detected")
    expect(not FORBIDDEN_NOTION_HEADERS.intersection(set(notion_headers)), "NOTION_EXPORT_VIEW exposes forbidden fields")
    print("workbook shape ok")


def validate_registry_shape(registry: dict[str, Any]) -> None:
    expect(REGISTRY_PATH.exists(), f"Registry missing: {REGISTRY_PATH}")
    expect(registry.get("registry_id") == "BOSMAX_COPYWRITING_ID_RESOLVER_V1", "registry_id drift detected")
    expect(bool(registry.get("copy_packs")), "copy_packs cannot be empty")
    expect(bool(registry.get("alias_map")), "alias_map cannot be empty")
    expect(registry.get("samples", {}).get("sample_copywriting_id") == "BOSMAX_SERUM_CP_0001", "sample_copywriting_id drift detected")
    print("registry shape ok")


def validate_aliases(registry: dict[str, Any]) -> None:
    seen: dict[str, str] = {}
    for row in registry.get("alias_map", []):
        normalized = normalize(row.get("normalized_alias")) or normalize_alias(row.get("alias"))
        canonical = normalize(row.get("canonical_id"))
        expect(normalized and canonical, "alias_map rows must include normalized_alias and canonical_id")
        previous = seen.get(normalized)
        expect(previous in (None, canonical), f"alias collision detected for {normalized}: {previous} vs {canonical}")
        seen[normalized] = canonical
    print("alias map ok")


def validate_runtime_records(registry: dict[str, Any]) -> None:
    copy_packs = registry.get("copy_packs", [])
    pack_index = {normalize(pack.get("copywriting_id")): pack for pack in copy_packs}
    expect("BOSMAX_SERUM_CP_0001" in pack_index, "BOSMAX_SERUM_CP_0001 missing")

    for pack in copy_packs:
        assert_copywriting_pack_runtime(
            pack,
            expected_template_lane=pack.get("lane"),
            expected_template_silo=pack.get("silo_key"),
        )
    print("runtime record invariants ok")


def validate_runtime_negative_paths(registry: dict[str, Any]) -> None:
    sample = resolve_copywriting_id(
        "Bosmax_serum_0001",
        expected_template_lane="STEALTH",
        expected_template_silo="male_health_stealth_01",
    )
    expect(sample["copywriting_id"] == "BOSMAX_SERUM_CP_0001", "alias resolution failed for Bosmax_serum_0001")

    try:
        resolve_copywriting_id("BOSMAX_SERUM_CP_9999", expected_template_lane="STEALTH")
    except ResolverError:
        pass
    else:
        fail("Missing Copywriting_ID must fail closed")

    try:
        resolve_copywriting_id("BOSMAX_SERUM_CP_0001", expected_template_lane="DIRECT")
    except ResolverError:
        pass
    else:
        fail("Lane mismatch must fail closed")

    try:
        resolve_copywriting_id(
            "BOSMAX_SERUM_CP_0001",
            expected_template_lane="STEALTH",
            manual_override=True,
            review_status="READY",
        )
    except ResolverError:
        pass
    else:
        fail("Unsafe manual override must fail closed")

    invalid_status_pack = dict(sample)
    invalid_status_pack["status"] = "DRAFT"
    try:
        assert_copywriting_pack_runtime(
            invalid_status_pack,
            expected_template_lane="STEALTH",
            expected_template_silo="male_health_stealth_01",
        )
    except ResolverError:
        pass
    else:
        fail("Invalid status must fail closed")

    missing_provenance_pack = dict(sample)
    missing_provenance_pack["source_script_node"] = ""
    try:
        assert_copywriting_pack_runtime(
            missing_provenance_pack,
            expected_template_lane="STEALTH",
            expected_template_silo="male_health_stealth_01",
        )
    except ResolverError:
        pass
    else:
        fail("Missing BOSMAX Serum provenance nodes must fail closed")

    print("negative fail-closed checks ok")


def main() -> None:
    registry = load_yaml_registry()
    validate_workbook_shape()
    validate_registry_shape(registry)
    validate_aliases(registry)
    validate_runtime_records(registry)
    validate_runtime_negative_paths(registry)
    print("VALIDATION PASSED: copywriting ID resolver")


if __name__ == "__main__":
    main()
