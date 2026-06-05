from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from stealth_copy_authority import (
    ROOT,
    STEALTH_SOURCE_HEADERS,
    build_family_sheet_rows,
    build_product_sheet_rows,
)

LEGACY_WORKBOOK = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_v1.xlsx"
OUTPUT_WORKBOOK = ROOT / "BOSMAX_PRODUCT_COPYWRITING_LIBRARY_FAMILY_v2.xlsx"
FASTMOSS_WORKBOOK = ROOT / "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx"

FLAGSHIP_SHEET_BOSMAX = "PRODUCT_BOSMAX_SERUM"
FLAGSHIP_SHEET_MWCB = "PRODUCT_MW_CAP_BURUNG"
FAMILY_MASTER_SHEET = "DEDUPED_PRODUCT_FAMILIES"
RAW_FASTMOSS_SHEET = "RAW_FASTMOSS_PRODUCTS"
FAMILY_MAPPING_SHEET = "PRODUCT_FAMILY_MAPPING"

WORKING_HEADERS = [
    "Row_ID",
    "Family_Code",
    "Family_Name",
    "Product_ID_Optional",
    "Product_Name_Optional",
    "SKU_Optional",
    "Category",
    "Sub_Category",
    "Product_Type",
    "UOM",
    "Product_Size",
    "Product_Scale",
    "Type_of_Content",
    "Silo_Key",
    "Angle_ID",
    "MCA_ID",
    "Compliance_Risk",
    "Angle",
    "Hook_ID",
    "Hook",
    "Pain_or_Friction",
    "USP_1",
    "USP_2",
    "USP_3",
    "CTA_ID",
    "CTA",
    "Copywriting_Formula",
    "Authority_Source",
    *STEALTH_SOURCE_HEADERS,
    "Fastmoss_Reference",
    "Status",
    "Notes",
]

RAW_FASTMOSS_HEADERS = [
    "Rank",
    "Product_Name",
    "Shop_Name",
    "Category",
    "Sub_Category",
    "Product_Type",
    "Raw_Category",
    "Avg_Price_RM",
    "Commission_Rate",
    "Orders",
    "Order_Growth",
    "Total_Units_Sold",
    "Total_Revenue_RM",
    "Product_Status",
    "Copywriting_Angle",
    "Hook",
    "USP_1",
    "USP_2",
    "USP_3",
    "Body",
    "CTA",
    "Fastmoss_Reference",
]

FAMILY_MAPPING_HEADERS = [
    "Rank",
    "Product_Name",
    "Category",
    "Sub_Category",
    "Product_Type",
    "Assigned_Family_Code",
    "Assigned_Family_Name",
    "Worksheet_Name",
    "Type_of_Content",
    "Silo_Key",
    "Mapping_Confidence",
    "Mapping_Reason",
    "Representative_Function",
    "Fastmoss_Reference",
]

FAMILY_MASTER_HEADERS = [
    "Family_Code",
    "Family_Name",
    "Worksheet_Name",
    "Type_of_Content",
    "Silo_Key_Default",
    "Default_Formula",
    "Category_Family",
    "Commercial_Mechanic",
    "Mapped_Product_Count",
    "Representative_Product_Types",
    "Representative_Products",
    "Flagship_Link",
    "Authority_Source",
    "Notes",
]

INDEX_HEADERS = ["Sheet_Name", "Role", "Fill_Owner", "Description"]

READ_ME_ROWS = [
    (
        "Workbook Role",
        "Operator-facing copywriting library. Raw Fastmoss listings are only source signals; reusable output authority is organized by deduped product family.",
    ),
    (
        "Architecture",
        "Read in order: RAW_FASTMOSS_PRODUCTS -> PRODUCT_FAMILY_MAPPING -> DEDUPED_PRODUCT_FAMILIES -> flagship sheets -> family library sheets.",
    ),
    (
        "Deduping Rule",
        "Collapse brands and listings that sell the same underlying product mechanism into one product family. Example: many men perfume brands -> one men perfume family sheet.",
    ),
    (
        "Authority Boundary",
        "SCRIPT_REGISTRY_UNIFIED.md and SCRIPT_VARIANT_LIBRARY.md remain runtime authority for sensitive dialogue lanes. This workbook is a structured fill surface only.",
    ),
    (
        "Row Contract",
        "One row = one coherent angle-hook-pain/friction arc with one USP triplet, one CTA, and one chosen formula.",
    ),
    (
        "Antigravity Scope",
        "Antigravity fills only Type_of_Content, Silo_Key, Angle, Hook, Pain_or_Friction, USP_1-3, CTA, Copywriting_Formula, Notes, and optional IDs/Status inside product or family library sheets.",
    ),
    (
        "Do Not Touch",
        "Do not rewrite RAW_FASTMOSS_PRODUCTS, PRODUCT_FAMILY_MAPPING, DEDUPED_PRODUCT_FAMILIES, flagship product truth, or family codes without explicit architect approval.",
    ),
    (
        "Strategic Families",
        "Sensitive male external oil, traditional remedy oil, men perfume, women perfume, unisex perfume, hair shampoo, hair oil, tudung bawal, tisu-related lanes, and kambing perap remain first-class routes.",
    ),
]

FORMULA_OPTIONS = ["AIDA", "PAS", "HSO", "HPAS", "SAVAGE_HPAS"]
TYPE_OF_CONTENT_OPTIONS = ["DIRECT", "STEALTH"]
STATUS_OPTIONS = ["SEED_READY", "ANTIGRAVITY_FILL_PENDING", "REVIEW_REQUIRED", "APPROVED", "LOCKED"]
STRATEGIC_FAMILY_CODES = {
    "FAMILY_MALE_EXT_SENSITIVE_OIL",
    "FAMILY_TRADITIONAL_REMEDY_OIL",
    "FAMILY_MEN_PERFUME",
    "FAMILY_WOMEN_PERFUME",
    "FAMILY_UNISEX_PERFUME",
    "FAMILY_HAIR_SHAMPOO",
    "FAMILY_HAIR_OIL",
    "FAMILY_TUDUNG_BAWAL",
    "FAMILY_KAMBING_PERAP",
    "FAMILY_MENS_COLLARED_TSHIRT",
    "FAMILY_TISU_GENERAL",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
RAW_FILL = PatternFill("solid", fgColor="7F6000")
MAPPING_FILL = PatternFill("solid", fgColor="BF9000")
MASTER_FILL = PatternFill("solid", fgColor="5B9BD5")
FLAGSHIP_FILL = PatternFill("solid", fgColor="C00000")
FAMILY_FILL = PatternFill("solid", fgColor="70AD47")
WHITE_FONT = Font(color="FFFFFF", bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


@dataclass(frozen=True)
class FastmossRow:
    rank: int
    product_name: str
    shop_name: str
    category: str
    sub_category: str
    product_type: str
    raw_category: str
    avg_price_rm: Any
    commission_rate: Any
    orders: Any
    order_growth: Any
    total_units_sold: Any
    total_revenue_rm: Any
    product_status: str
    copywriting_angle: str
    hook: str
    usp_1: str
    usp_2: str
    usp_3: str
    body: str
    cta: str
    fastmoss_reference: str

    @property
    def search_text(self) -> str:
        return normalize_spaces(
            " ".join(
                [
                    self.product_name,
                    self.category,
                    self.sub_category,
                    self.product_type,
                    self.raw_category,
                ]
            )
        ).lower()


@dataclass(frozen=True)
class FamilyDefinition:
    code: str
    name: str
    sheet_name: str
    type_of_content: str
    silo_key: str
    default_formula: str
    category_family: str
    commercial_mechanic: str
    flagship_link: str
    authority_source: str
    notes: str


@dataclass(frozen=True)
class FamilyAssignment:
    family_code: str
    reason: str
    confidence: str


def normalize_spaces(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def derive_pain_from_body(body: str, fallback_hook: str) -> str:
    body_text = normalize_spaces(body)
    hook_text = normalize_spaces(fallback_hook)
    if body_text:
        match = re.search(r"Ramai orang ada masalah (.+?)\.", body_text, flags=re.IGNORECASE)
        if match:
            candidate = normalize_spaces(match.group(1))
            if candidate:
                return candidate[0].upper() + candidate[1:] + "."

        if "mula dengan masalah ini:" in body_text:
            after = body_text.split("mula dengan masalah ini:", 1)[1]
            candidate = normalize_spaces(after.split(". ", 1)[0])
            if candidate:
                return f"Bila {candidate[0].lower() + candidate[1:]}, orang terus cari pilihan yang lebih praktikal."

        first_sentence = normalize_spaces(body_text.split(". ", 1)[0])
        if first_sentence:
            return first_sentence if first_sentence.endswith(".") else f"{first_sentence}."

    if hook_text:
        lowered = hook_text.rstrip("?!")
        return f"Bila {lowered[0].lower() + lowered[1:]}, rutin terus terasa terganggu."

    return ""


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "_", value.upper()).strip("_")
    return re.sub(r"_+", "_", cleaned) or "GENERIC"


def sheet_name_for_family(code: str) -> str:
    if len(code) <= 31:
        return code
    return code[:31]


def contains_any(text: str, keywords: list[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def product_yaml(path: Path) -> dict[str, Any]:
    return yaml.safe_load(path.read_text(encoding="utf-8"))


def style_header(ws: Worksheet, fill: PatternFill) -> None:
    for cell in ws[1]:
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.alignment = WRAP_ALIGNMENT


def set_standard_layout(ws: Worksheet) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        max_length = 0
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(normalize_spaces(cell.value)))
            cell.alignment = WRAP_ALIGNMENT
        ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)


def create_sheet_with_headers(wb: Workbook, title: str, headers: list[str], fill: PatternFill) -> Worksheet:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    style_header(ws, fill)
    return ws


def write_rows(ws: Worksheet, rows: list[list[Any]]) -> None:
    for row in rows:
        ws.append(row)
    set_standard_layout(ws)


def add_validations(ws: Worksheet) -> None:
    header_positions = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    if "Type_of_Content" not in header_positions:
        return

    type_validation = DataValidation(type="list", formula1='"DIRECT,STEALTH"', allow_blank=True)
    formula_validation = DataValidation(type="list", formula1='"AIDA,PAS,HSO,HPAS,SAVAGE_HPAS"', allow_blank=True)
    status_validation = DataValidation(type="list", formula1='"SEED_READY,ANTIGRAVITY_FILL_PENDING,REVIEW_REQUIRED,APPROVED,LOCKED"', allow_blank=True)
    ws.add_data_validation(type_validation)
    ws.add_data_validation(formula_validation)
    ws.add_data_validation(status_validation)

    type_col = get_column_letter(header_positions["Type_of_Content"])
    formula_col = get_column_letter(header_positions["Copywriting_Formula"])
    status_col = get_column_letter(header_positions["Status"])
    type_validation.add(f"{type_col}2:{type_col}5000")
    formula_validation.add(f"{formula_col}2:{formula_col}5000")
    status_validation.add(f"{status_col}2:{status_col}5000")


def load_fastmoss_copy_map() -> list[FastmossRow]:
    wb = load_workbook(FASTMOSS_WORKBOOK, read_only=True, data_only=True)
    ws = wb["Copywriting_Product_Map"]
    rows: list[FastmossRow] = []
    for raw in ws.iter_rows(min_row=5, values_only=True):
        if not raw or raw[0] is None:
            continue
        rank = int(raw[0])
        rows.append(
            FastmossRow(
                rank=rank,
                product_name=normalize_spaces(raw[1]),
                shop_name=normalize_spaces(raw[2]),
                category=normalize_spaces(raw[3]),
                sub_category=normalize_spaces(raw[4]),
                product_type=normalize_spaces(raw[5]),
                raw_category=normalize_spaces(raw[6]),
                avg_price_rm=raw[7],
                commission_rate=raw[8],
                orders=raw[9],
                order_growth=raw[10],
                total_units_sold=raw[11],
                total_revenue_rm=raw[12],
                product_status=normalize_spaces(raw[13]),
                copywriting_angle=normalize_spaces(raw[14]),
                hook=normalize_spaces(raw[15]),
                usp_1=normalize_spaces(raw[16]),
                usp_2=normalize_spaces(raw[17]),
                usp_3=normalize_spaces(raw[18]),
                body=normalize_spaces(raw[19]),
                cta=normalize_spaces(raw[20]),
                fastmoss_reference=f"FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map#rank={rank}",
            )
        )
    wb.close()
    return rows


def build_family_catalog() -> dict[str, FamilyDefinition]:
    families = [
        FamilyDefinition(
            code="FAMILY_MALE_EXT_SENSITIVE_OIL",
            name="Male Sensitive External Oil",
            sheet_name="FAMILY_MALE_EXT_OIL",
            type_of_content="STEALTH",
            silo_key="male_health_stealth_01",
            default_formula="SAVAGE_HPAS",
            category_family="Health & Wellness",
            commercial_mechanic="Same stealth male external-oil function sold under many brands and pack names.",
            flagship_link=FLAGSHIP_SHEET_BOSMAX,
            authority_source="products/BOSMAX_SERUM.yaml + products/MAVERIX_MAXOIL.yaml + script registry",
            notes="Collapse Bosmax Serum, Maverix Maxoil, Big Boss-type lanes, and future sensitive external male oil brands into one copy family.",
        ),
        FamilyDefinition(
            code="FAMILY_TRADITIONAL_REMEDY_OIL",
            name="Traditional Remedy Oil",
            sheet_name="FAMILY_TRAD_REMEDY_OIL",
            type_of_content="DIRECT",
            silo_key="traditional_remedy_direct",
            default_formula="PAS",
            category_family="Health & Wellness",
            commercial_mechanic="Household-relief traditional oil sold through pain-relief, comfort, and heritage positioning.",
            flagship_link=FLAGSHIP_SHEET_MWCB,
            authority_source="products/CAP_BURUNG_MINYAK.yaml + products/JUNGLE_GIRL_MINYAK.yaml",
            notes="Collapse direct minyak urut, minyak angin, and traditional remedy oils that sell through practical relief rather than stealth intimacy.",
        ),
        FamilyDefinition(
            code="FAMILY_MEN_PERFUME",
            name="Perfume Lelaki",
            sheet_name="FAMILY_MEN_PERFUME",
            type_of_content="DIRECT",
            silo_key="men_perfume_direct",
            default_formula="AIDA",
            category_family="Beauty & Personal Care",
            commercial_mechanic="Many men perfume brands, one repeatable confidence-and-attraction perfume copy engine.",
            flagship_link="",
            authority_source="FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map",
            notes="Strategic family retained even when current Fastmoss slice has few or zero men-specific rows.",
        ),
        FamilyDefinition(
            code="FAMILY_WOMEN_PERFUME",
            name="Perfume Wanita",
            sheet_name="FAMILY_WOMEN_PERFUME",
            type_of_content="DIRECT",
            silo_key="women_perfume_direct",
            default_formula="AIDA",
            category_family="Beauty & Personal Care",
            commercial_mechanic="Many women perfume brands, one repeatable attraction-confidence-freshness perfume family.",
            flagship_link="",
            authority_source="FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map",
            notes="Use for fragrance lanes aimed at women and female self-confidence positioning.",
        ),
        FamilyDefinition(
            code="FAMILY_UNISEX_PERFUME",
            name="Perfume Unisex",
            sheet_name="FAMILY_UNISEX_PERFUME",
            type_of_content="DIRECT",
            silo_key="unisex_perfume_direct",
            default_formula="AIDA",
            category_family="Beauty & Personal Care",
            commercial_mechanic="Fragrance sold on freshness, cleanliness, and easy daily confidence without strict gender coding.",
            flagship_link="",
            authority_source="FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map",
            notes="Acts as the neutral perfume family when listing copy is not strongly gendered.",
        ),
        FamilyDefinition(
            code="FAMILY_HAIR_SHAMPOO",
            name="Shampoo Rambut",
            sheet_name="FAMILY_HAIR_SHAMPOO",
            type_of_content="DIRECT",
            silo_key="hair_shampoo_direct",
            default_formula="PAS",
            category_family="Beauty & Personal Care",
            commercial_mechanic="Scalp-cleaning and smooth-hair promise sold across shampoo brands.",
            flagship_link="",
            authority_source="FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map",
            notes="Keep distinct from hair oil and treatment mask families.",
        ),
        FamilyDefinition(
            code="FAMILY_HAIR_OIL",
            name="Minyak Rambut",
            sheet_name="FAMILY_HAIR_OIL",
            type_of_content="DIRECT",
            silo_key="hair_oil_direct",
            default_formula="PAS",
            category_family="Beauty & Personal Care",
            commercial_mechanic="Hair-growth, anti-hair-fall, or nourishing-oil promise sold as topical hair oil.",
            flagship_link="",
            authority_source="FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map",
            notes="Includes castor oil and similar topical rambut-growth lanes.",
        ),
        FamilyDefinition(
            code="FAMILY_TUDUNG_BAWAL",
            name="Tudung Bawal",
            sheet_name="FAMILY_TUDUNG_BAWAL",
            type_of_content="DIRECT",
            silo_key="tudung_bawal_direct",
            default_formula="AIDA",
            category_family="Muslim Fashion",
            commercial_mechanic="Many bawal designs, one reusable shaping-easy-styling copy engine.",
            flagship_link="",
            authority_source="FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map",
            notes="Collapse printed and plain bawal listings into one family if the wearing logic is the same.",
        ),
        FamilyDefinition(
            code="FAMILY_KAMBING_PERAP",
            name="Kambing Perap",
            sheet_name="FAMILY_KAMBING_PERAP",
            type_of_content="DIRECT",
            silo_key="food_kambing_perap_direct",
            default_formula="AIDA",
            category_family="Food & Beverages",
            commercial_mechanic="Pre-marinated meat sold through taste, convenience, and celebration suitability.",
            flagship_link="",
            authority_source="FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map",
            notes="Strategic family retained because the user explicitly wants kambing perap as a reusable product family.",
        ),
        FamilyDefinition(
            code="FAMILY_MENS_COLLARED_TSHIRT",
            name="T-Shirt Berkolar Lelaki",
            sheet_name="FAMILY_MENS_COLLARED_T",
            type_of_content="DIRECT",
            silo_key="mens_collared_tshirt_direct",
            default_formula="AIDA",
            category_family="Menswear & Underwear",
            commercial_mechanic="Men tops sold through neatness, smart-casual look, and easy daily wear.",
            flagship_link="",
            authority_source="Future Fastmoss mapping or manual family registration",
            notes="Placeholder strategic family for future polo/collared-men-top imports, even if current top-300 slice has zero rows.",
        ),
        FamilyDefinition(
            code="FAMILY_TISU_GENERAL",
            name="Tisu",
            sheet_name="FAMILY_TISU_GENERAL",
            type_of_content="DIRECT",
            silo_key="tisu_general_direct",
            default_formula="AIDA",
            category_family="Home Supplies",
            commercial_mechanic="Disposable wipe/tissue convenience lane retained as a future umbrella family.",
            flagship_link="",
            authority_source="Future Fastmoss mapping or manual family registration",
            notes="Use only when a generic tisu family is genuinely needed. Specific tissue lanes remain more useful when function differs.",
        ),
    ]
    return {family.code: family for family in families}


def fallback_family_definition(row: FastmossRow) -> FamilyDefinition:
    type_slug = slugify(row.product_type)[:18]
    code = f"FAMILY_{type_slug}"
    name = normalize_spaces(row.product_type) or "Generic Product Family"
    family_name = name.title() if name.islower() else name
    return FamilyDefinition(
        code=code,
        name=family_name,
        sheet_name=sheet_name_for_family(code),
        type_of_content="DIRECT",
        silo_key=f"generic_{slugify(name).lower()}",
        default_formula="AIDA",
        category_family=row.category,
        commercial_mechanic="Auto-generated fallback family from Fastmoss type; refine only if a broader commercial family becomes obvious.",
        flagship_link="",
        authority_source="FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map",
        notes="Fallback family generated from a low-frequency type. Manual review can collapse this further later.",
    )


def classify_family(row: FastmossRow) -> FamilyAssignment:
    text = row.search_text
    type_text = row.product_type.lower()
    sub_text = row.sub_category.lower()
    category_text = row.category.lower()
    perfume_context = "perfume" in type_text or sub_text == "perfume"
    hair_context = sub_text == "haircare & styling" or category_text == "beauty & personal care" and contains_any(text, ["hair", "rambut"])

    if contains_any(
        text,
        [
            "maverix",
            "bosmax",
            "maxoil",
            "big boss",
            "men's care",
            "male health",
            "vitaliti lelaki",
            "bahagian sensitif",
            "intimate care oil",
        ],
    ) or (sub_text == "men's care" and "bath & body care" in type_text):
        return FamilyAssignment("FAMILY_MALE_EXT_SENSITIVE_OIL", "Sensitive male external-oil lane detected from brand/product-lane keywords.", "HIGH")

    if contains_any(
        text,
        [
            "traditional remedy",
            "minyak angin",
            "minyak urut",
            "massage oil",
            "habbatus",
            "liniment",
        ],
    ) or type_text == "body & massage oil":
        return FamilyAssignment("FAMILY_TRADITIONAL_REMEDY_OIL", "Traditional remedy / massage-oil function detected.", "HIGH")

    if "car fragrance" in type_text or contains_any(text, ["pewangi kereta", "air freshener", "car perfume"]):
        return FamilyAssignment("FAMILY_CAR_FRAGRANCE", "Vehicle fragrance lane detected.", "HIGH")

    if perfume_context and ("women's perfume" in type_text or contains_any(text, ["women perfume", "for women", "ladies perfume", "perfume wanita", "perfume perempuan", "wangian wanita"])):
        return FamilyAssignment("FAMILY_WOMEN_PERFUME", "Women-specific perfume signals detected.", "HIGH")

    if perfume_context and "unisex perfume" in type_text:
        return FamilyAssignment("FAMILY_UNISEX_PERFUME", "Unisex perfume type detected.", "HIGH")

    if perfume_context:
        if contains_any(text, ["for men", "lelaki", "men perfume", "masculine scent", "for him"]):
            return FamilyAssignment("FAMILY_MEN_PERFUME", "Men-specific perfume wording detected.", "HIGH")
        return FamilyAssignment("FAMILY_UNISEX_PERFUME", "Generic perfume without strong gender marker routed to unisex perfume family.", "MEDIUM")

    if hair_context and contains_any(text, ["hair oil", "minyak rambut", "castor oil"]):
        return FamilyAssignment("FAMILY_HAIR_OIL", "Hair oil function detected.", "HIGH")

    if hair_context and ("hair dye" in type_text or contains_any(text, ["pewarna rambut", "hair colour", "hair color", "tutup uban"])):
        return FamilyAssignment("FAMILY_HAIR_DYE", "Hair dye / color lane detected.", "HIGH")

    if hair_context and ("hair treatments/scalp treatments" in type_text or contains_any(text, ["keratin", "topeng rawatan", "serum rambut", "rebonding", "rawatan rambut"])):
        return FamilyAssignment("FAMILY_HAIR_TREATMENT", "Hair treatment or repair-mask lane detected.", "HIGH")

    if hair_context and ("shampoo & conditioner" in type_text or contains_any(text, ["syampu", "shampoo", "hair conditioner"])):
        return FamilyAssignment("FAMILY_HAIR_SHAMPOO", "Hair shampoo / conditioner lane detected.", "HIGH")

    if "beauty supplement" in type_text:
        return FamilyAssignment("FAMILY_BEAUTY_SUPPLEMENT", "Beauty supplement lane detected.", "HIGH")

    if sub_text == "food supplements" or "wellness supplements" in type_text:
        return FamilyAssignment("FAMILY_WELLNESS_SUPPLEMENT", "General wellness supplement lane detected.", "HIGH")

    if type_text == "body wash & soap":
        return FamilyAssignment("FAMILY_BODY_SOAP", "Body soap / body wash lane detected.", "HIGH")

    if type_text == "facial cleansers":
        return FamilyAssignment("FAMILY_SKINCARE_CLEANSER", "Facial cleanser lane detected.", "HIGH")

    if "facial sunscreen" in type_text:
        return FamilyAssignment("FAMILY_SKINCARE_SUNSCREEN", "Sunscreen lane detected.", "HIGH")

    if type_text == "serums & essences":
        return FamilyAssignment("FAMILY_SKINCARE_SERUM", "Skincare serum lane detected.", "HIGH")

    if type_text == "moisturizers & mists":
        return FamilyAssignment("FAMILY_SKINCARE_MIST", "Hydration mist / moisturizer lane detected.", "HIGH")

    if type_text == "face masks":
        return FamilyAssignment("FAMILY_FACE_MASK", "Face mask lane detected.", "HIGH")

    if type_text == "lip treatments":
        return FamilyAssignment("FAMILY_LIP_CARE", "Lip treatment lane detected.", "HIGH")

    if type_text == "eye treatments":
        return FamilyAssignment("FAMILY_EYE_CARE", "Eye care lane detected.", "HIGH")

    if type_text == "diapers":
        return FamilyAssignment("FAMILY_BABY_DIAPER", "Baby diaper lane detected.", "HIGH")

    if type_text == "adult diapers":
        return FamilyAssignment("FAMILY_ADULT_DIAPER", "Adult diaper lane detected.", "HIGH")

    if type_text == "wipes & holders":
        return FamilyAssignment("FAMILY_BABY_WIPES", "Baby wipes lane detected.", "HIGH")

    if type_text == "baby bottles & accessories":
        return FamilyAssignment("FAMILY_BABY_FEEDING_BOTTLE", "Baby feeding bottle lane detected.", "HIGH")

    if type_text == "baby skincare":
        return FamilyAssignment("FAMILY_BABY_SKINCARE", "Baby skincare lane detected.", "HIGH")

    if type_text == "cleaning cloths" and contains_any(text, ["tisu", "tissue", "paper towel"]):
        return FamilyAssignment("FAMILY_KITCHEN_TISSUE", "Kitchen tissue / disposable wipe lane detected.", "HIGH")

    if type_text == "makeup remover":
        return FamilyAssignment("FAMILY_MAKEUP_REMOVER_WIPES", "Makeup-remover tissue lane detected.", "HIGH")

    if "hijabs" in sub_text or "tudung" in text:
        if "square hijabs" in type_text or "bawal" in text:
            return FamilyAssignment("FAMILY_TUDUNG_BAWAL", "Bawal/square hijab lane detected.", "HIGH")
        if "instant hijab" in type_text:
            return FamilyAssignment("FAMILY_TUDUNG_INSTANT", "Instant hijab lane detected.", "HIGH")
        if "underscarves" in type_text or contains_any(text, ["anak tudung", "snow cap", "inner hijab"]):
            return FamilyAssignment("FAMILY_INNER_TUDUNG", "Inner tudung lane detected.", "HIGH")

    if type_text == "socks":
        return FamilyAssignment("FAMILY_SOCKS", "Socks lane detected.", "HIGH")

    if type_text in {"underwear", "sports underwear"}:
        return FamilyAssignment("FAMILY_MENS_UNDERWEAR", "Men underwear lane detected.", "HIGH")

    if type_text == "bras":
        return FamilyAssignment("FAMILY_WOMENS_BRA", "Women bra lane detected.", "HIGH")

    if type_text == "t" and category_text == "menswear & underwear":
        return FamilyAssignment("FAMILY_MENS_TSHIRT", "Men T-shirt lane detected.", "HIGH")

    if type_text in {"blouses & shirts", "shirts & blouses", "islamic tracksuits"} or contains_any(text, ["jersey muslimah", "jersi muslimah", "microfiber", "athleisure"]):
        return FamilyAssignment("FAMILY_WOMENS_JERSEY_TOP", "Women jersey / Muslimah top lane detected.", "MEDIUM")

    if type_text == "trousers":
        if category_text == "menswear & underwear":
            return FamilyAssignment("FAMILY_MENS_LONG_PANTS", "Men long-pants lane detected.", "HIGH")
        return FamilyAssignment("FAMILY_WOMENS_LONG_PANTS", "Women long-pants lane detected.", "HIGH")

    if type_text == "shorts":
        if category_text == "menswear & underwear":
            return FamilyAssignment("FAMILY_MENS_SHORTS", "Men shorts lane detected.", "HIGH")
        return FamilyAssignment("FAMILY_WOMENS_SHORTS", "Women shorts lane detected.", "HIGH")

    if type_text == "bedding sets":
        return FamilyAssignment("FAMILY_BEDDING_SET", "Bedding-set lane detected.", "HIGH")

    if type_text == "sheets & pillowcases":
        return FamilyAssignment("FAMILY_BEDSHEET_SET", "Bedsheet / pillowcase lane detected.", "HIGH")

    if type_text == "curtains":
        return FamilyAssignment("FAMILY_CURTAINS", "Curtain lane detected.", "HIGH")

    if type_text == "carpets, mats & rugs":
        return FamilyAssignment("FAMILY_CARPETS", "Carpet/mat lane detected.", "HIGH")

    if type_text in {"storage boxes & bins", "storage holders & racks"}:
        return FamilyAssignment("FAMILY_STORAGE_BOX", "Storage/organization lane detected.", "HIGH")

    if type_text in {"household cleaners", "trash bags"}:
        return FamilyAssignment("FAMILY_HOUSEHOLD_CLEANER", "Household cleaning lane detected.", "MEDIUM")

    if type_text == "pest & weed control":
        return FamilyAssignment("FAMILY_PEST_CONTROL", "Pest-control lane detected.", "HIGH")

    if type_text == "popcorn":
        return FamilyAssignment("FAMILY_POPCORN", "Popcorn snack lane detected.", "HIGH")

    if type_text == "cooking sauces":
        return FamilyAssignment("FAMILY_COOKING_SAUCE", "Cooking sauce / sambal lane detected.", "HIGH")

    if type_text in {"canned, jarred & packaged foods", "instant noodles", "instant food"}:
        return FamilyAssignment("FAMILY_INSTANT_FOOD", "Instant or ready-to-eat food lane detected.", "MEDIUM")

    if type_text == "frozen food" or contains_any(text, ["kambing perap"]):
        return FamilyAssignment("FAMILY_KAMBING_PERAP", "Kambing perap / marinated-frozen-food lane detected.", "HIGH")

    if category_text == "beauty & personal care":
        if sub_text == "makeup":
            return FamilyAssignment("FAMILY_MAKEUP_COSMETICS", "General makeup lane fallback.", "MEDIUM")
        if sub_text == "bath & body care":
            return FamilyAssignment("FAMILY_BODY_CARE_FRESHNESS", "General bath/body-care lane fallback.", "MEDIUM")
        if sub_text == "haircare & styling":
            return FamilyAssignment("FAMILY_HAIR_CARE_GENERAL", "General hair-care lane fallback.", "MEDIUM")
        if sub_text == "feminine care":
            return FamilyAssignment("FAMILY_FEMININE_CARE", "Feminine-care lane fallback.", "MEDIUM")
        if sub_text == "hand, foot & nail care":
            return FamilyAssignment("FAMILY_NAIL_CARE", "Nail-care lane fallback.", "MEDIUM")
        if sub_text == "nasal & oral care":
            return FamilyAssignment("FAMILY_ORAL_CARE", "Oral-care lane fallback.", "MEDIUM")

    if category_text == "health":
        if sub_text == "medical supplies":
            return FamilyAssignment("FAMILY_HOME_HEALTH_TEST", "Home health-test lane fallback.", "MEDIUM")

    if category_text == "food & beverages":
        if sub_text == "snacks":
            return FamilyAssignment("FAMILY_SNACKS", "General snack lane fallback.", "MEDIUM")
        if sub_text == "staples & cooking essentials":
            return FamilyAssignment("FAMILY_COOKING_ESSENTIALS", "Cooking-essentials lane fallback.", "MEDIUM")
        if sub_text == "drinks":
            return FamilyAssignment("FAMILY_DRINKS", "Drink lane fallback.", "MEDIUM")
        if sub_text == "fresh & frozen food":
            return FamilyAssignment("FAMILY_FROZEN_FOOD", "Frozen-food lane fallback.", "MEDIUM")

    if category_text == "phones & electronics":
        if "smart watches" in type_text:
            return FamilyAssignment("FAMILY_SMARTWATCH", "Smartwatch lane fallback.", "MEDIUM")
        return FamilyAssignment("FAMILY_GADGET_ACCESSORY", "General gadget-accessory lane fallback.", "MEDIUM")

    if category_text == "automotive & motorcycle":
        return FamilyAssignment("FAMILY_AUTO_ACCESSORY", "General automotive-accessory lane fallback.", "MEDIUM")

    if category_text == "household appliances":
        return FamilyAssignment("FAMILY_HOME_APPLIANCE", "General home-appliance lane fallback.", "MEDIUM")

    if category_text == "pet supplies":
        return FamilyAssignment("FAMILY_PET_FOOD_ACCESSORY", "Pet-supplies lane fallback.", "MEDIUM")

    if category_text == "books, magazines & audio":
        return FamilyAssignment("FAMILY_BOOKS_REFERENCE", "Books/reference lane fallback.", "MEDIUM")

    if category_text == "computers & office equipment":
        return FamilyAssignment("FAMILY_STATIONERY_GIFT", "Stationery/gift lane fallback.", "MEDIUM")

    if category_text == "kitchenware":
        return FamilyAssignment("FAMILY_KITCHENWARE", "Kitchenware lane fallback.", "MEDIUM")

    if category_text == "home improvement":
        return FamilyAssignment("FAMILY_HOME_IMPROVEMENT", "Home-improvement lane fallback.", "MEDIUM")

    if category_text == "home supplies":
        if sub_text == "home decor" or sub_text == "festive & party supplies":
            return FamilyAssignment("FAMILY_HOME_DECOR_GIFT", "Home-decor or party-gift lane fallback.", "MEDIUM")
        if sub_text == "bathroom supplies":
            return FamilyAssignment("FAMILY_BATHROOM_SUPPLY", "Bathroom-supply lane fallback.", "MEDIUM")
        return FamilyAssignment("FAMILY_HOME_UTILITY", "General home-utility lane fallback.", "MEDIUM")

    if category_text == "textiles & soft furnishings":
        return FamilyAssignment("FAMILY_HOME_TEXTILE_MISC", "General home-textile lane fallback.", "MEDIUM")

    if category_text == "muslim fashion":
        if sub_text == "women's islamic clothing" or sub_text == "islamic sportswear":
            return FamilyAssignment("FAMILY_MUSLIMAH_CLOTHING", "Muslimah clothing lane fallback.", "MEDIUM")
        if sub_text == "islamic accessories":
            return FamilyAssignment("FAMILY_HIJAB_ACCESSORY", "Hijab accessory lane fallback.", "MEDIUM")
        return FamilyAssignment("FAMILY_MUSLIM_FASHION_MISC", "General Muslim fashion lane fallback.", "MEDIUM")

    if category_text == "womenswear & underwear":
        if "sleepwear" in sub_text:
            return FamilyAssignment("FAMILY_WOMENS_SLEEPWEAR", "Women sleepwear lane fallback.", "MEDIUM")
        if "tops" in sub_text:
            return FamilyAssignment("FAMILY_WOMENS_TOPS", "Women tops lane fallback.", "MEDIUM")
        if "bottoms" in sub_text:
            return FamilyAssignment("FAMILY_WOMENS_BOTTOMS", "Women bottoms lane fallback.", "MEDIUM")
        return FamilyAssignment("FAMILY_WOMENS_FASHION_MISC", "General women-fashion lane fallback.", "MEDIUM")

    if category_text == "menswear & underwear":
        if "tops" in sub_text:
            return FamilyAssignment("FAMILY_MENS_TOPS", "Men tops lane fallback.", "MEDIUM")
        if "bottoms" in sub_text:
            return FamilyAssignment("FAMILY_MENS_BOTTOMS", "Men bottoms lane fallback.", "MEDIUM")
        return FamilyAssignment("FAMILY_MENS_FASHION_MISC", "General men-fashion lane fallback.", "MEDIUM")

    if category_text == "sports & outdoor":
        return FamilyAssignment("FAMILY_SPORTSWEAR_MISC", "Sports/outdoor apparel lane fallback.", "MEDIUM")

    if category_text == "baby & maternity":
        return FamilyAssignment("FAMILY_BABY_ESSENTIALS", "General baby-essentials lane fallback.", "MEDIUM")

    if category_text == "fashion accessories":
        return FamilyAssignment("FAMILY_FASHION_ACCESSORY", "Fashion-accessory lane fallback.", "MEDIUM")

    if category_text == "shoes":
        return FamilyAssignment("FAMILY_FOOTWEAR", "Footwear lane fallback.", "MEDIUM")

    if category_text == "toys & hobbies":
        return FamilyAssignment("FAMILY_TOYS_CRAFTS", "Toys/crafts lane fallback.", "MEDIUM")

    return FamilyAssignment("", "No strategic family rule matched; fallback family required.", "LOW")


def ensure_dynamic_family(families: dict[str, FamilyDefinition], code: str, row: FastmossRow) -> None:
    if code in families:
        return
    families[code] = fallback_family_definition(row)


def dynamic_catalog_extension(families: dict[str, FamilyDefinition]) -> None:
    extra_families = [
        FamilyDefinition("FAMILY_CAR_FRAGRANCE", "Pewangi Kereta", "FAMILY_CAR_FRAGRANCE", "DIRECT", "car_fragrance_direct", "AIDA", "Automotive & Motorcycle", "Fragrance sold through cabin freshness and long-lasting scent benefit.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Keep separate from body perfume because usage context differs."),
        FamilyDefinition("FAMILY_HAIR_TREATMENT", "Rawatan Rambut", "FAMILY_HAIR_TREATMENT", "DIRECT", "hair_treatment_direct", "PAS", "Beauty & Personal Care", "Hair repair and smoothness promise sold through treatment/mask logic.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Separate from shampoo and hair oil."),
        FamilyDefinition("FAMILY_HAIR_DYE", "Pewarna Rambut", "FAMILY_HAIR_DYE", "DIRECT", "hair_dye_direct", "PAS", "Beauty & Personal Care", "Cover-grey and color-refresh hair lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Useful for tutup uban messaging."),
        FamilyDefinition("FAMILY_BEAUTY_SUPPLEMENT", "Supplement Kecantikan", "FAMILY_BEAUTY_SUPP", "DIRECT", "beauty_supplement_direct", "AIDA", "Health", "Beauty-from-within supplement lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Keep separate from general wellness supplement."),
        FamilyDefinition("FAMILY_WELLNESS_SUPPLEMENT", "Supplement Kesihatan", "FAMILY_WELLNESS_SUPP", "DIRECT", "wellness_supplement_direct", "PAS", "Health", "General wellness supplement lane with health-maintenance positioning.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Broad family for chewables, detox, and wellness support."),
        FamilyDefinition("FAMILY_BODY_SOAP", "Sabun Badan", "FAMILY_BODY_SOAP", "DIRECT", "body_soap_direct", "PAS", "Beauty & Personal Care", "Soap/body-wash lane sold through cleanliness and visible-skin comfort.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Keep separate from facial cleanser."),
        FamilyDefinition("FAMILY_SKINCARE_CLEANSER", "Pencuci Muka", "FAMILY_SKINCARE_CLEANSER", "DIRECT", "skincare_cleanser_direct", "PAS", "Beauty & Personal Care", "Facial cleanser lane sold through brightening, oil-control, or acne-clearing benefit.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Skincare cleanser family."),
        FamilyDefinition("FAMILY_SKINCARE_SUNSCREEN", "Sunscreen", "FAMILY_SKIN_SUNSCREEN", "DIRECT", "skincare_sunscreen_direct", "AIDA", "Beauty & Personal Care", "Daily UV-defense lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Separate from serum and cleanser."),
        FamilyDefinition("FAMILY_SKINCARE_SERUM", "Serum Wajah", "FAMILY_SKINCARE_SERUM", "DIRECT", "skincare_serum_direct", "AIDA", "Beauty & Personal Care", "Face serum lane sold through brightening and glow improvement.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Serum-specific face-care family."),
        FamilyDefinition("FAMILY_SKINCARE_MIST", "Mist Penjagaan Kulit", "FAMILY_SKINCARE_MIST", "DIRECT", "skincare_mist_direct", "AIDA", "Beauty & Personal Care", "Hydration mist / serum spray lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Useful for instant-hydration and glow copy."),
        FamilyDefinition("FAMILY_FACE_MASK", "Mask Wajah", "FAMILY_FACE_MASK", "DIRECT", "face_mask_direct", "AIDA", "Beauty & Personal Care", "Mask lane sold through hydration or soothing logic.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Face-mask family."),
        FamilyDefinition("FAMILY_LIP_CARE", "Penjagaan Bibir", "FAMILY_LIP_CARE", "DIRECT", "lip_care_direct", "AIDA", "Beauty & Personal Care", "Lip treatment lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Lip-serum and lip-care family."),
        FamilyDefinition("FAMILY_EYE_CARE", "Penjagaan Mata", "FAMILY_EYE_CARE", "DIRECT", "eye_care_direct", "AIDA", "Beauty & Personal Care", "Eye-care hydration line-smoothing lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Eye-balm and eye-treatment family."),
        FamilyDefinition("FAMILY_BABY_DIAPER", "Lampin Bayi", "FAMILY_BABY_DIAPER", "DIRECT", "baby_diaper_direct", "PAS", "Baby & Maternity", "Baby diaper lane sold through comfort, leak protection, and value-pack logic.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Collapse multiple diaper brands into one family."),
        FamilyDefinition("FAMILY_ADULT_DIAPER", "Lampin Dewasa", "FAMILY_ADULT_DIAPER", "DIRECT", "adult_diaper_direct", "PAS", "Beauty & Personal Care", "Adult diaper lane sold through absorbency and comfort.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Low-frequency but distinct function family."),
        FamilyDefinition("FAMILY_BABY_WIPES", "Tisu Basah Bayi", "FAMILY_BABY_WIPES", "DIRECT", "baby_wipes_direct", "AIDA", "Baby & Maternity", "Baby wipes lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Keep separate from kitchen tissue."),
        FamilyDefinition("FAMILY_BABY_FEEDING_BOTTLE", "Botol Bayi", "FAMILY_BABY_BOTTLE", "DIRECT", "baby_bottle_direct", "AIDA", "Baby & Maternity", "Bottle-feeding lane sold through anti-colic and newborn suitability.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Baby bottle family."),
        FamilyDefinition("FAMILY_BABY_SKINCARE", "Penjagaan Kulit Bayi", "FAMILY_BABY_SKINCARE", "DIRECT", "baby_skincare_direct", "AIDA", "Baby & Maternity", "Baby skincare family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Baby-care topical family."),
        FamilyDefinition("FAMILY_KITCHEN_TISSUE", "Tisu Dapur", "FAMILY_KITCHEN_TISSUE", "DIRECT", "kitchen_tissue_direct", "AIDA", "Home Supplies", "Disposable cleaning paper/tissue lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Useful for wipe, absorb, and convenience copy."),
        FamilyDefinition("FAMILY_MAKEUP_REMOVER_WIPES", "Tisu Makeup Remover", "FAMILY_MAKEUP_REMOVER", "DIRECT", "makeup_remover_wipes_direct", "AIDA", "Beauty & Personal Care", "Convenient makeup-remover tissue lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Separate from generic tisu."),
        FamilyDefinition("FAMILY_TUDUNG_INSTANT", "Tudung Instant", "FAMILY_TUDUNG_INSTANT", "DIRECT", "tudung_instant_direct", "AIDA", "Muslim Fashion", "Quick-wear hijab lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Different wearing logic from bawal."),
        FamilyDefinition("FAMILY_INNER_TUDUNG", "Inner Tudung", "FAMILY_INNER_TUDUNG", "DIRECT", "inner_tudung_direct", "AIDA", "Muslim Fashion", "Inner scarf / anak tudung lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Support layer family for hijab-wearers."),
        FamilyDefinition("FAMILY_MENS_TSHIRT", "T-Shirt Lelaki", "FAMILY_MENS_TSHIRT", "DIRECT", "mens_tshirt_direct", "AIDA", "Menswear & Underwear", "Men tops sold through comfort and easy daily wear.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Different from future collared-shirt family."),
        FamilyDefinition("FAMILY_WOMENS_JERSEY_TOP", "Top/Jersi Wanita", "FAMILY_WOMEN_JERSEY", "DIRECT", "women_jersey_top_direct", "AIDA", "Womenswear & Underwear", "Women jersey and Muslimah activewear top lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Covers microfiber Muslimah jersey and similar easy-wear tops."),
        FamilyDefinition("FAMILY_MENS_UNDERWEAR", "Seluar Dalam Lelaki", "FAMILY_MENS_UNDERWEAR", "DIRECT", "mens_underwear_direct", "AIDA", "Menswear & Underwear", "Men underwear comfort lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Comfort-fit men undergarment family."),
        FamilyDefinition("FAMILY_WOMENS_BRA", "Bra Wanita", "FAMILY_WOMENS_BRA", "DIRECT", "womens_bra_direct", "AIDA", "Womenswear & Underwear", "Women bra comfort/support lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Support, coverage, comfort family."),
        FamilyDefinition("FAMILY_SOCKS", "Stokin", "FAMILY_SOCKS", "DIRECT", "socks_direct", "AIDA", "Fashion", "Socks lane sold through comfort and value-pack simplicity.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Shared across men/women listings because use-case is the same."),
        FamilyDefinition("FAMILY_MENS_LONG_PANTS", "Seluar Panjang Lelaki", "FAMILY_MENS_PANTS", "DIRECT", "mens_pants_direct", "AIDA", "Menswear & Underwear", "Men long-pants lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Track/jogger/casual long-pants family."),
        FamilyDefinition("FAMILY_WOMENS_LONG_PANTS", "Seluar Panjang Wanita", "FAMILY_WOMENS_PANTS", "DIRECT", "womens_pants_direct", "AIDA", "Womenswear & Underwear", "Women long-pants lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Bootcut/wide-leg everyday pants family."),
        FamilyDefinition("FAMILY_MENS_SHORTS", "Seluar Pendek Lelaki", "FAMILY_MENS_SHORTS", "DIRECT", "mens_shorts_direct", "AIDA", "Menswear & Underwear", "Men shorts lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Casual/sports shorts family."),
        FamilyDefinition("FAMILY_WOMENS_SHORTS", "Seluar Pendek Wanita", "FAMILY_WOMEN_SHORTS", "DIRECT", "womens_shorts_direct", "AIDA", "Womenswear & Underwear", "Women shorts lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Casual shorts family."),
        FamilyDefinition("FAMILY_BEDDING_SET", "Set Bedding", "FAMILY_BEDDING_SET", "DIRECT", "bedding_set_direct", "AIDA", "Textiles & Soft Furnishings", "Complete bedding-set lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Covers full bedsheet/comforter set offers."),
        FamilyDefinition("FAMILY_BEDSHEET_SET", "Cadar & Sarung Bantal", "FAMILY_BEDSHEET_SET", "DIRECT", "bedsheet_set_direct", "AIDA", "Textiles & Soft Furnishings", "Bedsheet and pillowcase lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Separate from full bedding set."),
        FamilyDefinition("FAMILY_CURTAINS", "Langsir", "FAMILY_CURTAINS", "DIRECT", "curtain_direct", "AIDA", "Textiles & Soft Furnishings", "Curtain lane sold through room transformation and blackout comfort.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Home textile transformation family."),
        FamilyDefinition("FAMILY_CARPETS", "Karpet", "FAMILY_CARPETS", "DIRECT", "carpet_direct", "AIDA", "Textiles & Soft Furnishings", "Carpet lane sold through comfort and room-upgrade effect.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Home-comfort floor textile family."),
        FamilyDefinition("FAMILY_STORAGE_BOX", "Storage Organizer", "FAMILY_STORAGE_BOX", "DIRECT", "storage_organizer_direct", "AIDA", "Home Supplies", "Home organization lane sold through space-saving clarity.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Storage box and organizer family."),
        FamilyDefinition("FAMILY_HOUSEHOLD_CLEANER", "Pembersih Rumah", "FAMILY_HOUSE_CLEANER", "DIRECT", "household_cleaner_direct", "PAS", "Home Supplies", "Home-cleaning lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Cleaner, detergent, and trash-bag style practical lane."),
        FamilyDefinition("FAMILY_PEST_CONTROL", "Kawalan Serangga", "FAMILY_PEST_CONTROL", "DIRECT", "pest_control_direct", "PAS", "Home Supplies", "Pest-control lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Lipas/semut/nyamuk control family."),
        FamilyDefinition("FAMILY_POPCORN", "Popcorn", "FAMILY_POPCORN", "DIRECT", "popcorn_direct", "AIDA", "Food & Beverages", "Snack lane sold through sedap, rangup, and gifting suitability.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Popcorn family."),
        FamilyDefinition("FAMILY_COOKING_SAUCE", "Sos & Sambal", "FAMILY_COOKING_SAUCE", "DIRECT", "cooking_sauce_direct", "AIDA", "Food & Beverages", "Flavor booster lane sold through sedap, mudah, and menu-upgrade logic.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Cooking sauce and sambal family."),
        FamilyDefinition("FAMILY_INSTANT_FOOD", "Makanan Segera", "FAMILY_INSTANT_FOOD", "DIRECT", "instant_food_direct", "AIDA", "Food & Beverages", "Ready-to-eat or quick-serve food lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Instant convenience food family."),
        FamilyDefinition("FAMILY_MAKEUP_COSMETICS", "Makeup", "FAMILY_MAKEUP", "DIRECT", "makeup_direct", "AIDA", "Beauty & Personal Care", "Color-cosmetic lane sold through beauty transformation and finish.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback umbrella for makeup types not split further."),
        FamilyDefinition("FAMILY_BODY_CARE_FRESHNESS", "Penjagaan Badan", "FAMILY_BODY_CARE", "DIRECT", "body_care_direct", "AIDA", "Beauty & Personal Care", "Bath/body-care freshness lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Umbrella for bath/body products outside specific soap or remedy families."),
        FamilyDefinition("FAMILY_HAIR_CARE_GENERAL", "Penjagaan Rambut", "FAMILY_HAIR_CARE", "DIRECT", "hair_care_direct", "PAS", "Beauty & Personal Care", "General hair-care umbrella for anything not classified as shampoo, oil, dye, or treatment.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "General hair-care fallback."),
        FamilyDefinition("FAMILY_FEMININE_CARE", "Penjagaan Intim Wanita", "FAMILY_FEMININE_CARE", "STEALTH", "female_care_stealth", "PAS", "Beauty & Personal Care", "Women intimate-care lane handled carefully and euphemistically.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Sensitive but non-explicit feminine-care fallback."),
        FamilyDefinition("FAMILY_NAIL_CARE", "Penjagaan Kuku", "FAMILY_NAIL_CARE", "DIRECT", "nail_care_direct", "AIDA", "Beauty & Personal Care", "Nail polish and nail styling lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Nail-care family."),
        FamilyDefinition("FAMILY_ORAL_CARE", "Penjagaan Mulut", "FAMILY_ORAL_CARE", "DIRECT", "oral_care_direct", "PAS", "Beauty & Personal Care", "Tooth and oral-care lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Oral-care family."),
        FamilyDefinition("FAMILY_HOME_HEALTH_TEST", "Alat Ujian Kesihatan", "FAMILY_HEALTH_TEST", "DIRECT", "health_test_direct", "AIDA", "Health", "Home health monitoring and self-test lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Health-test family."),
        FamilyDefinition("FAMILY_SNACKS", "Snek", "FAMILY_SNACKS", "DIRECT", "snack_direct", "AIDA", "Food & Beverages", "General snack lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Umbrella snack family."),
        FamilyDefinition("FAMILY_COOKING_ESSENTIALS", "Bahan Masakan", "FAMILY_COOK_ESSENTIALS", "DIRECT", "cooking_essentials_direct", "AIDA", "Food & Beverages", "Cooking essentials and pantry-lane umbrella.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Pantry/cooking fallback family."),
        FamilyDefinition("FAMILY_DRINKS", "Minuman", "FAMILY_DRINKS", "DIRECT", "drinks_direct", "AIDA", "Food & Beverages", "Drink lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "General drink family."),
        FamilyDefinition("FAMILY_FROZEN_FOOD", "Makanan Sejuk Beku", "FAMILY_FROZEN_FOOD", "DIRECT", "frozen_food_direct", "AIDA", "Food & Beverages", "Frozen-food lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Frozen-food umbrella outside kambing perap."),
        FamilyDefinition("FAMILY_GADGET_ACCESSORY", "Aksesori Gadget", "FAMILY_GADGET_ACCESSORY", "DIRECT", "gadget_accessory_direct", "AIDA", "Phones & Electronics", "Gadget-accessory umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Accessories, holders, and cable-related products."),
        FamilyDefinition("FAMILY_SMARTWATCH", "Smartwatch", "FAMILY_SMARTWATCH", "DIRECT", "smartwatch_direct", "AIDA", "Phones & Electronics", "Smartwatch lane.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Wearable-device family."),
        FamilyDefinition("FAMILY_AUTO_ACCESSORY", "Aksesori Automotif", "FAMILY_AUTO_ACCESSORY", "DIRECT", "auto_accessory_direct", "AIDA", "Automotive & Motorcycle", "Automotive-accessory umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Car-care and mount/accessory family."),
        FamilyDefinition("FAMILY_HOME_APPLIANCE", "Peralatan Rumah", "FAMILY_HOME_APPLIANCE", "DIRECT", "home_appliance_direct", "AIDA", "Household Appliances", "Home-appliance umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Portable fans, vacuums, and similar tools."),
        FamilyDefinition("FAMILY_PET_FOOD_ACCESSORY", "Produk Haiwan Peliharaan", "FAMILY_PET_PRODUCTS", "DIRECT", "pet_products_direct", "AIDA", "Pet Supplies", "Pet-food and pet-accessory umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Pet supplies family."),
        FamilyDefinition("FAMILY_BOOKS_REFERENCE", "Buku & Rujukan", "FAMILY_BOOKS_REFERENCE", "DIRECT", "books_reference_direct", "AIDA", "Books, Magazines & Audio", "Book and reference-content family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Book/reference family."),
        FamilyDefinition("FAMILY_STATIONERY_GIFT", "Stationery & Hadiah", "FAMILY_STATIONERY", "DIRECT", "stationery_gift_direct", "AIDA", "Computers & Office Equipment", "Stationery and small-gift umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Stationery/gifting family."),
        FamilyDefinition("FAMILY_KITCHENWARE", "Kitchenware", "FAMILY_KITCHENWARE", "DIRECT", "kitchenware_direct", "AIDA", "Kitchenware", "Kitchenware umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Kitchen utensil and cookware family."),
        FamilyDefinition("FAMILY_HOME_IMPROVEMENT", "Home Improvement", "FAMILY_HOME_IMPROVE", "DIRECT", "home_improvement_direct", "AIDA", "Home Improvement", "Home-improvement umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Lighting, wall, and garden helper family."),
        FamilyDefinition("FAMILY_HOME_DECOR_GIFT", "Dekorasi & Hadiah Rumah", "FAMILY_HOME_DECOR", "DIRECT", "home_decor_gift_direct", "AIDA", "Home Supplies", "Home decor and giftable house-item umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Home decor/party gift family."),
        FamilyDefinition("FAMILY_BATHROOM_SUPPLY", "Keperluan Bilik Air", "FAMILY_BATHROOM", "DIRECT", "bathroom_supply_direct", "AIDA", "Home Supplies", "Bathroom-supply umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Bathroom utility family."),
        FamilyDefinition("FAMILY_HOME_UTILITY", "Keperluan Rumah", "FAMILY_HOME_UTILITY", "DIRECT", "home_utility_direct", "AIDA", "Home Supplies", "General home-utility umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback for home supplies."),
        FamilyDefinition("FAMILY_HOME_TEXTILE_MISC", "Tekstil Rumah", "FAMILY_HOME_TEXTILE", "DIRECT", "home_textile_direct", "AIDA", "Textiles & Soft Furnishings", "General home-textile umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback for remaining home textiles."),
        FamilyDefinition("FAMILY_MUSLIMAH_CLOTHING", "Pakaian Muslimah", "FAMILY_MUSLIMAH_WEAR", "DIRECT", "muslimah_clothing_direct", "AIDA", "Muslim Fashion", "Muslimah clothing umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "For Islamic clothing not routed to bawal or inner-tudung."),
        FamilyDefinition("FAMILY_HIJAB_ACCESSORY", "Aksesori Hijab", "FAMILY_HIJAB_ACCESSORY", "DIRECT", "hijab_accessory_direct", "AIDA", "Muslim Fashion", "Hijab accessory family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Pins and brooches family."),
        FamilyDefinition("FAMILY_MUSLIM_FASHION_MISC", "Fesyen Muslim", "FAMILY_MUSLIM_FASHION", "DIRECT", "muslim_fashion_direct", "AIDA", "Muslim Fashion", "General Muslim fashion umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback for Muslim fashion."),
        FamilyDefinition("FAMILY_WOMENS_SLEEPWEAR", "Baju Tidur Wanita", "FAMILY_WOMEN_SLEEP", "DIRECT", "women_sleepwear_direct", "AIDA", "Womenswear & Underwear", "Women sleepwear family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Sleepwear family."),
        FamilyDefinition("FAMILY_WOMENS_TOPS", "Top Wanita", "FAMILY_WOMEN_TOPS", "DIRECT", "women_tops_direct", "AIDA", "Womenswear & Underwear", "Women tops umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback women-top family."),
        FamilyDefinition("FAMILY_WOMENS_BOTTOMS", "Bottom Wanita", "FAMILY_WOMEN_BOTTOMS", "DIRECT", "women_bottoms_direct", "AIDA", "Womenswear & Underwear", "Women bottoms umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback women-bottom family."),
        FamilyDefinition("FAMILY_WOMENS_FASHION_MISC", "Fesyen Wanita", "FAMILY_WOMEN_FASHION", "DIRECT", "women_fashion_direct", "AIDA", "Womenswear & Underwear", "General women-fashion umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback women-fashion family."),
        FamilyDefinition("FAMILY_MENS_TOPS", "Top Lelaki", "FAMILY_MEN_TOPS", "DIRECT", "men_tops_direct", "AIDA", "Menswear & Underwear", "Men tops umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback men-top family."),
        FamilyDefinition("FAMILY_MENS_BOTTOMS", "Bottom Lelaki", "FAMILY_MEN_BOTTOMS", "DIRECT", "men_bottoms_direct", "AIDA", "Menswear & Underwear", "Men bottoms umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback men-bottom family."),
        FamilyDefinition("FAMILY_MENS_FASHION_MISC", "Fesyen Lelaki", "FAMILY_MEN_FASHION", "DIRECT", "men_fashion_direct", "AIDA", "Menswear & Underwear", "General men-fashion umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback men-fashion family."),
        FamilyDefinition("FAMILY_SPORTSWEAR_MISC", "Sportswear", "FAMILY_SPORTSWEAR", "DIRECT", "sportswear_direct", "AIDA", "Sports & Outdoor", "General sportswear umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Sports/outdoor wear family."),
        FamilyDefinition("FAMILY_BABY_ESSENTIALS", "Keperluan Bayi", "FAMILY_BABY_ESSENTIALS", "DIRECT", "baby_essentials_direct", "AIDA", "Baby & Maternity", "General baby-essentials umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Fallback baby family."),
        FamilyDefinition("FAMILY_FASHION_ACCESSORY", "Aksesori Fesyen", "FAMILY_FASHION_ACCESSORY", "DIRECT", "fashion_accessory_direct", "AIDA", "Fashion Accessories", "Fashion-accessory umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "General fashion-accessory family."),
        FamilyDefinition("FAMILY_FOOTWEAR", "Footwear", "FAMILY_FOOTWEAR", "DIRECT", "footwear_direct", "AIDA", "Shoes", "Footwear family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "Footwear family."),
        FamilyDefinition("FAMILY_TOYS_CRAFTS", "Toys & Crafts", "FAMILY_TOYS_CRAFTS", "DIRECT", "toys_crafts_direct", "AIDA", "Toys & Hobbies", "Toys/crafts umbrella family.", "", "FASTMOSS_COMBINED_10_FILES_WORKBOOK.xlsx::Copywriting_Product_Map", "General toys/crafts family."),
    ]
    for family in extra_families:
        families.setdefault(family.code, family)


def load_existing_flagship_rows(sheet_name: str, family_code: str, family_name: str) -> list[list[Any]]:
    source_workbooks = [OUTPUT_WORKBOOK, LEGACY_WORKBOOK]
    workbook_path = next((path for path in source_workbooks if path.exists()), None)
    if workbook_path is None:
        return []

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]
    header_row = [normalize_spaces(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header_index = {header: idx for idx, header in enumerate(header_row)}
    rows: list[list[Any]] = []

    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or all(value is None for value in raw):
            continue

        if "Family_Code" in header_index:
            source = {header: raw[idx] if idx < len(raw) else "" for header, idx in header_index.items()}
            row_values = [source.get(header, "") for header in WORKING_HEADERS]
            row_values[1] = family_code
            row_values[2] = family_name
            rows.append(row_values)
            continue

        source = {header: raw[idx] if idx < len(raw) else "" for header, idx in header_index.items()}
        rows.append(
            [
                source.get("Row_ID", ""),
                family_code,
                family_name,
                source.get("Product_ID_Optional", ""),
                source.get("Product_Name_Optional", ""),
                source.get("SKU_Optional", ""),
                source.get("Category", ""),
                source.get("Sub_Category", ""),
                source.get("Product_Type", ""),
                source.get("UOM", ""),
                source.get("Product_Size", ""),
                source.get("Product_Scale", ""),
                source.get("Type_of_Content", ""),
                source.get("Silo_Key", ""),
                source.get("Angle_ID", ""),
                source.get("Angle", ""),
                source.get("Hook_ID", ""),
                source.get("Hook", ""),
                source.get("Pain_or_Friction", ""),
                source.get("USP_1", ""),
                source.get("USP_2", ""),
                source.get("USP_3", ""),
                source.get("CTA_ID", ""),
                source.get("CTA", ""),
                source.get("Copywriting_Formula", ""),
                source.get("Authority_Source", ""),
                source.get("Source_Script_Node", ""),
                source.get("Source_Variant_Hook_Node", ""),
                source.get("Source_Variant_Problem_Node", ""),
                source.get("Source_Variant_Solution_Node", ""),
                source.get("Source_Variant_CTA_Node", ""),
                source.get("Fastmoss_Reference", ""),
                source.get("Status", ""),
                source.get("Notes", ""),
            ]
        )

    wb.close()
    return rows


def default_flagship_bosmax_rows() -> list[list[Any]]:
    return [[row.get(header, "") for header in WORKING_HEADERS] for row in build_product_sheet_rows()]


def default_flagship_mwcb_rows() -> list[list[Any]]:
    product = product_yaml(ROOT / "products" / "CAP_BURUNG_MINYAK.yaml")
    variant = product["variants"][0]
    authority = "products/CAP_BURUNG_MINYAK.yaml"
    starter_rows = [
        (
            "ANG_001",
            "Minyak rumah yang orang capai bila badan rasa tak sedap dan perlukan sapuan cepat.",
            "HOOK_001",
            "Benda macam ni biasanya tak perlu intro panjang sebab sekali tengok terus orang tahu kegunaan dia dalam rumah.",
            "Masalah dia selalu datang tanpa amaran: badan terasa tak sedap, kepala berat, atau angin naik masa orang dah sibuk dengan rutin lain.",
            "Profil minyak tradisional buat positioning melegakan dan menyegarkan lebih mudah dipercayai.",
            "Saiz 30ml WG40 sesuai untuk simpan dalam beg, kereta, atau laci rumah.",
            "Heritage 1958 kuat untuk copy yang tekan unsur turun-temurun dan kepercayaan keluarga.",
            "CTA_001",
            "Simpan satu botol di rumah atau dalam kereta sebelum waktu perlu baru tercari-cari.",
            "PAS",
            "Seeded household relief row.",
        ),
        (
            "ANG_002",
            "Produk warisan yang senang dijual bila orang nampak botol kecil tapi gunaannya banyak.",
            "HOOK_002",
            "Kalau rumah Melayu ada satu botol minyak serbaguna, memang jenis macam ni yang selalu jadi standby.",
            "Friction biasa berlaku bila keperluan kecil datang berulang kali, tetapi orang malas simpan terlalu banyak botol untuk masalah yang berbeza-beza.",
            "Visual botol klasik terus bantu angle petua lama yang masih dipakai sampai sekarang.",
            "Copy boleh masuk lane keluarga, travel, dan urut ringan tanpa jadi terlalu sempit.",
            "Cap merah dan label warisan bantu rasa trusted dan familiar pada pembeli.",
            "CTA_002",
            "Klik beg kuning dan standby satu botol untuk kegunaan rumah.",
            "AIDA",
            "Seeded heritage direct row.",
        ),
    ]

    rows: list[list[Any]] = []
    for idx, row in enumerate(starter_rows, start=1):
        rows.append(
            [
                f"{FLAGSHIP_SHEET_MWCB}_R{idx:03d}",
                "FAMILY_TRADITIONAL_REMEDY_OIL",
                "Traditional Remedy Oil",
                product["product_id"],
                product["product_name"],
                variant["variant_id"],
                product["category"],
                product["sub_category"],
                product["product_type"],
                "Bottle",
                variant["variant_name"],
                variant["scale_anchor_descriptor"],
                "DIRECT",
                "traditional_remedy_direct",
                f"{FLAGSHIP_SHEET_MWCB}_{row[0]}",
                row[1],
                f"{FLAGSHIP_SHEET_MWCB}_{row[2]}",
                row[3],
                row[4],
                row[5],
                row[6],
                f"{FLAGSHIP_SHEET_MWCB}_{row[8]}",
                row[9],
                row[10],
                authority,
                "",
                "",
                "",
                "",
                "",
                product.get("fastmoss_product_name", "") or "No confirmed direct Fastmoss listing locked yet.",
                "SEED_READY",
                row[11],
            ]
        )
    return rows


def build_raw_rows(rows: list[FastmossRow]) -> list[list[Any]]:
    return [
        [
            row.rank,
            row.product_name,
            row.shop_name,
            row.category,
            row.sub_category,
            row.product_type,
            row.raw_category,
            row.avg_price_rm,
            row.commission_rate,
            row.orders,
            row.order_growth,
            row.total_units_sold,
            row.total_revenue_rm,
            row.product_status,
            row.copywriting_angle,
            row.hook,
            row.usp_1,
            row.usp_2,
            row.usp_3,
            row.body,
            row.cta,
            row.fastmoss_reference,
        ]
        for row in rows
    ]


def build_mapping_and_grouped_rows(
    rows: list[FastmossRow],
    families: dict[str, FamilyDefinition],
) -> tuple[list[list[Any]], dict[str, list[FastmossRow]]]:
    mapping_rows: list[list[Any]] = []
    grouped: dict[str, list[FastmossRow]] = defaultdict(list)

    for row in rows:
        assignment = classify_family(row)
        family_code = assignment.family_code
        if not family_code:
            fallback = fallback_family_definition(row)
            families.setdefault(fallback.code, fallback)
            family_code = fallback.code
            assignment = FamilyAssignment(family_code, f"{assignment.reason} Fallback generated from product type '{row.product_type}'.", "LOW")

        ensure_dynamic_family(families, family_code, row)
        family = families[family_code]
        grouped[family_code].append(row)
        mapping_rows.append(
            [
                row.rank,
                row.product_name,
                row.category,
                row.sub_category,
                row.product_type,
                family.code,
                family.name,
                family.sheet_name,
                family.type_of_content,
                family.silo_key,
                assignment.confidence,
                assignment.reason,
                family.commercial_mechanic,
                row.fastmoss_reference,
            ]
        )

    return mapping_rows, grouped


def build_family_master_rows(
    families: dict[str, FamilyDefinition],
    grouped: dict[str, list[FastmossRow]],
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for code in sorted(families):
        family = families[code]
        source_rows = sorted(grouped.get(code, []), key=lambda item: item.rank)
        representative_products = " | ".join(item.product_name for item in source_rows[:4])
        product_types = " | ".join(sorted({item.product_type for item in source_rows[:4]}))
        rows.append(
            [
                family.code,
                family.name,
                family.sheet_name,
                family.type_of_content,
                family.silo_key,
                family.default_formula,
                family.category_family,
                family.commercial_mechanic,
                len(source_rows),
                product_types,
                representative_products,
                family.flagship_link,
                family.authority_source,
                family.notes,
            ]
        )
    return rows


def seed_family_library_rows(
    families: dict[str, FamilyDefinition],
    grouped: dict[str, list[FastmossRow]],
) -> dict[str, list[list[Any]]]:
    seeded: dict[str, list[list[Any]]] = {}
    for code, family in families.items():
        if code == "FAMILY_MALE_EXT_SENSITIVE_OIL":
            seeded[code] = [[row.get(header, "") for header in WORKING_HEADERS] for row in build_family_sheet_rows()]
            continue

        rows: list[list[Any]] = []
        source_rows = sorted(grouped.get(code, []), key=lambda item: item.rank)
        for idx, row in enumerate(source_rows[:3], start=1):
            rows.append(
                [
                    f"{family.sheet_name}_R{idx:03d}",
                    family.code,
                    family.name,
                    "",
                    row.product_name,
                    "",
                    row.category,
                    row.sub_category,
                    row.product_type,
                    "VARIES",
                    "VARIES_BY_BRAND",
                    "SEE_ACTUAL_PRODUCT_TRUTH",
                    family.type_of_content,
                    family.silo_key,
                    f"{family.code}_ANG_{idx:03d}",
                    row.copywriting_angle,
                    f"{family.code}_HOOK_{idx:03d}",
                    row.hook,
                    derive_pain_from_body(row.body, row.hook),
                    row.usp_1,
                    row.usp_2,
                    row.usp_3,
                    f"{family.code}_CTA_{idx:03d}",
                    row.cta,
                    family.default_formula,
                    family.authority_source,
                    "",
                    "",
                    "",
                    "",
                    "",
                    row.fastmoss_reference,
                    "SEED_READY",
                    "Representative Fastmoss row extracted for this deduped family. Generalize beyond the brand name when expanding.",
                ]
            )
        seeded[code] = rows
    return seeded


def write_index_sheet(wb: Workbook, family_rows_for_sheets: list[list[Any]]) -> None:
    ws = wb.active
    ws.title = "INDEX"
    ws.append(INDEX_HEADERS)
    style_header(ws, HEADER_FILL)
    rows = [
        ["INDEX", "Governance", "System", "Navigation sheet for workbook structure and ownership."],
        ["README_OR_RULES", "Governance", "System", "Operating rules, fill boundaries, and family-deduping logic."],
        [RAW_FASTMOSS_SHEET, "Source Intake", "System", "All 300 raw Fastmoss listings retained exactly as intake evidence."],
        [FAMILY_MAPPING_SHEET, "Routing", "System", "Per-listing assignment from raw Fastmoss product into one deduped product family."],
        [FAMILY_MASTER_SHEET, "Master Taxonomy", "System", "One row per deduped product family after collapsing same-function brands/listings."],
        [FLAGSHIP_SHEET_BOSMAX, "Flagship Product", "Human + Antigravity", "Sensitive male external-oil flagship sheet tied to Bosmax Serum product truth."],
        [FLAGSHIP_SHEET_MWCB, "Flagship Product", "Human + Antigravity", "Traditional remedy flagship sheet tied to Minyak Warisan Cap Burung product truth."],
    ]
    for row in family_rows_for_sheets:
        sheet_name = row[2]
        rows.append([sheet_name, "Family Library", "Antigravity", f"Reusable copy library for {row[1]} family."])
    write_rows(ws, rows)


def write_readme_sheet(wb: Workbook) -> None:
    ws = create_sheet_with_headers(wb, "README_OR_RULES", ["Rule", "Detail"], HEADER_FILL)
    for rule, detail in READ_ME_ROWS:
        ws.append([rule, detail])
    set_standard_layout(ws)


def tint_second_row(ws: Worksheet) -> None:
    if ws.max_row >= 2:
        for cell in ws[2]:
            cell.fill = SUB_HEADER_FILL


def build_workbook() -> None:
    families = build_family_catalog()
    dynamic_catalog_extension(families)

    fastmoss_rows = load_fastmoss_copy_map()
    raw_rows = build_raw_rows(fastmoss_rows)
    mapping_rows, grouped = build_mapping_and_grouped_rows(fastmoss_rows, families)
    family_master_rows = build_family_master_rows(families, grouped)
    family_library_rows = seed_family_library_rows(families, grouped)

    selected_family_rows = [
        row for row in family_master_rows
        if (isinstance(row[8], int) and row[8] >= 3) or row[0] in STRATEGIC_FAMILY_CODES
    ]

    wb = Workbook()
    write_index_sheet(wb, selected_family_rows)
    write_readme_sheet(wb)

    raw_ws = create_sheet_with_headers(wb, RAW_FASTMOSS_SHEET, RAW_FASTMOSS_HEADERS, RAW_FILL)
    write_rows(raw_ws, raw_rows)

    mapping_ws = create_sheet_with_headers(wb, FAMILY_MAPPING_SHEET, FAMILY_MAPPING_HEADERS, MAPPING_FILL)
    write_rows(mapping_ws, mapping_rows)

    master_ws = create_sheet_with_headers(wb, FAMILY_MASTER_SHEET, FAMILY_MASTER_HEADERS, MASTER_FILL)
    write_rows(master_ws, family_master_rows)

    bosmax_rows = load_existing_flagship_rows(
        FLAGSHIP_SHEET_BOSMAX,
        "FAMILY_MALE_EXT_SENSITIVE_OIL",
        "Male Sensitive External Oil",
    ) or default_flagship_bosmax_rows()
    bosmax_ws = create_sheet_with_headers(wb, FLAGSHIP_SHEET_BOSMAX, WORKING_HEADERS, FLAGSHIP_FILL)
    write_rows(bosmax_ws, bosmax_rows)
    add_validations(bosmax_ws)
    tint_second_row(bosmax_ws)

    mwcb_rows = load_existing_flagship_rows(
        FLAGSHIP_SHEET_MWCB,
        "FAMILY_TRADITIONAL_REMEDY_OIL",
        "Traditional Remedy Oil",
    ) or default_flagship_mwcb_rows()
    mwcb_ws = create_sheet_with_headers(wb, FLAGSHIP_SHEET_MWCB, WORKING_HEADERS, FLAGSHIP_FILL)
    write_rows(mwcb_ws, mwcb_rows)
    add_validations(mwcb_ws)
    tint_second_row(mwcb_ws)

    family_codes_in_order = [row[0] for row in selected_family_rows]
    for code in family_codes_in_order:
        family = families[code]
        ws = create_sheet_with_headers(wb, family.sheet_name, WORKING_HEADERS, FAMILY_FILL)
        write_rows(ws, family_library_rows.get(code, []))
        add_validations(ws)
        tint_second_row(ws)

    if OUTPUT_WORKBOOK.exists():
        OUTPUT_WORKBOOK.unlink()
    wb.save(OUTPUT_WORKBOOK)


if __name__ == "__main__":
    build_workbook()
    print(f"Wrote workbook: {OUTPUT_WORKBOOK}")
